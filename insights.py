# -*- coding: utf-8 -*-
"""insights.py

Минимально-стабильная реализация генерации листа "инсайты".

Цель: убрать проблему с битой кодировкой и всегда писать нормальный русский текст.
Блоки 1–5 сейчас выводятся в компактном виде (при необходимости расширим).
Блок 6 делает расчеты по ТЗ и заполняет таблицу + сводку флагов формулами.

Вход: bytes Excel (.xlsx/.xlsm) с листами W, M, Wt, Mt и опционально общ/кред.
Выход: bytes Excel.
"""

from __future__ import annotations

import io
import re
import zipfile
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# -------- styles --------
C_RED = "CC0000"
C_GREEN = "006600"
C_ORANGE = "CC8800"
C_GRAY = "666666"

FILL_HDR = PatternFill("solid", fgColor="D9E1F2")
FILL_PROB = PatternFill("solid", fgColor="FCE4EC")
FILL_RISK = PatternFill("solid", fgColor="FFF2CC")

A_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
A_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=False)
A_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)

NUM = "#,##0"
PCT = "0.0%"
MULT = '0.0"x"'

F_DEF = Font(name="Calibri", size=11)
F_HDR = Font(name="Calibri", size=11, bold=True)
F_TITLE = Font(name="Calibri", size=14, bold=True)
F_GRAY_ITALIC = Font(name="Calibri", size=9, italic=True, color=C_GRAY)
F_RED_BOLD = Font(name="Calibri", size=11, bold=True, color=C_RED)
F_GREEN_BOLD = Font(name="Calibri", size=11, bold=True, color=C_GREEN)
F_ORANGE_BOLD = Font(name="Calibri", size=11, bold=True, color=C_ORANGE)


def _detect_keep_vba(file_bytes: bytes) -> bool:
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
            return "xl/vbaProject.bin" in set(zf.namelist())
    except Exception:
        return False


def _load_wb(file_bytes: bytes):
    return load_workbook(io.BytesIO(file_bytes), data_only=False, keep_vba=_detect_keep_vba(file_bytes))


def _save_wb(wb) -> bytes:
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _as_text(v) -> str:
    return "" if v is None else str(v)


def _to_float(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return 0.0
    s = s.replace("\u00A0", "").replace("\u202F", "").replace(" ", "")
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


def _to_intish(v) -> Optional[int]:
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, int):
        return int(v)
    if isinstance(v, float):
        return int(v)
    s = str(v).strip()
    if not s or s.startswith("="):
        return None
    m = re.search(r"-?\d+", s)
    if not m:
        return None
    try:
        return int(m.group(0))
    except Exception:
        return None


def _write_row(ws, r: int, values: Sequence[object], *, start_col: int = 2,
               font: Optional[Font] = None, fill: Optional[PatternFill] = None,
               align: Optional[Alignment] = None, num_fmt: Optional[str] = None):
    for i, v in enumerate(values):
        c = ws.cell(row=r, column=start_col + i, value=v)
        if font is not None:
            c.font = font
        if fill is not None:
            c.fill = fill
        if align is not None:
            c.alignment = align
        if num_fmt is not None:
            c.number_format = num_fmt


def _set_cell(ws, r: int, c: int, v=None, *, font: Optional[Font] = None,
              fill: Optional[PatternFill] = None, align: Optional[Alignment] = None,
              num_fmt: Optional[str] = None):
    cell = ws.cell(row=r, column=c)
    if v is not None:
        cell.value = v
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if align is not None:
        cell.alignment = align
    if num_fmt is not None:
        cell.number_format = num_fmt
    return cell


# -------- period parsing (общ!A2) --------
_MONTHS = {
    "январь": 1, "января": 1,
    "февраль": 2, "февраля": 2,
    "март": 3, "марта": 3,
    "апрель": 4, "апреля": 4,
    "май": 5, "мая": 5,
    "июнь": 6, "июня": 6,
    "июль": 7, "июля": 7,
    "август": 8, "августа": 8,
    "сентябрь": 9, "сентября": 9,
    "октябрь": 10, "октября": 10,
    "ноябрь": 11, "ноября": 11,
    "декабрь": 12, "декабря": 12,
}


def _parse_obsh_period_from_a2(a2: str) -> Optional[Tuple[int, int, int, int, str]]:
    s = _as_text(a2)
    if not s:
        return None
    s_low = s.lower()
    if "за" in s_low:
        s2 = s[s_low.find("за") + 2 :].strip()
    else:
        s2 = s.strip()
    s2 = s2.replace("г.", "").replace("г", "").strip(" .")

    m = re.fullmatch(r"(\d{4})", s2.strip())
    if m:
        y = int(m.group(1))
        return y, 1, y, 12, s2

    m = re.search(r"(\d{2})\.(\d{2})\.(\d{4})\s*[-–]\s*(\d{2})\.(\d{2})\.(\d{4})", s2)
    if m:
        m0, y0 = int(m.group(2)), int(m.group(3))
        m1, y1 = int(m.group(5)), int(m.group(6))
        return y0, m0, y1, m1, s2

    m = re.search(r"([а-я]+)\s+(\d{4})\s*[-–]\s*([а-я]+)\s+(\d{4})", s2, flags=re.I)
    if m:
        m0s, y0 = m.group(1).lower(), int(m.group(2))
        m1s, y1 = m.group(3).lower(), int(m.group(4))
        if m0s in _MONTHS and m1s in _MONTHS:
            return y0, _MONTHS[m0s], y1, _MONTHS[m1s], s2

    m = re.search(r"([а-я]+)\s+(\d{4})", s2, flags=re.I)
    if m:
        ms, y = m.group(1).lower(), int(m.group(2))
        if ms in _MONTHS:
            mm = _MONTHS[ms]
            return y, mm, y, mm, s2

    m = re.search(r"(\d{4})", s2)
    if m:
        y = int(m.group(1))
        return y, 1, y, 12, s2

    return None


def _month_iter(y0: int, m0: int, y1: int, m1: int) -> List[Tuple[int, int]]:
    out: List[Tuple[int, int]] = []
    y, m = y0, m0
    while (y < y1) or (y == y1 and m <= m1):
        out.append((y, m))
        m += 1
        if m == 13:
            m = 1
            y += 1
    return out


def _fmt_period_short(y0: int, m0: int, y1: int, m1: int) -> str:
    if y0 == y1 and m0 == m1:
        return f"{m0:02d}.{y0}"
    return f"{m0:02d}.{y0} – {m1:02d}.{y1}"


def _ym_key(ym: Tuple[int, int]) -> int:
    return ym[0] * 12 + ym[1]


def _iter_month_rows(ws, *, start_row: int, year_col: int, month_col: int) -> Iterable[Tuple[int, int, int]]:
    r = start_row
    max_r = int(ws.max_row or 1)
    while r <= max_r:
        y = _to_intish(ws.cell(row=r, column=year_col).value)
        if y is None:
            break
        m = _to_intish(ws.cell(row=r, column=month_col).value)
        if m is None or not (1 <= m <= 12):
            break
        yield r, int(y), int(m)
        r += 1


def _m_revenue_value(ws_m, rr: int) -> float:
    g = ws_m.cell(row=rr, column=7).value
    if isinstance(g, (int, float)):
        return float(g)
    return sum(_to_float(ws_m.cell(row=rr, column=c).value) for c in (17, 18, 19))


def _sum_m_revenue_for_period(ws_m, months: List[Tuple[int, int]], *, times_1000: bool) -> float:
    want = {f"{y:04d}_{m:02d}" for (y, m) in months}
    total = 0.0
    for rr, y, m in _iter_month_rows(ws_m, start_row=12, year_col=2, month_col=3):
        if f"{y:04d}_{m:02d}" in want:
            total += _m_revenue_value(ws_m, rr)
    return total * (1000.0 if times_1000 else 1.0)


def _sum_w_cfo_for_period(ws_w, months: List[Tuple[int, int]]) -> float:
    want = {f"{y:04d}_{m:02d}" for (y, m) in months}
    total = 0.0
    for rr, y, m in _iter_month_rows(ws_w, start_row=12, year_col=2, month_col=3):
        if f"{y:04d}_{m:02d}" in want:
            total += _to_float(ws_w.cell(row=rr, column=6).value)
    return total


def _last_real_month_m(ws_m) -> Optional[Tuple[int, int]]:
    last = None
    for rr, y, m in _iter_month_rows(ws_m, start_row=12, year_col=2, month_col=3):
        if abs(_m_revenue_value(ws_m, rr)) > 0:
            last = (y, m)
    return last


def _last_real_month_w(ws_w) -> Optional[Tuple[int, int]]:
    last = None
    for rr, y, m in _iter_month_rows(ws_w, start_row=12, year_col=2, month_col=3):
        cfo = _to_float(ws_w.cell(row=rr, column=6).value)
        clients = _to_float(ws_w.cell(row=rr, column=7).value)
        if abs(cfo) > 0 or abs(clients) > 0:
            last = (y, m)
    return last


def _parse_obsh_accounts(ws_obsh) -> Tuple[Dict[str, Dict[str, float]], Optional[int]]:
    acc: Dict[str, Dict[str, float]] = {}
    itogo_row: Optional[int] = None
    max_r = int(ws_obsh.max_row or 1)
    for r in range(6, max_r + 1):
        a = _as_text(ws_obsh.cell(row=r, column=1).value).strip()
        if not a:
            continue
        if a.strip().lower() == "итого":
            itogo_row = r
            break
        m = re.match(r'^"?(\d{4})', a)
        if not m:
            continue
        code = m.group(1)
        acc[code] = {
            "E": _to_float(ws_obsh.cell(row=r, column=5).value),
            "F": _to_float(ws_obsh.cell(row=r, column=6).value),
            "G": _to_float(ws_obsh.cell(row=r, column=7).value),
            "H": _to_float(ws_obsh.cell(row=r, column=8).value),
        }
    return acc, itogo_row



# =============================================================================
# Blocks 1-5 (по ТЗ)
# =============================================================================

FILL_HIGH = PatternFill("solid", fgColor="E8F5E9")
FILL_LOW  = PatternFill("solid", fgColor="FDE8E8")


def _period(y: int, m: int) -> str:
    return f"{y:04d}_{m:02d}"


def _is_next_month(prev: Tuple[int, int], cur: Tuple[int, int]) -> bool:
    py, pm = prev
    cy, cm = cur
    return (cy == py and cm == pm + 1) or (cy == py + 1 and pm == 12 and cm == 1)


def _parse_tables_generic(
    ws,
    *,
    start_row: int,
    name_col: int,
    first_month_col: int,
    total_col: int,
    header_month: str = "2025_01",
    include_tables: Optional[set[str]] = None,
    exclude_tables: Optional[set[str]] = None,
    skip_names: Optional[set[str]] = None,
) -> List[dict]:
    """Парсинг блок-таблиц Wt/кред.

    Заголовок таблицы: (B непусто) и (C == '2025_01').
    Контрагенты: строки до 'Доля', исключая Топ/Всего/Доля и (в кред) Тело/Проценты/Платежи.
    """

    if skip_names is None:
        skip_names = {"Топ", "Всего", "Доля", "Тело", "Проценты", "Платежи", ""}

    out: List[dict] = []
    r = start_row
    max_r = int(ws.max_row or 1)

    while r <= max_r:
        tbl_name = _as_text(ws.cell(row=r, column=name_col).value).strip()
        first_month = _as_text(ws.cell(row=r, column=first_month_col).value).strip()

        if tbl_name and tbl_name not in skip_names and (first_month == header_month or re.match(r"^\d{4}_\d{2}$", first_month)):
            if include_tables is not None and tbl_name not in include_tables:
                rr = r + 1
                while rr <= max_r and _as_text(ws.cell(row=rr, column=name_col).value).strip() != "Доля":
                    rr += 1
                r = rr + 3
                continue

            if exclude_tables is not None and tbl_name in exclude_tables:
                rr = r + 1
                while rr <= max_r and _as_text(ws.cell(row=rr, column=name_col).value).strip() != "Доля":
                    rr += 1
                r = rr + 3
                continue

            cps: List[dict] = []
            rr = r + 1
            while rr <= max_r:
                nm = _as_text(ws.cell(row=rr, column=name_col).value).strip()
                if nm == "Доля":
                    break
                if nm and nm not in skip_names:
                    total = _to_float(ws.cell(row=rr, column=total_col).value)
                    cps.append({"name": nm, "total": total})
                rr += 1

            if cps:
                out.append({"name": tbl_name, "counterparties": cps})

            r = rr + 3
        else:
            r += 1

    return out


def _write_block1(ws, row: int, *, ws_wt, ws_cred=None) -> int:
    _write_row(ws, row, ["1. ПЕРЕСЕЧЕНИЯ КОНТРАГЕНТОВ (Wt / кред)"], font=F_HDR, align=A_LEFT)
    row += 1
    _write_row(ws, row, ["Контрагенты, присутствующие в 2+ таблицах"], font=F_GRAY_ITALIC, align=A_LEFT)
    row += 1
    _write_row(ws, row, ["Контрагент", "Лист", "Таблица", "Направление", "Итого"], font=F_HDR, fill=FILL_HDR, align=A_LEFT)
    row += 1

    wt_excl = {"Прочая ДЗ/КЗ Inflow", "Прочая ДЗ/КЗ Outflow"}
    wt_tables = _parse_tables_generic(ws_wt, start_row=3, name_col=2, first_month_col=3, total_col=29, exclude_tables=wt_excl)

    cred_tables: List[dict] = []
    if ws_cred is not None:
        cred_include = {"Краткосрочные кредиты (Netto)", "Долгосрочные кредиты (Netto)"}
        cred_tables = _parse_tables_generic(ws_cred, start_row=5, name_col=2, first_month_col=3, total_col=29, include_tables=cred_include)

    occ: Dict[str, List[Tuple[str, str, float]]] = {}
    for t in wt_tables:
        for cp in t["counterparties"]:
            occ.setdefault(cp["name"], []).append(("Wt", t["name"], float(cp["total"])))
    for t in cred_tables:
        for cp in t["counterparties"]:
            occ.setdefault(cp["name"], []).append(("кред", t["name"], float(cp["total"])))

    inter: Dict[str, List[Tuple[str, str, float]]] = {}
    for name, items in occ.items():
        uniq = {(sh, tbl) for (sh, tbl, _) in items}
        if len(uniq) >= 2:
            inter[name] = items

    if not inter:
        _write_row(ws, row, ["Данных нет"], font=F_GRAY_ITALIC, align=A_LEFT)
        return row + 3

    company_tbls = {
        "Клиенты",
        "Пост-ки",
        "Прочая ДЗ/КЗ",
        "Прочие займы",
        "Краткосрочные кредиты (Netto)",
        "Долгосрочные кредиты (Netto)",
    }
    people_tbls = {"Подотчет", "3350", "73(&)"}

    companies: List[str] = []
    persons: List[str] = []

    for name, items in inter.items():
        tbls = {tbl for _, tbl, _ in items}
        if tbls.issubset(people_tbls):
            persons.append(name)
        elif tbls & company_tbls:
            companies.append(name)
        else:
            companies.append(name)

    def write_group(names: List[str], *, sep: Optional[str] = None) -> None:
        nonlocal row
        if not names:
            return
        if sep:
            _write_row(ws, row, [sep], font=F_GRAY_ITALIC, align=A_LEFT)
            row += 1
        for nm in sorted(names):
            items = inter[nm]
            first = True
            for sh, tbl, total in items:
                direction = "Приход" if total >= 0 else "Расход"
                _write_row(ws, row, [nm if first else "", sh, tbl, direction, total], align=A_LEFT)
                ws.cell(row=row, column=6).number_format = NUM
                ws.cell(row=row, column=6).font = F_GREEN_BOLD if total >= 0 else F_RED_BOLD
                first = False
                row += 1

    write_group(companies)
    write_group(persons, sep="Физические лица (подотчет + персонал):")

    return row + 2


def _write_block2(ws, row: int, *, ws_m) -> int:
    _write_row(ws, row, ["2. РЕГУЛЯРНОСТЬ ВЫРУЧКИ (M)"], font=F_HDR, align=A_LEFT)
    row += 1

    rows: List[Tuple[int, int, int, float]] = []
    for rr, y, m in _iter_month_rows(ws_m, start_row=12, year_col=2, month_col=3):
        rows.append((rr, y, m, float(_m_revenue_value(ws_m, rr))))

    if not rows:
        _write_row(ws, row, ["Данных нет"], font=F_GRAY_ITALIC, align=A_LEFT)
        return row + 3

    last_idx = None
    for i, (_rr, _y, _m, v) in enumerate(rows):
        if abs(v) > 0:
            last_idx = i
    if last_idx is None:
        _write_row(ws, row, ["Данных нет"], font=F_GRAY_ITALIC, align=A_LEFT)
        return row + 3

    rows = rows[: last_idx + 1]

    nonzero = [v for *_rest, v in rows if v > 0]
    if not nonzero:
        _write_row(ws, row, ["Данных нет"], font=F_GRAY_ITALIC, align=A_LEFT)
        return row + 3

    avg = sum(nonzero) / len(nonzero)
    thr = 0.10 * avg

    # breaks
    breaks: List[Tuple[str, str, int]] = []
    cur_start: Optional[Tuple[int, int]] = None
    cur_len = 0
    prev_ym: Optional[Tuple[int, int]] = None

    for (_rr, y, m, v) in rows:
        ym = (y, m)
        low = v < thr
        if low:
            if cur_start is None:
                cur_start = ym
                cur_len = 1
            else:
                if prev_ym is not None and _is_next_month(prev_ym, ym):
                    cur_len += 1
                else:
                    breaks.append((_period(*cur_start), _period(*prev_ym), cur_len))
                    cur_start = ym
                    cur_len = 1
        else:
            if cur_start is not None and prev_ym is not None:
                breaks.append((_period(*cur_start), _period(*prev_ym), cur_len))
                cur_start = None
                cur_len = 0
        prev_ym = ym

    if cur_start is not None and prev_ym is not None:
        breaks.append((_period(*cur_start), _period(*prev_ym), cur_len))

    if not breaks:
        _write_row(ws, row, ["Перерывов не выявлено"], font=F_GREEN_BOLD, align=A_LEFT)
        row += 1
    else:
        for b0, b1, ln in breaks:
            _write_row(ws, row, [f"Перерывы: {b0} — {b1} ({ln} мес. выручка < порога)"], font=F_HDR, align=A_LEFT)
            row += 1

    # anomalies
    revs = [v for *_rest, v in rows]
    yms = [(y, m) for (_rr, y, m, _v) in rows]
    probs: List[Tuple[str, float, float, str]] = []

    for i in range(1, len(revs)):
        prev_vals = revs[max(0, i - 3) : i]
        avg3 = sum(prev_vals) / len(prev_vals) if prev_vals else 0.0
        if avg3 <= 0:
            continue
        v = revs[i]
        if v == 0 and avg3 > 0:
            probs.append((_period(*yms[i]), v, avg3, "Нулевая выручка"))
        elif v < 0.3 * avg3:
            pct = int(round((1 - (v / avg3)) * 100))
            probs.append((_period(*yms[i]), v, avg3, f"Падение {pct}%"))

    if probs:
        _write_row(ws, row, ["Период", "Выручка", "Ср. 3 мес.", "Статус"], font=F_HDR, fill=FILL_HDR, align=A_LEFT)
        row += 1
        for per, v, a3, st in probs:
            _write_row(ws, row, [per, v, a3, st], align=A_LEFT)
            ws.cell(row=row, column=3).number_format = NUM
            ws.cell(row=row, column=4).number_format = NUM
            ws.cell(row=row, column=5).font = F_RED_BOLD
            row += 1
    else:
        _write_row(ws, row, ["Проблемных месяцев не выявлено"], font=F_GREEN_BOLD, align=A_LEFT)
        row += 1

    return row + 2


def _write_block3(ws, row: int, *, ws_w) -> int:
    _write_row(ws, row, ["3. РЕГУЛЯРНОСТЬ ПЕРСОНАЛА И НАЛОГОВ (W)"], font=F_HDR, align=A_LEFT)
    row += 1
    _write_row(ws, row, ["Аномалии: падение >50% от скользящего среднего за 3 мес."], font=F_GRAY_ITALIC, align=A_LEFT)
    row += 1

    rows: List[Tuple[int, int, int, float, float, float, float]] = []
    for rr, y, m in _iter_month_rows(ws_w, start_row=12, year_col=2, month_col=3):
        cfo = _to_float(ws_w.cell(row=rr, column=6).value)
        clients = _to_float(ws_w.cell(row=rr, column=7).value)
        pers = _to_float(ws_w.cell(row=rr, column=18).value)   # R
        taxes = _to_float(ws_w.cell(row=rr, column=19).value)  # S
        rows.append((rr, y, m, cfo, clients, pers, taxes))

    if not rows:
        _write_row(ws, row, ["Данных нет"], font=F_GRAY_ITALIC, align=A_LEFT)
        return row + 3

    last_idx = None
    for i, (_rr, _y, _m, cfo, clients, *_rest) in enumerate(rows):
        if abs(cfo) > 0 or abs(clients) > 0:
            last_idx = i
    if last_idx is None:
        _write_row(ws, row, ["Данных нет"], font=F_GRAY_ITALIC, align=A_LEFT)
        return row + 3

    rows = rows[: last_idx + 1]
    yms = [(y, m) for (_rr, y, m, *_rest) in rows]
    pers_vals = [pers for (*_a, pers, _t) in rows]
    tax_vals = [taxes for (*_a, _p, taxes) in rows]

    def anomalies(vals: List[float]) -> List[Tuple[str, float, float, str]]:
        out: List[Tuple[str, float, float, str]] = []
        for i in range(1, len(vals)):
            prev_vals = vals[max(0, i - 3) : i]
            avg3 = sum(prev_vals) / len(prev_vals) if prev_vals else 0.0
            if avg3 <= 1000:
                continue
            v = vals[i]
            if v == 0:
                out.append((_period(*yms[i]), v, avg3, "Нулевой платеж"))
            elif v < 0.5 * avg3:
                pct = int(round((1 - (v / avg3)) * 100))
                out.append((_period(*yms[i]), v, avg3, f"Падение {pct}%"))
        return out

    # Personnel
    _write_row(ws, row, ["Персонал"], font=F_HDR, align=A_LEFT)
    row += 1
    pers_an = anomalies(pers_vals)
    if not pers_an:
        _write_row(ws, row, ["Аномалий не выявлено"], font=F_GREEN_BOLD, align=A_LEFT)
        row += 1
    else:
        _write_row(ws, row, ["Период", "Сумма", "Ср. 3 мес.", "Статус"], font=F_HDR, fill=FILL_PROB, align=A_LEFT)
        row += 1
        for per, v, a3, st in pers_an:
            _write_row(ws, row, [per, v, a3, st], align=A_LEFT)
            ws.cell(row=row, column=3).number_format = NUM
            ws.cell(row=row, column=4).number_format = NUM
            ws.cell(row=row, column=5).font = F_RED_BOLD
            row += 1

    # Taxes
    _write_row(ws, row, ["Налоги"], font=F_HDR, align=A_LEFT)
    row += 1
    tax_an = anomalies(tax_vals)
    if not tax_an:
        _write_row(ws, row, ["Аномалий не выявлено"], font=F_GREEN_BOLD, align=A_LEFT)
        row += 1
    else:
        _write_row(ws, row, ["Период", "Сумма", "Ср. 3 мес.", "Статус"], font=F_HDR, fill=FILL_PROB, align=A_LEFT)
        row += 1
        for per, v, a3, st in tax_an:
            _write_row(ws, row, [per, v, a3, st], align=A_LEFT)
            ws.cell(row=row, column=3).number_format = NUM
            ws.cell(row=row, column=4).number_format = NUM
            ws.cell(row=row, column=5).font = F_RED_BOLD
            row += 1

    return row + 2


def _parse_mt(ws_mt) -> Tuple[List[Tuple[str, float]], float]:
    cps: List[Tuple[str, float]] = []
    total = 0.0
    max_r = int(ws_mt.max_row or 1)

    for r in range(6, max_r + 1):
        name = _as_text(ws_mt.cell(row=r, column=2).value).strip()
        if not name:
            break
        if name == "Всего":
            total = _to_float(ws_mt.cell(row=r, column=29).value)
            break
        if name in {"Топ", "Доля"}:
            break
        cps.append((name, _to_float(ws_mt.cell(row=r, column=29).value)))

    if total == 0.0:
        for r in range(6, max_r + 1):
            name = _as_text(ws_mt.cell(row=r, column=2).value).strip()
            if name == "Всего":
                total = _to_float(ws_mt.cell(row=r, column=29).value)
                break

    return cps, total


def _find_wt_table(ws_wt, table_name: str) -> Optional[int]:
    max_r = int(ws_wt.max_row or 1)
    for r in range(3, max_r + 1):
        nm = _as_text(ws_wt.cell(row=r, column=2).value).strip()
        first = _as_text(ws_wt.cell(row=r, column=3).value).strip()
        if nm == table_name and re.match(r"^\d{4}_\d{2}$", first):
            return r
    return None


def _parse_wt_clients(ws_wt) -> Tuple[List[Tuple[str, float]], float, List[float], List[float]]:
    """Контрагенты + итоги по таблице 'Клиенты' на Wt."""
    header_r = _find_wt_table(ws_wt, "Клиенты")
    if header_r is None:
        return [], 0.0, [0.0] * 12, [0.0] * 12

    cps: List[Tuple[str, float]] = []
    total_all = 0.0
    max_r = int(ws_wt.max_row or 1)

    r = header_r + 1
    while r <= max_r:
        nm = _as_text(ws_wt.cell(row=r, column=2).value).strip()
        if nm in {"Топ", "Всего", "Доля", ""}:
            break
        cps.append((nm, _to_float(ws_wt.cell(row=r, column=29).value)))
        r += 1

    rr = header_r + 1
    while rr <= max_r:
        nm = _as_text(ws_wt.cell(row=rr, column=2).value).strip()
        if nm == "Всего":
            total_all = _to_float(ws_wt.cell(row=rr, column=29).value)
            m2025 = [_to_float(ws_wt.cell(row=rr, column=c).value) for c in range(3, 15)]   # C..N
            m2026 = [_to_float(ws_wt.cell(row=rr, column=c).value) for c in range(15, 27)]  # O..Z
            return cps, total_all, m2025, m2026
        if nm == "Доля":
            break
        rr += 1

    return cps, total_all, [0.0] * 12, [0.0] * 12


def _write_block4(ws, row: int, *, ws_mt, ws_wt) -> int:
    _write_row(ws, row, ["4. КОНЦЕНТРАЦИЯ КЛИЕНТОВ"], font=F_HDR, align=A_LEFT)
    row += 1

    # Mt
    _write_row(ws, row, ["Выручка (Mt)"], font=F_HDR, align=A_LEFT)
    row += 1
    cps, total = _parse_mt(ws_mt)
    if total <= 0:
        _write_row(ws, row, ["Данных нет"], font=F_GRAY_ITALIC, align=A_LEFT)
        row += 2
    else:
        cps_sorted = sorted(cps, key=lambda x: x[1], reverse=True)
        _write_row(ws, row, ["Клиент", "Сумма", "Доля"], font=F_HDR, fill=FILL_HDR, align=A_LEFT)
        row += 1
        for nm, v in cps_sorted:
            _write_row(ws, row, [nm, v, (v / total) if total else 0], align=A_LEFT)
            ws.cell(row=row, column=3).number_format = NUM
            ws.cell(row=row, column=4).number_format = "0.0%"
            if total and (v / total) > 0.70:
                ws.cell(row=row, column=4).font = F_RED_BOLD
            row += 1
        top_nm, top_v = cps_sorted[0]
        share = (top_v / total) if total else 0.0
        if share > 0.70:
            _write_row(ws, row, [f"ВЫСОКАЯ КОНЦЕНТРАЦИЯ: {share:.1%} выручки — один клиент ({top_nm})"], font=F_HDR, align=A_LEFT)
            row += 1
        elif share > 0.50:
            _write_row(ws, row, [f"ЗНАЧИТЕЛЬНАЯ КОНЦЕНТРАЦИЯ: {share:.1%} выручки — один клиент ({top_nm})"], font=F_ORANGE_BOLD, align=A_LEFT)
            row += 1
        _write_row(ws, row, [f"Всего клиентов: {len(cps_sorted)}"], font=F_DEF, align=A_LEFT)
        row += 2

    # Wt
    _write_row(ws, row, ["Поступления от клиентов (Wt)"], font=F_HDR, align=A_LEFT)
    row += 1
    wt_cps, wt_total, _m25, _m26 = _parse_wt_clients(ws_wt)
    if wt_total <= 0:
        _write_row(ws, row, ["Данных нет"], font=F_GRAY_ITALIC, align=A_LEFT)
        row += 1
    else:
        wt_sorted = sorted(wt_cps, key=lambda x: x[1], reverse=True)
        _write_row(ws, row, ["Клиент", "Сумма", "Доля"], font=F_HDR, fill=FILL_HDR, align=A_LEFT)
        row += 1
        for nm, v in wt_sorted:
            _write_row(ws, row, [nm, v, (v / wt_total) if wt_total else 0], align=A_LEFT)
            ws.cell(row=row, column=3).number_format = NUM
            ws.cell(row=row, column=4).number_format = "0.0%"
            if wt_total and (v / wt_total) > 0.70:
                ws.cell(row=row, column=4).font = F_RED_BOLD
            row += 1
        top_nm, top_v = wt_sorted[0]
        share = (top_v / wt_total) if wt_total else 0.0
        if share > 0.70:
            _write_row(ws, row, [f"ВЫСОКАЯ КОНЦЕНТРАЦИЯ: {share:.1%} поступлений — один клиент ({top_nm})"], font=F_HDR, align=A_LEFT)
            row += 1
        elif share > 0.50:
            _write_row(ws, row, [f"ЗНАЧИТЕЛЬНАЯ КОНЦЕНТРАЦИЯ: {share:.1%} поступлений — один клиент ({top_nm})"], font=F_ORANGE_BOLD, align=A_LEFT)
            row += 1
        _write_row(ws, row, [f"Всего клиентов: {len(wt_sorted)}"], font=F_DEF, align=A_LEFT)
        row += 1

    return row + 2


def _write_block5(ws, row: int, *, ws_wt) -> int:
    _write_row(ws, row, ["5. СЕЗОННОСТЬ ПОСТУПЛЕНИЙ ОТ КЛИЕНТОВ (Wt)"], font=F_HDR, align=A_LEFT)
    row += 1

    _cps, _tot, m2025, m2026 = _parse_wt_clients(ws_wt)
    if not any(abs(v) > 0 for v in m2025):
        _write_row(ws, row, ["Данных нет"], font=F_GRAY_ITALIC, align=A_LEFT)
        return row + 3

    avg = sum(m2025) / 12.0
    peaks = [i + 1 for i, v in enumerate(m2025) if v > 1.5 * avg]
    lows = [i + 1 for i, v in enumerate(m2025) if v < 0.3 * avg]

    last26 = 0
    for i, v in enumerate(m2026, start=1):
        if abs(v) > 0:
            last26 = i

    peaks_s = ", ".join(str(x) for x in peaks) if peaks else "—"
    lows_s = ", ".join(str(x) for x in lows) if lows else "—"

    _write_row(ws, row, [f"Среднее за 2025: {int(round(avg)):,} / мес. Пиковые: {peaks_s}. Провалы: {lows_s}".replace(",", " ")], font=F_GRAY_ITALIC, align=A_LEFT)
    row += 1

    months_ru = ["", "янв", "фев", "мар", "апр", "май", "июн", "июл", "авг", "сен", "окт", "ноя", "дек"]
    _write_row(ws, row, months_ru, font=F_HDR, fill=FILL_HDR, align=A_CENTER)
    row += 1

    _set_cell(ws, row, 2, "2025", font=F_HDR, align=A_CENTER)
    for i, v in enumerate(m2025, start=1):
        c = ws.cell(row=row, column=2 + i)
        c.value = v
        c.number_format = NUM
        c.alignment = A_CENTER
        if i in lows:
            c.fill = FILL_LOW
            c.font = F_RED_BOLD
        if i in peaks:
            c.fill = FILL_HIGH
            c.font = F_GREEN_BOLD
    row += 1

    _set_cell(ws, row, 2, "2026", font=F_HDR, align=A_CENTER)
    for i, v in enumerate(m2026, start=1):
        c = ws.cell(row=row, column=2 + i)
        c.alignment = A_CENTER
        if last26 and i > last26:
            c.value = ""
            c.font = Font(name="Calibri", size=11, color="CCCCCC")
        else:
            c.value = v
            c.number_format = NUM
    row += 1

    _set_cell(ws, row, 2, "Ср. мес.", font=Font(name="Calibri", size=11, bold=True, italic=True, color=C_GRAY), align=A_CENTER)
    for i in range(1, 13):
        c = ws.cell(row=row, column=2 + i)
        c.value = avg
        c.number_format = NUM
        c.font = Font(name="Calibri", size=11, italic=True, color=C_GRAY)
        c.alignment = A_CENTER
    row += 1

    nonzero25 = [v for v in m2025 if v > 0]
    if nonzero25:
        mn = min(nonzero25)
        mx = max(nonzero25)
        spread = (mx / mn) if mn else 0.0
        _write_row(ws, row, [f"Поступления крайне неравномерны: мин. {int(round(mn)):,}, макс. {int(round(mx)):,} — разброс в {spread:.1f}x".replace(",", " ")], font=F_ORANGE_BOLD, align=A_LEFT)
        row += 1

    if lows:
        _write_row(ws, row, ["Риск кассового разрыва в периоды низких поступлений при сохранении фиксированных расходов"], font=F_HDR, align=A_LEFT)
        row += 1

    return row + 2





def _write_block6(ws, row: int, *, ws_obsh, ws_m, ws_w) -> int:
    STOP = "\u0421\u0422\u041e\u041f-\u0424\u0410\u041a\u0422\u041e\u0420"
    RED  = "\u041a\u0420\u0410\u0421\u041d\u042b\u0419 \u0424\u041b\u0410\u0413"
    YEL  = "\u0416\u0401\u041b\u0422\u042b\u0419 \u0424\u041b\u0410\u0413"
    POS  = "\u041f\u041e\u0417\u0418\u0422\u0418\u0412"

    p = _parse_obsh_period_from_a2(_as_text(ws_obsh["A2"].value)) if ws_obsh is not None else None
    if p is None:
        years = [y for _, y, _ in _iter_month_rows(ws_m, start_row=12, year_col=2, month_col=3)]
        yy = max(years) if years else 2025
        y0, m0, y1, m1, raw = yy, 1, yy, 12, str(yy)
    else:
        y0, m0, y1, m1, raw = p

    period_short = _fmt_period_short(y0, m0, y1, m1)
    months = _month_iter(y0, m0, y1, m1)

    last_m = _last_real_month_m(ws_m)
    last_w = _last_real_month_w(ws_w)
    last_common = last_m or last_w
    if last_m and last_w:
        last_common = last_m if _ym_key(last_m) <= _ym_key(last_w) else last_w
    if last_common:
        months = [ym for ym in months if _ym_key(ym) <= _ym_key(last_common)]

    acc_map, itogo_row = ({}, None)
    if ws_obsh is not None:
        acc_map, itogo_row = _parse_obsh_accounts(ws_obsh)

    def end_db(code: str) -> float:
        return float((acc_map.get(code) or {}).get("G", 0.0) or 0.0)

    def end_cr(code: str) -> float:
        return float((acc_map.get(code) or {}).get("H", 0.0) or 0.0)

    def turn_db(code: str) -> float:
        return float((acc_map.get(code) or {}).get("E", 0.0) or 0.0)

    def turn_cr(code: str) -> float:
        return float((acc_map.get(code) or {}).get("F", 0.0) or 0.0)

    balance_currency = 0.0
    if ws_obsh is not None and itogo_row is not None:
        balance_currency = _to_float(ws_obsh.cell(row=itogo_row, column=7).value)

    revenue_for_ratios = _sum_m_revenue_for_period(ws_m, months, times_1000=True)
    revenue_period_ths = _sum_m_revenue_for_period(ws_m, months, times_1000=False)
    cfo_period_ths = _sum_w_cfo_for_period(ws_w, months)

    base_year = int(y0)
    prev_year = base_year - 1

    def sum_m_year_ths(year: int) -> float:
        tot = 0.0
        for rr, y, m in _iter_month_rows(ws_m, start_row=12, year_col=2, month_col=3):
            if y == year:
                tot += _m_revenue_value(ws_m, rr)
        return tot

    rev_base = sum_m_year_ths(base_year) * 1000.0
    rev_prev = sum_m_year_ths(prev_year) * 1000.0

    rev_6010 = turn_cr("6010")
    cogs_7010 = turn_db("7010")
    sga_7110 = turn_db("7110")
    ga_7210 = turn_db("7210")
    op_profit = rev_6010 - cogs_7010 - sga_7110 - ga_7210

    oa_codes = ["1030", "1050", "1060", "1100", "1210", "1250", "1270", "1300", "1400", "1710", "1720", "1730"]
    ko_codes = ["3010", "3100", "3200", "3300", "3400", "3500"]

    _write_row(ws, row, ["6. \u0424\u0418\u041d\u0410\u041d\u0421\u041e\u0412\u042b\u0415 \u041f\u041e\u041a\u0410\u0417\u0410\u0422\u0415\u041b\u0418 \u0418 \u0424\u041b\u0410\u0413\u0418 (\u043e\u0431\u0449 / M / W)"], font=F_HDR, align=A_LEFT)
    row += 1
    _write_row(ws, row, [f"\u041e\u0421\u0412 \u0437\u0430 {raw}. \u0412\u044b\u0440\u0443\u0447\u043a\u0430 \u0434\u043b\u044f \u043a\u043e\u044d\u0444\u0444\u0438\u0446\u0438\u0435\u043d\u0442\u043e\u0432 \u2014 \u0438\u0437 M \u0437\u0430 \u0442\u043e\u0442 \u0436\u0435 \u043f\u0435\u0440\u0438\u043e\u0434 (\u00d71000, \u0442\u0433)"], font=Font(name="Calibri", size=9, italic=True), align=A_LEFT)
    row += 2
    _write_row(ws, row, ["\u041f\u043e\u043a\u0430\u0437\u0430\u0442\u0435\u043b\u044c", "\u0421\u0447\u0451\u0442 / \u0418\u0441\u0442\u043e\u0447\u043d\u0438\u043a", "\u0417\u043d\u0430\u0447\u0435\u043d\u0438\u0435", "\u041a\u043e\u044d\u0444\u0444\u0438\u0446\u0438\u0435\u043d\u0442", "\u0424\u043b\u0430\u0433"], font=F_HDR, fill=FILL_HDR, align=A_LEFT)
    row += 1

    equity_row = row
    _set_cell(ws, row, 2, "\u041a\u0430\u043f\u0438\u0442\u0430\u043b (Equity)", font=F_HDR, align=A_LEFT)
    _set_cell(ws, row, 3, "5030 + 5600", font=F_DEF, align=A_LEFT)
    row += 1
    r5030 = row
    _set_cell(ws, row, 2, "\u0423\u0441\u0442\u0430\u0432\u043d\u044b\u0439 \u043a\u0430\u043f\u0438\u0442\u0430\u043b", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 3, "5030", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 4, round(end_cr("5030")), font=F_DEF, align=A_RIGHT, num_fmt=NUM)
    row += 1
    r5600 = row
    _set_cell(ws, row, 2, "\u041d\u0435\u0440\u0430\u0441\u043f\u0440\u0435\u0434. \u043f\u0440\u0438\u0431\u044b\u043b\u044c", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 3, "5600", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 4, round(end_cr("5600")), font=F_DEF, align=A_RIGHT, num_fmt=NUM)
    _set_cell(ws, equity_row, 4, f"=SUM(D{r5030}:D{r5600})", font=F_HDR, align=A_RIGHT, num_fmt=NUM)
    _set_cell(ws, equity_row, 6, f"=IF(D{equity_row}<0,\"{STOP}\",\"OK\")", font=F_HDR, align=A_LEFT)
    row += 2

    bal_row = row
    _set_cell(ws, row, 2, "\u0412\u0430\u043b\u044e\u0442\u0430 \u0431\u0430\u043b\u0430\u043d\u0441\u0430", font=F_HDR, align=A_LEFT)
    _set_cell(ws, row, 3, "\u0418\u0442\u043e\u0433\u043e", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 4, round(balance_currency), font=F_DEF, align=A_RIGHT, num_fmt=NUM)
    row += 1

    vb_ratio_row = row
    _set_cell(ws, row, 2, "\u0412\u0430\u043b\u044e\u0442\u0430 \u0431\u0430\u043b\u0430\u043d\u0441\u0430 / \u0412\u044b\u0440\u0443\u0447\u043a\u0430", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 3, "\u0412\u0411 (\u043e\u0431\u0449) / \u0412\u044b\u0440\u0443\u0447\u043a\u0430 (M\u00d71000)", font=F_DEF, align=A_LEFT)
    rev_rat_row = row + 1
    _set_cell(ws, row, 5, f"=IF(D{rev_rat_row}=0,\"N/A\",D{bal_row}/D{rev_rat_row})", font=F_DEF, align=A_RIGHT, num_fmt=MULT)
    _set_cell(ws, row, 6, f"=IF(E{vb_ratio_row}=\"N/A\",\"\",IF(E{vb_ratio_row}<1,\"{YEL}\",\"OK\"))", font=F_DEF, align=A_LEFT)
    row += 1

    _set_cell(ws, row, 2, "\u0412\u044b\u0440\u0443\u0447\u043a\u0430 \u0434\u043b\u044f \u043a\u043e\u044d\u0444\u0444\u0438\u0446\u0438\u0435\u043d\u0442\u043e\u0432", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 3, f"M (\u0432\u044b\u0440\u0443\u0447\u043a\u0430 {period_short}) \u00d71000", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 4, round(revenue_for_ratios), font=F_DEF, align=A_RIGHT, num_fmt=NUM)
    row += 2

    _set_cell(ws, row, 2, "\u0414\u043e\u043b\u0433\u043e\u0432\u0430\u044f \u043d\u0430\u0433\u0440\u0443\u0437\u043a\u0430", font=F_HDR, align=A_LEFT)
    row += 1

    debt_row = row
    _set_cell(ws, row, 2, "\u041e\u0431\u0449\u0438\u0439 \u0434\u043e\u043b\u0433 (3010+4010)", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 3, "3010 + 4010", font=F_DEF, align=A_LEFT)
    row += 1

    r3010 = row
    _set_cell(ws, row, 2, "  \u041a\u0440\u0430\u0442\u043a\u043e\u0441\u0440\u043e\u0447\u043d\u044b\u0439 \u0434\u043e\u043b\u0433", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 3, "3010", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 4, round(end_cr("3010")), font=F_DEF, align=A_RIGHT, num_fmt=NUM)
    row += 1

    r4010 = row
    _set_cell(ws, row, 2, "  \u0414\u043e\u043b\u0433\u043e\u0441\u0440\u043e\u0447\u043d\u044b\u0439 \u0434\u043e\u043b\u0433", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 3, "4010", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 4, round(end_cr("4010")), font=F_DEF, align=A_RIGHT, num_fmt=NUM)
    _set_cell(ws, debt_row, 4, f"=SUM(D{r3010}:D{r4010})", font=F_DEF, align=A_RIGHT, num_fmt=NUM)
    row += 1

    debt_eq_row = row
    _set_cell(ws, row, 2, "\u0414\u043e\u043b\u0433 / \u041a\u0430\u043f\u0438\u0442\u0430\u043b", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 5, f"=IF(D{equity_row}=0,\"N/A\",D{debt_row}/D{equity_row})", font=F_HDR, align=A_RIGHT, num_fmt=MULT)
    _set_cell(ws, row, 6, f"=IF(E{debt_eq_row}=\"N/A\",\"\",IF(E{debt_eq_row}>10,\"{RED}\",\"OK\"))", font=F_DEF, align=A_LEFT)
    row += 1

    sh_debt_row = row
    _set_cell(ws, row, 2, "\u041a\u0440\u0430\u0442\u043a. \u0434\u043e\u043b\u0433 / \u041a\u0430\u043f\u0438\u0442\u0430\u043b", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 3, "3010 / \u041a\u0430\u043f\u0438\u0442\u0430\u043b", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 4, f"=D{r3010}", font=F_DEF, align=A_RIGHT, num_fmt=NUM)
    _set_cell(ws, row, 5, f"=IF(D{equity_row}=0,\"N/A\",D{r3010}/D{equity_row})", font=F_HDR, align=A_RIGHT, num_fmt=MULT)
    _set_cell(ws, row, 6, f"=IF(E{sh_debt_row}=\"N/A\",\"\",IF(E{sh_debt_row}>10,\"{YEL}\",\"OK\"))", font=F_DEF, align=A_LEFT)
    row += 2

    dz_row = row
    _set_cell(ws, row, 2, "\u0414\u0417 \u043f\u043e\u043a\u0443\u043f\u0430\u0442\u0435\u043b\u0435\u0439 (1210+1730)", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 3, "1210 + 1730", font=F_DEF, align=A_LEFT)
    row += 1

    dz_1210_row = row
    _set_cell(ws, row, 2, "  1210", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 3, "1210", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 4, round(end_db("1210")), font=F_DEF, align=A_RIGHT, num_fmt=NUM)
    row += 1

    dz_1730_row = row
    _set_cell(ws, row, 2, "  1730", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 3, "1730", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 4, round(end_db("1730")), font=F_DEF, align=A_RIGHT, num_fmt=NUM)

    _set_cell(ws, dz_row, 4, f"=SUM(D{dz_1210_row}:D{dz_1730_row})", font=F_DEF, align=A_RIGHT, num_fmt=NUM)
    _set_cell(ws, dz_row, 5, f"=IF(D{rev_rat_row}=0,\"N/A\",D{dz_row}/D{rev_rat_row})", font=F_DEF, align=A_RIGHT, num_fmt=PCT)
    _set_cell(ws, dz_row, 6, f"=IF(E{dz_row}=\"N/A\",\"\",IF(E{dz_row}>0.5,\"{RED}\",IF(E{dz_row}>0.3,\"{YEL}\",\"OK\")))", font=F_DEF, align=A_LEFT)
    row += 1

    kz_row = row
    _set_cell(ws, row, 2, "\u041a\u0417 \u043f\u043e\u0441\u0442\u0430\u0432\u0449\u0438\u043a\u0430\u043c (3310)", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 3, "3310", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 4, round(end_cr("3310")), font=F_DEF, align=A_RIGHT, num_fmt=NUM)
    _set_cell(ws, row, 5, f"=IF(D{rev_rat_row}=0,\"N/A\",D{kz_row}/D{rev_rat_row})", font=F_DEF, align=A_RIGHT, num_fmt=PCT)
    _set_cell(ws, row, 6, "OK", font=F_HDR, align=A_LEFT)
    row += 1

    advg_row = row
    _set_cell(ws, row, 2, "\u0410\u0432\u0430\u043d\u0441\u044b \u0432\u044b\u0434\u0430\u043d\u043d\u044b\u0435 (1710)", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 3, "1710", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 4, round(end_db("1710")), font=F_DEF, align=A_RIGHT, num_fmt=NUM)
    _set_cell(ws, row, 5, f"=IF(D{rev_rat_row}=0,\"N/A\",D{advg_row}/D{rev_rat_row})", font=F_DEF, align=A_RIGHT, num_fmt=PCT)
    _set_cell(ws, row, 6, f"=IF(E{advg_row}=\"N/A\",\"\",IF(E{advg_row}>0.5,\"{RED}\",IF(E{advg_row}>0.3,\"{YEL}\",\"OK\")))", font=F_DEF, align=A_LEFT)
    row += 1

    advr_row = row
    _set_cell(ws, row, 2, "\u0410\u0432\u0430\u043d\u0441\u044b \u043f\u043e\u043b\u0443\u0447\u0435\u043d\u043d\u044b\u0435 (3510)", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 3, "3510", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 4, round(end_cr("3510")), font=F_DEF, align=A_RIGHT, num_fmt=NUM)
    _set_cell(ws, row, 5, f"=IF(D{rev_rat_row}=0,\"N/A\",D{advr_row}/D{rev_rat_row})", font=F_DEF, align=A_RIGHT, num_fmt=PCT)
    _set_cell(ws, row, 6, "OK", font=F_HDR, align=A_LEFT)
    row += 2

    chok_row = row
    _set_cell(ws, row, 2, "\u041e\u0431\u043e\u0440\u043e\u0442\u043d\u044b\u0439 \u043a\u0430\u043f\u0438\u0442\u0430\u043b (\u0427\u041e\u041a)", font=F_HDR, align=A_LEFT)
    _set_cell(ws, row, 3, "\u041e\u0410 \u2212 \u041a\u041e", font=F_DEF, align=A_LEFT)
    row += 1

    oa_row = row
    _set_cell(ws, row, 2, "  \u041e\u0431\u043e\u0440\u043e\u0442\u043d\u044b\u0435 \u0430\u043a\u0442\u0438\u0432\u044b", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 4, 0, font=F_DEF, align=A_RIGHT, num_fmt=NUM)
    row += 1

    oa_acc_rows = []
    for code in oa_codes:
        rr = row
        _set_cell(ws, rr, 2, f"    {code}", font=F_DEF, align=A_LEFT)
        _set_cell(ws, rr, 3, code, font=F_DEF, align=A_LEFT)
        _set_cell(ws, rr, 4, round(end_db(code)), font=F_DEF, align=A_RIGHT, num_fmt=NUM)
        oa_acc_rows.append(rr)
        row += 1

    _set_cell(ws, oa_row, 4, "=SUM(" + ",".join(f"D{r}" for r in oa_acc_rows) + ")", font=F_DEF, align=A_RIGHT, num_fmt=NUM)

    ko_row = row
    _set_cell(ws, row, 2, "  \u041a\u0440\u0430\u0442\u043a\u043e\u0441\u0440\u043e\u0447\u043d\u044b\u0435 \u043e\u0431\u044f\u0437\u0430\u0442\u0435\u043b\u044c\u0441\u0442\u0432\u0430", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 4, 0, font=F_DEF, align=A_RIGHT, num_fmt=NUM)
    row += 1

    ko_acc_rows = []
    for code in ko_codes:
        rr = row
        _set_cell(ws, rr, 2, f"    {code}", font=F_DEF, align=A_LEFT)
        _set_cell(ws, rr, 3, code, font=F_DEF, align=A_LEFT)
        _set_cell(ws, rr, 4, round(end_cr(code)), font=F_DEF, align=A_RIGHT, num_fmt=NUM)
        ko_acc_rows.append(rr)
        row += 1

    _set_cell(ws, ko_row, 4, "=SUM(" + ",".join(f"D{r}" for r in ko_acc_rows) + ")", font=F_DEF, align=A_RIGHT, num_fmt=NUM)

    _set_cell(ws, chok_row, 4, f"=D{oa_row}-D{ko_row}", font=F_HDR, align=A_RIGHT, num_fmt=NUM)
    _set_cell(ws, chok_row, 6, f"=IF(D{chok_row}<0,\"{STOP}\",\"OK\")", font=F_HDR, align=A_LEFT)
    row += 2

    _set_cell(ws, row, 2, "P&L (\u043e\u0431\u0449 / M)", font=F_HDR, align=A_LEFT)
    row += 1

    rev_obsh_row = row
    _set_cell(ws, row, 2, "\u0412\u044b\u0440\u0443\u0447\u043a\u0430 (\u0438\u0437 \u043e\u0431\u0449)", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 3, "6010", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 4, round(rev_6010), font=F_DEF, align=A_RIGHT, num_fmt=NUM)
    row += 1

    rev_base_row = row
    _set_cell(ws, row, 2, f"\u0412\u044b\u0440\u0443\u0447\u043a\u0430 {base_year} (M)", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 3, f"M (G, \u0441\u0443\u043c\u043c\u0430 {base_year})", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 4, round(rev_base), font=F_DEF, align=A_RIGHT, num_fmt=NUM)
    row += 1

    rev_prev_row = row
    _set_cell(ws, row, 2, f"\u0412\u044b\u0440\u0443\u0447\u043a\u0430 {prev_year} (M)", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 3, f"M (G, \u0441\u0443\u043c\u043c\u0430 {prev_year})", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 4, round(rev_prev), font=F_DEF, align=A_RIGHT, num_fmt=NUM)
    row += 1

    yoy_row = row
    _set_cell(ws, row, 2, "\u0414\u0438\u043d\u0430\u043c\u0438\u043a\u0430 YoY", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 3, f"{base_year} vs {prev_year}", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 5, f"=IF(D{rev_prev_row}=0,\"N/A\",D{rev_base_row}/D{rev_prev_row}-1)", font=F_HDR, align=A_RIGHT, num_fmt=PCT)
    _set_cell(ws, row, 6, f"=IF(E{yoy_row}=\"N/A\",\"\",IF(E{yoy_row}<-0.3,\"{STOP}\",IF(E{yoy_row}<-0.1,\"{RED}\",IF(E{yoy_row}<0,\"{YEL}\",\"OK\"))))", font=F_DEF, align=A_LEFT)
    row += 2

    op_margin_row = row
    _set_cell(ws, row, 2, "\u041e\u043f\u0435\u0440\u0430\u0446\u0438\u043e\u043d\u043d\u0430\u044f \u0440\u0435\u043d\u0442\u0430\u0431\u0435\u043b\u044c\u043d\u043e\u0441\u0442\u044c", font=F_HDR, align=A_LEFT)
    _set_cell(ws, row, 3, "(6010\u22127010\u22127110\u22127210)/6010", font=F_DEF, align=A_LEFT)
    op_profit_row = row + 1
    _set_cell(ws, row, 5, f"=IF(D{rev_obsh_row}=0,\"N/A\",D{op_profit_row}/D{rev_obsh_row})", font=F_HDR, align=A_RIGHT, num_fmt=PCT)
    _set_cell(ws, row, 6, f"=IF(E{op_margin_row}=\"N/A\",\"\",IF(E{op_margin_row}<=0.02,\"{RED}\",\"OK\"))", font=F_DEF, align=A_LEFT)
    row += 1
    _set_cell(ws, row, 2, "  \u041e\u043f\u0435\u0440\u0430\u0446\u0438\u043e\u043d\u043d\u0430\u044f \u043f\u0440\u0438\u0431\u044b\u043b\u044c", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 4, round(op_profit), font=F_DEF, align=A_RIGHT, num_fmt=NUM)
    row += 2

    cfo_ratio_row = row
    _set_cell(ws, row, 2, f"CFO / \u0412\u044b\u0440\u0443\u0447\u043a\u0430 (\u0437\u0430 {period_short})", font=F_HDR, align=A_LEFT)
    _set_cell(ws, row, 3, "W (CFO) / M (\u0412\u044b\u0440\u0443\u0447\u043a\u0430)", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 5, f"=IF(D{row+2}=0,\"N/A\",D{row+1}/D{row+2})", font=F_HDR, align=A_RIGHT, num_fmt=PCT)
    _set_cell(ws, row, 6, f"=IF(E{cfo_ratio_row}=\"N/A\",\"\",IF(E{cfo_ratio_row}>0.5,\"{POS}\",IF(E{cfo_ratio_row}>=0.2,\"OK\",IF(E{cfo_ratio_row}<0,\"{RED}\",\"{YEL}\"))))", font=F_DEF, align=A_LEFT)
    row += 1
    _set_cell(ws, row, 2, "  CFO (\u0442\u044b\u0441. \u0442\u0435\u043d\u0433\u0435)", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 4, round(cfo_period_ths), font=F_DEF, align=A_RIGHT, num_fmt=NUM)
    row += 1
    _set_cell(ws, row, 2, f"  \u0412\u044b\u0440\u0443\u0447\u043a\u0430 {period_short} (\u0442\u044b\u0441. \u0442\u0435\u043d\u0433\u0435)", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 4, round(revenue_period_ths), font=F_DEF, align=A_RIGHT, num_fmt=NUM)
    row += 2

    _write_row(ws, row, ["\u0420\u0410\u0421\u0427\u0401\u0422 \u041b\u0418\u041c\u0418\u0422\u0410"], font=F_HDR, fill=FILL_RISK, align=A_LEFT)
    row += 1

    lim5_row = row
    _set_cell(ws, row, 2, "  5% \u00d7 \u0412\u044b\u0440\u0443\u0447\u043a\u0430", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 4, f"=0.05*D{rev_rat_row}", font=F_DEF, align=A_RIGHT, num_fmt=NUM)
    row += 1

    lim10_row = row
    _set_cell(ws, row, 2, "  10% \u00d7 \u0412\u0430\u043b\u044e\u0442\u0430 \u0431\u0430\u043b\u0430\u043d\u0441\u0430", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 4, f"=0.10*D{bal_row}", font=F_DEF, align=A_RIGHT, num_fmt=NUM)
    row += 1

    _set_cell(ws, row, 2, "  \u041a\u0430\u043f\u0438\u0442\u0430\u043b", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 4, f"=D{equity_row}", font=F_DEF, align=A_RIGHT, num_fmt=NUM)
    row += 1

    min_row = row
    _set_cell(ws, row, 2, "  min(\u043b\u0438\u043c\u0438\u0442\u044b)", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 4, f"=MIN(D{lim5_row},D{lim10_row},D{equity_row})", font=F_HDR, align=A_RIGHT, num_fmt=NUM)
    row += 1

    minus3010_row = row
    _set_cell(ws, row, 2, "  (\u2212) \u041a\u0440\u0430\u0442\u043a\u043e\u0441\u0440\u043e\u0447\u043d\u044b\u0439 \u0434\u043e\u043b\u0433 (3010)", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 4, f"=-D{r3010}", font=F_HDR, align=A_RIGHT, num_fmt=NUM)
    row += 1

    lease_row = row
    _set_cell(ws, row, 2, "  (\u2212) \u041b\u0438\u0437\u0438\u043d\u0433 (4150/7)", font=F_DEF, align=A_LEFT)
    _set_cell(ws, row, 4, round(end_cr("4150") / 7.0) if end_cr("4150") else 0, font=F_DEF, align=A_RIGHT, num_fmt=NUM)
    row += 2

    limit_total_row = row
    _set_cell(ws, row, 2, "  \u041b\u0418\u041c\u0418\u0422 \u0418\u0422\u041e\u0413\u041e", font=Font(name="Calibri", size=12, bold=True), align=A_LEFT)
    _set_cell(ws, row, 4, f"=D{min_row}+D{minus3010_row}-D{lease_row}", font=Font(name="Calibri", size=12, bold=True), align=A_RIGHT, num_fmt=NUM)
    _set_cell(ws, row, 6, f"=IF(D{limit_total_row}<0,\"{STOP}\",\"\")", font=Font(name="Calibri", size=12, bold=True), align=A_LEFT)
    row += 2
    _write_row(ws, row, ["\u0421\u0412\u041e\u0414\u041a\u0410 \u0424\u041b\u0410\u0413\u041e\u0412", "", "\u041f\u0440\u043e\u0432\u0435\u0440\u043a\u0430", "\u0417\u043d\u0430\u0447\u0435\u043d\u0438\u0435", "\u0420\u0435\u0437\u0443\u043b\u044c\u0442\u0430\u0442"], font=F_HDR, fill=FILL_HDR, align=A_LEFT)
    row += 1

    summary_start = row

    def _sum_row(cat, check, value_formula, yes_formula, *, cat_color, fill_yes=None, num_fmt=None):
        nonlocal row
        _set_cell(ws, row, 2, cat, font=Font(name="Calibri", size=11, bold=True), align=A_LEFT)
        _set_cell(ws, row, 4, check, font=F_DEF, align=A_LEFT)
        _set_cell(ws, row, 5, value_formula, font=F_DEF, align=A_RIGHT, num_fmt=num_fmt)
        _set_cell(ws, row, 6, f"=IF({yes_formula},\"\u0414\u0410\",\"\u041d\u0415\u0422\")", font=F_HDR, align=A_CENTER)
        if fill_yes is not None:
            ws.cell(row=row, column=6).fill = fill_yes
        row += 1

    _sum_row("\u0421\u0422\u041e\u041f", "\u041e\u0442\u0440\u0438\u0446\u0430\u0442\u0435\u043b\u044c\u043d\u044b\u0439 \u043a\u0430\u043f\u0438\u0442\u0430\u043b", f"=D{equity_row}", f"D{equity_row}<0", cat_color=C_RED, num_fmt=NUM)
    _sum_row("\u0421\u0422\u041e\u041f", "\u041f\u0430\u0434\u0435\u043d\u0438\u0435 \u0432\u044b\u0440\u0443\u0447\u043a\u0438 >30%", f"=E{yoy_row}", f"AND(E{yoy_row}<>\"N/A\",E{yoy_row}<-0.3)", cat_color=C_RED, num_fmt=PCT)
    _sum_row("\u0421\u0422\u041e\u041f", "\u041e\u0442\u0440\u0438\u0446\u0430\u0442\u0435\u043b\u044c\u043d\u044b\u0439 \u043b\u0438\u043c\u0438\u0442", f"=D{limit_total_row}", f"D{limit_total_row}<0", cat_color=C_RED, num_fmt=NUM)

    _sum_row("\u041a\u0420\u0410\u0421\u041d\u042b\u0419", "\u041e\u043f. \u0440\u0435\u043d\u0442\u0430\u0431\u0435\u043b\u044c\u043d\u043e\u0441\u0442\u044c \u2264 2%", f"=E{op_margin_row}", f"AND(E{op_margin_row}<>\"N/A\",E{op_margin_row}<=0.02)", cat_color=C_RED, num_fmt=PCT)
    _sum_row("\u041a\u0420\u0410\u0421\u041d\u042b\u0419", "\u0414\u043e\u043b\u0433/\u041a\u0430\u043f\u0438\u0442\u0430\u043b >10x", f"=E{debt_eq_row}", f"AND(E{debt_eq_row}<>\"N/A\",E{debt_eq_row}>10)", cat_color=C_RED, num_fmt=MULT)

    _sum_row("\u0416\u0401\u041b\u0422\u042b\u0419", "\u041a\u0440\u0430\u0442\u043a. \u0434\u043e\u043b\u0433/\u041a\u0430\u043f\u0438\u0442\u0430\u043b >10", f"=E{sh_debt_row}", f"AND(E{sh_debt_row}<>\"N/A\",E{sh_debt_row}>10)", cat_color=C_ORANGE, num_fmt=MULT)
    _sum_row("\u0416\u0401\u041b\u0422\u042b\u0419", "\u0414\u0417 / \u0412\u044b\u0440\u0443\u0447\u043a\u0430 >30%", f"=E{dz_row}", f"AND(E{dz_row}<>\"N/A\",E{dz_row}>0.3)", cat_color=C_ORANGE, num_fmt=PCT)
    _sum_row("\u0416\u0401\u041b\u0422\u042b\u0419", "\u0410\u0432\u0430\u043d\u0441\u044b \u0432\u044b\u0434\u0430\u043d\u043d\u044b\u0435 >30% \u0432\u044b\u0440\u0443\u0447\u043a\u0438", f"=E{advg_row}", f"AND(E{advg_row}<>\"N/A\",E{advg_row}>0.3)", cat_color=C_ORANGE, num_fmt=PCT)
    _sum_row("\u0416\u0401\u041b\u0422\u042b\u0419", "\u041e\u0442\u0440\u0438\u0446\u0430\u0442\u0435\u043b\u044c\u043d\u044b\u0439 \u0427\u041e\u041a", f"=D{chok_row}", f"D{chok_row}<0", cat_color=C_ORANGE, num_fmt=NUM)

    _sum_row("\u041f\u041e\u0417\u0418\u0422\u0418\u0412", "CFO / \u0412\u044b\u0440\u0443\u0447\u043a\u0430 > 50%", f"=E{cfo_ratio_row}", f"AND(E{cfo_ratio_row}<>\"N/A\",E{cfo_ratio_row}>0.5)", cat_color=C_GREEN, num_fmt=PCT)

    summary_end = row - 1

    row += 1

    stop_cnt_row = row
    _set_cell(ws, row, 2, "\u0421\u0442\u043e\u043f-\u0444\u0430\u043a\u0442\u043e\u0440\u043e\u0432:", font=F_HDR, align=A_LEFT)
    _set_cell(ws, row, 4, f"=COUNTIFS(B{summary_start}:B{summary_end},\"\u0421\u0422\u041e\u041f\",F{summary_start}:F{summary_end},\"\u0414\u0410\")", font=F_HDR, align=A_RIGHT)
    row += 1

    red_cnt_row = row
    _set_cell(ws, row, 2, "\u041a\u0440\u0430\u0441\u043d\u044b\u0445 \u0444\u043b\u0430\u0433\u043e\u0432:", font=F_HDR, align=A_LEFT)
    _set_cell(ws, row, 4, f"=COUNTIFS(B{summary_start}:B{summary_end},\"\u041a\u0420\u0410\u0421\u041d\u042b\u0419\",F{summary_start}:F{summary_end},\"\u0414\u0410\")", font=F_HDR, align=A_RIGHT)
    row += 1

    yellow_cnt_row = row
    _set_cell(ws, row, 2, "\u0416\u0451\u043b\u0442\u044b\u0445 \u0444\u043b\u0430\u0433\u043e\u0432:", font=F_HDR, align=A_LEFT)
    _set_cell(ws, row, 4, f"=COUNTIFS(B{summary_start}:B{summary_end},\"\u0416\u0401\u041b\u0422\u042b\u0419\",F{summary_start}:F{summary_end},\"\u0414\u0410\")", font=F_HDR, align=A_RIGHT)
    row += 1

    x_ref = f"D{stop_cnt_row}"
    y_ref = f"D{red_cnt_row}"
    z_ref = f"D{yellow_cnt_row}"
    _set_cell(ws, row, 2,
        f"=IF({x_ref}>0,\"\u041f\u043e \u0447\u0435\u043a-\u043b\u0438\u0441\u0442\u0443: \u041d\u0415 \u043f\u0440\u043e\u0445\u043e\u0434\u0438\u0442 \u043f\u0440\u0435\u0441\u043a\u043e\u0440\u0438\u043d\u0433 (\u0435\u0441\u0442\u044c \u0441\u0442\u043e\u043f-\u0444\u0430\u043a\u0442\u043e\u0440\u044b)\",IF({y_ref}>2,\"\u041f\u043e \u0447\u0435\u043a-\u043b\u0438\u0441\u0442\u0443: \u041d\u0415 \u043f\u0440\u043e\u0445\u043e\u0434\u0438\u0442 \u043f\u0440\u0435\u0441\u043a\u043e\u0440\u0438\u043d\u0433 (\u0431\u043e\u043b\u0435\u0435 2 \u043a\u0440\u0430\u0441\u043d\u044b\u0445 \u0444\u043b\u0430\u0433\u043e\u0432)\",IF({z_ref}>3,\"\u041f\u043e \u0447\u0435\u043a-\u043b\u0438\u0441\u0442\u0443: \u041d\u0415 \u043f\u0440\u043e\u0445\u043e\u0434\u0438\u0442 \u043f\u0440\u0435\u0441\u043a\u043e\u0440\u0438\u043d\u0433 (\u0431\u043e\u043b\u0435\u0435 3 \u0436\u0451\u043b\u0442\u044b\u0445 \u0444\u043b\u0430\u0433\u043e\u0432)\",\"\u041f\u043e \u0447\u0435\u043a-\u043b\u0438\u0441\u0442\u0443: \u041f\u0420\u041e\u0425\u041e\u0414\u0418\u0422 \u043f\u0440\u0435\u0441\u043a\u043e\u0440\u0438\u043d\u0433\")))",
        font=F_HDR, align=A_LEFT)
    row += 1

    return row




def generate_insights(file_bytes: bytes) -> bytes:
    """Создаёт лист "инсайты". Если уже есть — не пересоздаёт."""
    wb = _load_wb(file_bytes)

    required = {"W", "M", "Wt", "Mt"}
    if not required.issubset(set(wb.sheetnames)):
        return file_bytes

    if "инсайты" in wb.sheetnames:
        return _save_wb(wb)

    ws = wb.create_sheet("инсайты", 0)

    # widths
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 50
    for col in range(3, 15):
        ws.column_dimensions[get_column_letter(col)].width = 16

    # title
    _write_row(ws, 1, ["ИНСАЙТЫ"], font=F_TITLE, align=A_LEFT)

    row = 3
    ws_wt = wb["Wt"]
    ws_mt = wb["Mt"]
    ws_m = wb["M"]
    ws_w = wb["W"]
    ws_cred = wb["кред"] if "кред" in wb.sheetnames else None

    row = _write_block1(ws, row, ws_wt=ws_wt, ws_cred=ws_cred)
    row = _write_block2(ws, row, ws_m=ws_m)
    row = _write_block3(ws, row, ws_w=ws_w)
    row = _write_block4(ws, row, ws_mt=ws_mt, ws_wt=ws_wt)
    row = _write_block5(ws, row, ws_wt=ws_wt)

    ws_obsh = wb["общ"] if "общ" in wb.sheetnames else None
    if ws_obsh is None:
        _write_row(ws, row, ["Блок 6 недоступен: лист «общ» не найден"] , font=F_HDR, align=A_LEFT)
        row += 2
    else:
        row = _write_block6(ws, row, ws_obsh=ws_obsh, ws_m=ws_m, ws_w=ws_w)

    return _save_wb(wb)
