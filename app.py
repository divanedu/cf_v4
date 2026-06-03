import re
import io
import time
import os
import tempfile
import zipfile
from copy import copy as _copy
from collections import defaultdict
from decimal import Decimal, ROUND_HALF_UP
from typing import Any, Callable, Dict, Iterable, List, Optional, Set, Tuple

import requests as _requests_lib

import pandas as pd
import streamlit as st

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string


# =========================
# Вспомогательные функции
# =========================
def safe_sheet_name(name: str) -> str:
    """Правила Excel для имени листа: максимум 31 символ, нельзя : \\ / ? * [ ]"""
    banned = [":", "\\", "/", "?", "*", "[", "]"]
    for ch in banned:
        name = name.replace(ch, "")
    name = (name or "").strip() or "лист"
    return name[:31]

def has_vba_project(file_bytes: bytes) -> bool:
    """True если в файле есть проект VBA (xlsm). Используем для правильного расширения выходного файла."""
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
            return "xl/vbaProject.bin" in set(zf.namelist())
    except Exception:
        return False


def safe_filename(name: str) -> str:
    banned = ["<", ">", ":", "\"", "/", "\\", "|", "?", "*"]
    for ch in banned:
        name = (name or "").replace(ch, "")
    name = (name or "").strip() or "output.xlsx"
    return name


def split_prefix_suffix4(sheet_name: str) -> Tuple[str, str]:
    """(префикс, последние 4 символа)"""
    if len(sheet_name) < 4:
        return sheet_name, ""
    return sheet_name[:-4], sheet_name[-4:].lower()


def split_prefix_suffix2(sheet_name: str) -> Tuple[str, str]:
    """(префикс, последние 2 символа)"""
    if len(sheet_name) < 2:
        return sheet_name, ""
    return sheet_name[:-2], sheet_name[-2:].lower()


def normalize_prefix(prefix: str) -> str:
    return (prefix or "").strip()


def make_unique_sheet_title(wb, desired_title: str) -> str:
    base = safe_sheet_name(desired_title)
    if base not in wb.sheetnames:
        return base
    i = 1
    while True:
        suffix = f" ({i})"
        trimmed = base
        if len(trimmed) + len(suffix) > 31:
            trimmed = trimmed[: 31 - len(suffix)]
        candidate = f"{trimmed}{suffix}"
        if candidate not in wb.sheetnames:
            return candidate
        i += 1


def make_unique_with_fixed_suffix(wb, prefix: str, suffix: str) -> str:
    """
    Возвращает уникальное имя листа, которое ВСЕГДА заканчивается на `suffix` (например, '1210' или 'Wd').
    Если `prefix+suffix` уже существует, вставляет счётчик перед suffix: f'{prefix}{i}{suffix}'.
    """
    prefix = (prefix or "").strip()
    suffix = (suffix or "").strip()
    base = safe_sheet_name(prefix + suffix)
    if base.endswith(suffix) and base not in wb.sheetnames:
        return base

    i = 1
    while True:
        mid = str(i)
        # Следим за ограничениями Excel: общая длина <= 31 и окончание = суффикс (важно для логики поиска листов).
        max_prefix_len = 31 - len(mid) - len(suffix)
        p = prefix[:max_prefix_len] if max_prefix_len > 0 else ""
        candidate = safe_sheet_name(f"{p}{mid}{suffix}")
        if candidate.endswith(suffix) and candidate not in wb.sheetnames:
            return candidate
        i += 1


def load_wb_from_bytes(file_bytes: bytes, filename: str):
    ext = (filename or "").lower().rsplit(".", 1)[-1]
    keep_vba = ext == "xlsm"
    return load_workbook(io.BytesIO(file_bytes), keep_vba=keep_vba), keep_vba


def is_xls_filename(filename: str) -> bool:
    return (filename or "").lower().endswith(".xls") and not (filename or "").lower().endswith(".xlsx")


def convert_xls_to_xlsx_via_excel(xls_bytes: bytes, original_name: str) -> Tuple[str, bytes]:
    """
    Конвертирует байты .xls в байты .xlsx с помощью установленного Microsoft Excel (COM).
    Работает только на Windows, где установлен Excel и доступен pywin32.
    """
    if os.name != "nt":
        raise RuntimeError("Конвертация .xls поддерживается только на Windows с установленным Excel.")

    try:
        import pythoncom  # type: ignore
        import win32com.client  # type: ignore
    except Exception:
        raise RuntimeError("Для .xls нужно установить pywin32: pip install pywin32")

    pythoncom.CoInitialize()
    excel = None
    xls_path = None
    xlsx_path = None
    try:
        fd_xls, xls_path = tempfile.mkstemp(suffix=".xls")
        os.close(fd_xls)
        with open(xls_path, "wb") as f:
            f.write(xls_bytes)

        fd_xlsx, xlsx_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd_xlsx)

        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(xls_path, ReadOnly=True)
        try:
            # 51 = xlOpenXMLWorkbook (.xlsx) — формат сохранения для Excel COM
            wb.SaveAs(xlsx_path, FileFormat=51)
        finally:
            wb.Close(SaveChanges=False)
        excel.Quit()
        excel = None

        with open(xlsx_path, "rb") as f:
            out_bytes = f.read()

        base = (original_name or "input.xls")
        base = base[:-4] if base.lower().endswith(".xls") else base
        out_name = base + ".xlsx"
        return out_name, out_bytes
    finally:
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass
        for p in (xls_path, xlsx_path):
            if p and os.path.exists(p):
                try:
                    os.remove(p)
                except Exception:
                    pass


def ensure_openpyxl_bytes(filename: str, file_bytes: bytes, cache: Optional[Dict[Tuple[str, int], Tuple[str, bytes]]] = None) -> Tuple[str, bytes]:
    """
    Гарантирует, что на выходе будут байты, которые читает openpyxl (xlsx/xlsm).
    Если загружен .xls — конвертируем в .xlsx (локально на Windows через Excel COM).
    `cache` ключуется по (filename, len(bytes)), чтобы не конвертировать один и тот же файл повторно.
    """
    if not is_xls_filename(filename):
        return filename, file_bytes
    key = (filename, len(file_bytes))
    if cache is not None and key in cache:
        return cache[key]
    out = convert_xls_to_xlsx_via_excel(file_bytes, filename)
    if cache is not None:
        cache[key] = out
    return out


def copy_sheet(src_ws, dst_wb, new_title: str):
    dst_ws = dst_wb.create_sheet(title=new_title)
    dst_ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines
    dst_ws.freeze_panes = src_ws.freeze_panes

    for key, dim in src_ws.column_dimensions.items():
        dst = dst_ws.column_dimensions[key]
        dst.width = dim.width
        dst.hidden = dim.hidden
        dst.outlineLevel = dim.outlineLevel

    for idx, dim in src_ws.row_dimensions.items():
        dst = dst_ws.row_dimensions[idx]
        dst.height = dim.height
        dst.hidden = dim.hidden
        dst.outlineLevel = dim.outlineLevel

    for merged in list(src_ws.merged_cells.ranges):
        dst_ws.merge_cells(str(merged))

    # Быстрый путь: копируем только реально существующие ячейки (значения/стили),
    # чтобы не сканировать огромные "пустые" прямоугольники.
    for (_r, _c), cell in getattr(src_ws, "_cells", {}).items():
        if isinstance(cell, MergedCell):
            continue
        col_idx = cell.column if isinstance(cell.column, int) else column_index_from_string(str(cell.column))
        dst_cell = dst_ws.cell(row=cell.row, column=col_idx, value=cell.value)
        if cell.has_style:
            dst_cell.font = _copy(cell.font)
            dst_cell.border = _copy(cell.border)
            dst_cell.fill = _copy(cell.fill)
            dst_cell.number_format = cell.number_format
            dst_cell.protection = _copy(cell.protection)
            dst_cell.alignment = _copy(cell.alignment)
    return dst_ws


def excel_col_width_from_pixels(px: float) -> float:
    # Приблизительный перевод пикселей в "ширину колонки" Excel (по умолчанию).
    # Часто используют оценку: пиксели ~= width*7 + 5  => width ~= (px-5)/7
    try:
        px = float(px)
    except Exception:
        px = 48.0
    return max(0.0, round((px - 5.0) / 7.0, 2))


def set_all_columns_width(ws, px: float):
    width = excel_col_width_from_pixels(px)
    max_col = ws.max_column or 1
    # Ограничение на случай, если max_column "раздулся" из-за форматирования.
    max_col = min(max_col, 2000)
    for i in range(1, max_col + 1):
        col = get_column_letter(i)
        dim = ws.column_dimensions[col]
        dim.width = width


REGISTRY_COL_WIDTH_PX = 48.0
# Высота строк в Excel задаётся в пунктах (points). Требование: фиксированная высота (без авто-роста).
REGISTRY_ROW_HEIGHT_PT = 28.8


def format_registry_sheet(ws) -> None:
    """
    Реестр (Wr/Mr): фиксируем ширину всех колонок и высоту всех строк,
    чтобы Excel не "раздувал" строки из-за wrap_text.
    """
    set_all_columns_width(ws, REGISTRY_COL_WIDTH_PX)
    try:
        ws.sheet_format.defaultRowHeight = REGISTRY_ROW_HEIGHT_PT
    except Exception:
        pass

    rows: Set[int] = set(getattr(ws, "row_dimensions", {}).keys())
    for (r, _c), cell in getattr(ws, "_cells", {}).items():
        rows.add(int(r))
        # Принудительно отключаем перенос текста, иначе Excel может визуально увеличивать высоту строки.
        try:
            al = cell.alignment
            if al and getattr(al, "wrap_text", None):
                cell.alignment = Alignment(
                    horizontal=al.horizontal,
                    vertical=al.vertical,
                    textRotation=al.textRotation,
                    wrapText=False,
                    shrinkToFit=al.shrinkToFit,
                    indent=al.indent,
                    relativeIndent=al.relativeIndent,
                    justifyLastLine=al.justifyLastLine,
                    readingOrder=al.readingOrder,
                )
        except Exception:
            pass

    # Ставим высоту только тем строкам, которые реально существуют (в row_dimensions или есть ячейки),
    # чтобы не гонять цикл 1..max_row на больших листах.
    for r in rows:
        try:
            ws.row_dimensions[r].height = REGISTRY_ROW_HEIGHT_PT
        except Exception:
            pass


def merge_wh_m_into_analysis(analysis_wb, wh_bytes: Optional[bytes], wh_name: str, m_bytes: Optional[bytes], m_name: str) -> Dict[str, List[str]]:
    report = {"missing_wh": [], "missing_m": [], "copied": []}

    if m_bytes:
        m_wb, _ = load_wb_from_bytes(m_bytes, m_name)
        mapping_m = {
            "Итог": "M",
            "Реестр": "Mr",
            "Контрагенты": "Mt",
            "Договоры": "Md",
        }
        for src_title, dst_title in mapping_m.items():
            if src_title not in m_wb.sheetnames:
                report["missing_m"].append(src_title)
                continue
            new_title = make_unique_sheet_title(analysis_wb, dst_title)
            dst_ws = copy_sheet(m_wb[src_title], analysis_wb, new_title)
            if dst_title.lower() == "mr":
                format_registry_sheet(dst_ws)
            report["copied"].append(f"{m_name}:{src_title} -> {new_title}")

    if wh_bytes:
        wh_wb, _ = load_wb_from_bytes(wh_bytes, wh_name)
        mapping_wh = {
            "Итог": "W",
            "Реестр": "Wr",
            "Таблицы": "Wt",
            "Договоры": "Wd",
        }
        for src_title, dst_title in mapping_wh.items():
            if src_title not in wh_wb.sheetnames:
                report["missing_wh"].append(src_title)
                continue
            new_title = make_unique_sheet_title(analysis_wb, dst_title)
            dst_ws = copy_sheet(wh_wb[src_title], analysis_wb, new_title)
            if dst_title.lower() == "wr":
                format_registry_sheet(dst_ws)
            report["copied"].append(f"{wh_name}:{src_title} -> {new_title}")

        if "Кредиты" in wh_wb.sheetnames:
            ws_cred = wh_wb["Кредиты"]
            if ws_cred["AC6"].value not in (None, 0, 0.0, "0"):
                new_title = make_unique_sheet_title(analysis_wb, "кред")
                copy_sheet(ws_cred, analysis_wb, new_title)
                report["copied"].append(f"{wh_name}:Кредиты -> {new_title}")

    return report


def merge_wh_m_into_analysis_with_prefix(
    analysis_wb,
    wh_bytes: Optional[bytes],
    wh_name: str,
    m_bytes: Optional[bytes],
    m_name: str,
    prefix: str,
) -> Dict[str, List[str]]:
    """
    То же, что merge_wh_m_into_analysis, но принудительно добавляет префикс к импортируемым листам.
    Важно: сохраняем ожидаемые суффиксы (Wd/Md и т.п.), чтобы блок "Контракты" корректно находил пары листов.
    """
    report = {"missing_wh": [], "missing_m": [], "copied": []}
    p = (prefix or "").strip()

    if m_bytes:
        m_wb, _ = load_wb_from_bytes(m_bytes, m_name)
        mapping_m = {"Итог": "M", "Контрагенты": "Mt", "Реестр": "Mr", "Договоры": "Md"}
        for src_title, base in mapping_m.items():
            if src_title not in m_wb.sheetnames:
                report["missing_m"].append(src_title)
                continue
            new_title = make_unique_sheet_title(analysis_wb, f"{p}{base}")
            # Суффикс Md должен остаться Md, иначе не соберём пары Wd/Md при создании "контр".
            if base.lower().endswith("md"):
                new_title = make_unique_with_fixed_suffix(analysis_wb, p + base[:-2], "Md")
            dst_ws = copy_sheet(m_wb[src_title], analysis_wb, new_title)
            if base.lower() == "mr":
                format_registry_sheet(dst_ws)
            report["copied"].append(f"{m_name}:{src_title} -> {new_title}")

    if wh_bytes:
        wh_wb, _ = load_wb_from_bytes(wh_bytes, wh_name)
        mapping_wh = {"Итог": "W", "Таблицы": "Wt", "Реестр": "Wr", "Договоры": "Wd"}
        for src_title, base in mapping_wh.items():
            if src_title not in wh_wb.sheetnames:
                report["missing_wh"].append(src_title)
                continue
            new_title = make_unique_sheet_title(analysis_wb, f"{p}{base}")
            if base.lower().endswith("wd"):
                new_title = make_unique_with_fixed_suffix(analysis_wb, p + base[:-2], "Wd")
            dst_ws = copy_sheet(wh_wb[src_title], analysis_wb, new_title)
            if base.lower() == "wr":
                format_registry_sheet(dst_ws)
            report["copied"].append(f"{wh_name}:{src_title} -> {new_title}")

        if "Кредиты" in wh_wb.sheetnames:
            ws_cred = wh_wb["Кредиты"]
            if ws_cred["AC6"].value not in (None, 0, 0.0, "0"):
                new_title = make_unique_sheet_title(analysis_wb, f"{p}кред")
                copy_sheet(ws_cred, analysis_wb, new_title)
                report["copied"].append(f"{wh_name}:Кредиты -> {new_title}")

    return report


# =========================
# Очистка ОСВ (автоматически для ОСВ с "рандомным" именем файла)
# =========================
MAX_CELLS_PER_SHEET = 300_000
SORT_ACCOUNTS = {"1310", "1320", "1330"}  # только эти счета сортируем по колонке G
OSV_BAD_WORDS = [
    "Договор", "Догов", "Д/р о государственных закупках",
    "KZT", "RUB", "EUR", "USD", "Жетысу",
    "Головное подразделение",
    "Основной склад", "Основное подразделение", "Склад",
    "Обороты за",
    "Соглашение от", "Соглашение об", ",,,.", 
    "Б/н", " от ",
    "01.", "02.", "03.", "04.", "05.", "06.",
    "07.", "08.", "09.", "10.", "11.", "12.", "<...>", "<..>", "<.>",
    "основной склад",
    "вспомогательный склад",
    "резервный склад",
    "торговый зал",
    "подразделение",
    "филиал",
    "представительство",
    "департамент",
    "управление",
    "отдел",
    "сектор",
    "группа",
    "бюро",
    "служба",
    "цех",
    "участок",
    "бригада",
    "лаборатория",
    "архив",
    "канцелярия",
    "бухгалтерия",
    "кадры",
    "юрист",
    "охрана",
    "администрация",
    "руководство",
    "дирекция",
    "секретариат",
    "хозчасть",
    "хозяйственная часть",
    "административно-хозяйственный",
    "ахч",
    "хозу",
    "хоз отдел",
    "хозяйственный отдел",
    "материальный склад",
    "продуктовый склад",
    "товарный склад",
    "оптовый склад",
    "розничный склад",
    "центральный склад",
    "региональный склад",
    "распределительный центр",
    "логистический центр",
    "транспортный отдел",
    "экспедиция",
    "доставка",
    "перемещение",
    "инвентаризация",
    "пересчет",
    "списание",
    "оприходование",
    "поступление",
    "реализация",
    "возврат",
    "брак",
    "пересортица",
    "недостача",
    "излишек",
    "корректировка",
    "переоценка",
    "уценка",
    "наценка",
    "маркировка",
    "перемаркировка",
    "упаковка",
    "переупаковка",
    "фасовка",
    "перефасовка",
    "комплектация",
    "раскомплектация",
    "сборка",
    "разборка",
    "ремонт",
    "обслуживание",
    "техническое обслуживание",
    "то-1",
    "то-2",
    "сезонное обслуживание",
    "консервация",
    "расконсервация",
    "монтаж",
    "демонтаж",
    "наладка",
    "регулировка",
    "калибровка",
    "поверка",
    "испытание",
    "проверка",
    "контроль",
    "аудит",
    "ревизия",
    "мониторинг",
    "наблюдение",
    "анализ",
    "исследование",
    "экспертиза",
    "оценка",
    "сертификация",
    "лицензирование",
    "аккредитация",
    "аттестация",
    "рационализация",
    "модернизация",
    "реконструкция",
    "строительство",
    "ремонт",
    "ремонтно-строительные",
    "отделочные работы",
    "сантехнические работы",
    "электромонтажные работы",
    "сварочные работы",
    "погрузочно-разгрузочные",
    "такелажные работы",
    "уборочные работы", "Основной",
    "клининговые услуги",
    "дезинфекция",
    "дератизация",
    "дезинсекция",
    "вывоз мусора",
    "утилизация",
    "переработка",
    "хранение",
    "ответственное хранение",
    "аренда",
    "субаренда",
    "лизинг",
    "прокат",
    "наем",
    "поднаем",
    "пользование",
    "эксплуатация",
    "содержание",
    "обслуживание зданий",
    "обслуживание сооружений",
    "обслуживание оборудования",
    "обслуживание техники",
    "обслуживание транспорта",
    "обслуживание инвентаря",
    "обслуживание инструмента",
    "обслуживание оснастки",
    "обслуживание приспособлений",
    "смазка",
    "заправка",
    "зарядка",
    "подзарядка",
    "замена",
    "замена масла",
    "замена фильтров",
    "замена расходников",
    "замена запчастей",
    "замена комплектующих",
    "ремонт оборудования",
    "ремонт техники",
    "ремонт транспорта",
    "ремонт инвентаря",
    "ремонт инструмента",
    "ремонт оснастки",
    "ремонт приспособлений",
    "текущий ремонт",
    "капитальный ремонт",
    "плановый ремонт",
    "аварийный ремонт",
    "восстановительный ремонт",
    "профилактический ремонт",
]


def find_first_row_with_value(ws, value, col: int = 1) -> Optional[int]:
    target_raw = str(value).strip()
    target = target_raw.lower()
    target_digits = target_raw.isdigit()
    for r in range(1, (ws.max_row or 1) + 1):
        v = ws.cell(row=r, column=col).value
        if v is None:
            continue
        if target_digits and isinstance(v, (int, float)) and float(v).is_integer():
            if int(v) == int(target_raw):
                return r
            continue
        s = str(v).replace(" ", " ").replace(" ", " ").strip().lower()
        if s == target:
            return r
    return None


def find_first_row_contains(ws, substring: str, col: int = 1) -> Optional[int]:
    sub = (substring or "").lower()
    for r in range(1, (ws.max_row or 1) + 1):
        v = ws.cell(row=r, column=col).value
        if v and sub in str(v).lower():
            return r
    return None


def clear_outline_for_sheet(ws):
    last_row = max(ws.max_row or 1, 1)
    last_col = get_column_letter(max(ws.max_column or 1, 1))
    try:
        ws.ungroup_rows(1, last_row)
    except Exception:
        pass
    try:
        ws.ungroup_columns("A", last_col)
    except Exception:
        pass
    for r in range(1, last_row + 1):
        ws.row_dimensions[r].outlineLevel = 0
    for dim in ws.column_dimensions.values():
        dim.outlineLevel = 0


def remove_duplicate_rows(ws, start_row: int, end_row: int) -> bool:
    rows_data: Dict[Tuple[str, ...], List[int]] = {}
    rows_to_delete: Set[int] = set()
    for r in range(start_row, end_row + 1):
        if all((c.value is None or str(c.value).strip() == "") for c in ws[r]):
            continue
        row_values = tuple(str(cell.value).strip() if cell.value else "" for cell in ws[r])
        rows_data.setdefault(row_values, []).append(r)
    for rows in rows_data.values():
        if len(rows) >= 2:
            rows_to_delete.update(rows)
    if rows_to_delete:
        for r in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(r)
        return True
    return False


def get_account_number(ws) -> Optional[str]:
    """
    Extracts account number from the 2nd row.
    Some exports put the text in a merged cell that doesn't start at A2,
    so we scan multiple columns on row 2 and pick the first 4-digit match.
    """
    candidates: List[str] = []

    max_col = min(ws.max_column or 1, 60)
    for c in range(1, max_col + 1):
        v = ws.cell(row=2, column=c).value
        if v is None:
            continue
        if isinstance(v, (int, float)) and float(v).is_integer():
            candidates.append(str(int(v)))
        else:
            candidates.append(str(v))

    # Если 2-я строка часть объединённой шапки, а значение лежит в 1-й строке — берём левую-верхнюю ячейку merge-диапазона.
    for r in ws.merged_cells.ranges:
        min_col, min_row, max_col2, max_row2 = r.bounds
        if not (min_row <= 2 <= max_row2):
            continue
        tl = ws.cell(row=min_row, column=min_col).value
        if tl is None:
            continue
        candidates.append(str(tl))

    def _normalize(s: str) -> str:
        s = (s or "").replace(" ", " ").replace(" ", " ").strip()
        return s

    def _digits_compact(s: str) -> str:
        s = _normalize(s).replace(" ", "")
        return "".join(ch for ch in s if ch.isdigit())

    def _pick_from_text(s: str) -> Optional[str]:
        s_norm = _normalize(s)
        s_low = s_norm.lower()
        # Приоритет: 4 цифры сразу после слова "счет".
        idx = s_low.find("счет")
        if idx != -1:
            tail = s_norm[idx:]
            tail_compact = tail.replace(" ", "").replace(" ", "").replace(" ", "")
            m = re.search(r"(\\d{4})", tail_compact)
            if m:
                return m.group(1)
        # Фолбэк: любые 4 подряд идущие цифры (предварительно "сжимаем" пробелы между цифрами).
        compact = s_norm.replace(" ", "").replace(" ", "").replace(" ", "")
        m = re.search(r"(\\d{4})", compact)
        if m:
            return m.group(1)
        # Последний шанс: берём первые 4 цифры из строки, где оставили только цифры.
        d = _digits_compact(s_norm)
        if len(d) >= 4:
            return d[:4]
        return None

    for s in candidates:
        acc = _pick_from_text(s)
        if acc:
            return acc
    return None


def set_all_rows_height(ws, height: float = 12):
    for r in range(1, (ws.max_row or 1) + 1):
        ws.row_dimensions[r].height = height


_num_re = re.compile(r"[-+]?\\d+(?:[.,]\\d+)?")


def to_number(v) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    s = s.replace(" ", "").replace(" ", "").replace(" ", "")
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    m = _num_re.search(s)
    if not m:
        return None
    num = m.group(0).replace(",", ".")
    try:
        return float(num)
    except Exception:
        return None


def _cell_text(v) -> str:
    if v is None:
        return ""
    try:
        s = str(v)
    except Exception:
        return ""
    s = s.replace(" ", " ").replace(" ", " ").strip()
    return s


def extract_company_label_from_a1(ws) -> str:
    """
    For UX: show something human-readable per OSV sheet.
    Prefer A1; if empty, use first non-empty cell in row 1 (A..J).
    """
    s = _cell_text(ws.cell(row=1, column=1).value)
    if s:
        return s
    for c in range(1, 11):
        s = _cell_text(ws.cell(row=1, column=c).value)
        if s:
            return s
    return ""


def _short(s: str, n: int = 90) -> str:
    s = _cell_text(s)
    if len(s) <= n:
        return s
    return s[: n - 1].rstrip() + "…"


def sort_block_by_column(ws, start_row: int, end_row: int, col_index: int, descending: bool = True):
    if start_row >= end_row:
        return
    max_col = ws.max_column or 1
    rows = []
    for r in range(start_row, end_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        key = to_number(ws.cell(row=r, column=col_index).value)
        rows.append((key, row_vals))
    rows.sort(key=lambda x: (x[0] is not None, x[0] if x[0] is not None else 0.0), reverse=descending)
    for i, (_, row_vals) in enumerate(rows):
        rr = start_row + i
        for c in range(1, max_col + 1):
            ws.cell(row=rr, column=c).value = row_vals[c - 1]


def clean_osv_sheet_inplace(ws) -> Optional[str]:
    account_number = get_account_number(ws)
    if not account_number:
        set_all_rows_height(ws, 12)
        return None

    clear_outline_for_sheet(ws)
    if 5 in ws.row_dimensions:
        ws.row_dimensions[5].hidden = False

    total_cells = (ws.max_row or 1) * (ws.max_column or 1)
    if total_cells > MAX_CELLS_PER_SHEET:
        set_all_rows_height(ws, 12)
        return account_number

    last_col = ws.max_column or 1
    for mr in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mr))

    start_row = find_first_row_with_value(ws, account_number, col=1)
    end_row = find_first_row_with_value(ws, "Итого", col=1)
    if end_row is None:
        # Некоторые ОСВ выгрузки пишут "Итого:" или "Итого ..." в колонке A.
        for r in range(1, (ws.max_row or 1) + 1):
            v = ws.cell(row=r, column=1).value
            if v is None:
                continue
            s = str(v).replace(" ", " ").replace(" ", " ").strip().lower()
            if s.startswith("итого"):
                end_row = r
                break
    if not start_row or not end_row:
        set_all_rows_height(ws, 12)
        # Лист не выкидываем: номер счета мы всё равно знаем, и дальше он нужен в _Анализ.
        return account_number

    for r in range(ws.max_row or 1, start_row - 1, -1):
        v = ws.cell(row=r, column=1).value
        if v is None or str(v).strip() == "":
            ws.delete_rows(r)

    bad_words_lower = [bw.lower() for bw in OSV_BAD_WORDS]
    for r in range(ws.max_row or 1, start_row - 1, -1):
        v = ws.cell(row=r, column=1).value
        if v and any(bw in str(v).lower() for bw in bad_words_lower):
            ws.delete_rows(r)

    start_row = find_first_row_with_value(ws, account_number, col=1)
    end_row = find_first_row_with_value(ws, "Итого", col=1)
    if end_row is None:
        for r in range(1, (ws.max_row or 1) + 1):
            v = ws.cell(row=r, column=1).value
            if v is None:
                continue
            s = str(v).replace(" ", " ").replace(" ", " ").strip().lower()
            if s.startswith("итого"):
                end_row = r
                break
    if not start_row or not end_row:
        set_all_rows_height(ws, 12)
        return account_number

    if start_row + 1 <= end_row - 1:
        remove_duplicate_rows(ws, start_row + 1, end_row - 1)

    for r in range(start_row + 1, start_row + 2000):
        if r > (ws.max_row or 1):
            break
        cell = ws.cell(row=r, column=1)
        if cell.value is None:
            continue
        al = cell.alignment or Alignment()
        cell.alignment = Alignment(horizontal="left", vertical=al.vertical)

    for r in range(ws.max_row or 1, 0, -1):
        if all((c.value is None or str(c.value).strip() == "") for c in ws[r]):
            ws.delete_rows(r)

    ws.column_dimensions["A"].width = 50
    for col in ["B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col].width = 18

    if last_col == 8 and (ws.max_column or 1) >= 9:
        ws.delete_cols(9)
    elif last_col == 9 and (ws.max_column or 1) >= 8:
        ws.delete_cols(8)

    # =========================
    # Сортировка + контроль суммы после очистки (UX/качество данных)
    # =========================
    try:
        # Для счетов, у которых первые 2 цифры < 30: "Итого" берём из колонки G, иначе из H.
        two = int(str(account_number)[:2]) if str(account_number)[:2].isdigit() else 99
        total_col = 7 if two < 30 else 8  # G/H (после нормализации колонок)

        # Пересчитываем границы после удаления строк/колонок.
        start_row2 = find_first_row_with_value(ws, account_number, col=1)
        end_row2 = find_first_row_with_value(ws, "Итого", col=1)
        if end_row2 is None:
            for r in range(1, (ws.max_row or 1) + 1):
                v = ws.cell(row=r, column=1).value
                if v is None:
                    continue
                s = str(v).replace(" ", " ").replace(" ", " ").strip().lower()
                if s.startswith("итого"):
                    end_row2 = r
                    break
        if not start_row2 or not end_row2:
            start_row2, end_row2 = start_row, end_row

        # Формат чисел в блоке счета (включая строку счета и строку "Итого"):
        # от колонки C до "рабочей" колонки (G или H), только ПОСЛЕ очистки/нормализации.
        num_fmt_osv = "#,##0; -#,##0"
        for rr in range(int(start_row2), int(end_row2) + 1):
            for cc in range(3, int(total_col) + 1):  # C..G/H
                ws.cell(row=rr, column=cc).number_format = num_fmt_osv

        # Сортировка по убыванию по "рабочей" колонке (G для дебетовых <30, H для кредитовых >=30).
        data_start = start_row2 + 1
        data_end = end_row2 - 1
        if data_start <= data_end:
            sort_block_by_column(ws, data_start, data_end, col_index=total_col, descending=True)
            # После сортировки стили (в т.ч. красный шрифт) могут "остаться" на старых строках,
            # потому что мы переносим только значения. Сбрасываем цвет и заново красим отрицательные.
            for rr in range(data_start, data_end + 1):
                cell = ws.cell(row=rr, column=total_col)
                v = to_number(cell.value)
                base = cell.font or Font()
                if v is not None and float(v) < 0:
                    cell.font = Font(
                        name=base.name,
                        size=base.size,
                        bold=base.bold,
                        italic=base.italic,
                        underline=base.underline,
                        color="FFCC0000",
                    )
                else:
                    cell.font = Font(
                        name=base.name,
                        size=base.size,
                        bold=base.bold,
                        italic=base.italic,
                        underline=base.underline,
                        color=None,
                    )

        expected_total = to_number(ws.cell(row=start_row2, column=total_col).value) or 0.0

        actual_total = 0.0
        for rr in range(start_row2 + 1, end_row2):
            a = ws.cell(row=rr, column=1).value
            if a is not None and str(a).replace(" ", " ").replace(" ", " ").strip().lower().startswith("итого"):
                break
            v = to_number(ws.cell(row=rr, column=total_col).value)
            if v is not None:
                actual_total += float(v)

        if expected_total == 0.0:
            diff_pct = 0.0 if abs(actual_total) < 1e-9 else 1.0
        else:
            diff_pct = abs(actual_total - expected_total) / abs(expected_total)

        # % расхождения: пишем, но НЕ заливаем
        out_cell = ws.cell(row=1, column=total_col)
        out_cell.value = float(diff_pct)
        out_cell.number_format = "0.0%"
        out_cell.font = Font(name="Arial", size=9, bold=True, color="FF4B5563")
        out_cell.alignment = Alignment(horizontal="center", vertical="center")
        out_cell.fill = PatternFill(fill_type=None)

        # Статус: F1 если % в G1, либо G1 если % в H1
        status_col = 6 if total_col == 7 else 7  # F/G
        status_cell = ws.cell(row=1, column=status_col)
        status_cell.font = Font(name="Arial", size=9, bold=True, color="FF111827")
        status_cell.alignment = Alignment(horizontal="center", vertical="center")

        fill_match = PatternFill(fill_type="solid", fgColor="FFE4F0DD")     # мягкий зелёный
        fill_warn = PatternFill(fill_type="solid", fgColor="FFFFF2CC")      # мягкий жёлтый
        fill_check = PatternFill(fill_type="solid", fgColor="FFFCE4EC")     # мягкий красный

        if diff_pct <= 1e-12:
            status_cell.value = "MATCH"
            status_cell.fill = fill_match
        elif diff_pct < 0.10:
            status_cell.value = "MISMATCH"
            status_cell.fill = fill_warn
        else:
            status_cell.value = "CHECK"
            status_cell.fill = fill_check
    except Exception:
        pass

    countword = find_first_row_contains(ws, "Счет", col=1)
    if countword:
        for r in range(1, countword + 1):
            for cell in ws[r]:
                al = cell.alignment or Alignment()
                cell.alignment = Alignment(horizontal=al.horizontal, vertical=al.vertical, wrap_text=False)

    set_all_rows_height(ws, 12)
    return account_number


def add_cleaned_osv_files_to_analysis(analysis_wb, osv_files: Iterable[Tuple[str, bytes]]) -> Dict[str, List[str]]:
    report = {"added": [], "skipped": []}
    for fname, fbytes in osv_files:
        osv_wb, _ = load_wb_from_bytes(fbytes, fname)
        for ws in osv_wb.worksheets:
            clear_outline_for_sheet(ws)
        for ws in osv_wb.worksheets:
            acc = clean_osv_sheet_inplace(ws)
            if not acc:
                # Не выбрасываем лист: добавляем в _Анализ, но называем нейтрально, чтобы не ломать сборку.
                fallback = make_unique_sheet_title(analysis_wb, "OSV")
                copy_sheet(ws, analysis_wb, fallback)
                report["skipped"].append(f"{fname}:{ws.title} (не найден счет во 2-й строке)")
                report["added"].append(f"{fname}:{ws.title} -> {fallback}")
            else:
                # Суффикс счета должен остаться (важно: Сальдо ищет *1210/*1710/*3310/*3510 по названию листа).
                if acc.isdigit() and len(acc) == 4:
                    new_title = make_unique_with_fixed_suffix(analysis_wb, "", acc)
                else:
                    new_title = make_unique_sheet_title(analysis_wb, acc)
                copy_sheet(ws, analysis_wb, new_title)
                report["added"].append(f"{fname}:{ws.title} -> {new_title}")
    return report


def compute_availability_from_wb(wb) -> Dict[str, object]:
    saldo_suffixes = {"1210", "1710", "3310", "3510"}

    # Запасы: считаем доступными, если есть хотя бы один ОСВ-лист с 4-значным счётом, начинающимся на "13"
    inventory_accounts_found: List[str] = []
    inv_set: Set[str] = set()
    for sh in wb.sheetnames:
        _p, suf4 = split_prefix_suffix4(sh)
        if suf4.isdigit() and len(suf4) == 4 and suf4.startswith("13"):
            inv_set.add(suf4)
    if inv_set:
        inventory_accounts_found = sorted(inv_set)

    saldo_ok = any(split_prefix_suffix4(sh)[1] in saldo_suffixes for sh in wb.sheetnames)

    # "общ*" (общая ОСВ) — для чистой прибыли достаточно наличия хотя бы одного такого листа
    obsh_sheets = [sh for sh in wb.sheetnames if "общ" in str(sh).lower()]
    profit_ok = bool(obsh_sheets)

    prefix_to_pair: Dict[str, Dict[str, str]] = defaultdict(dict)
    for sh in wb.sheetnames:
        prefix, suf2 = split_prefix_suffix2(sh)
        if suf2 in ("wd", "md"):
            prefix = normalize_prefix(prefix)
            if suf2 not in prefix_to_pair[prefix]:
                prefix_to_pair[prefix][suf2] = sh
    contracts_prefixes = [p for p, d in prefix_to_pair.items() if "wd" in d and "md" in d]
    contracts_ok = bool(contracts_prefixes)

    # Инсайты: нужен набор листов W/M/Wt/Mt (возможно с префиксом).
    required_ins = {'w','m','wt','mt'}
    prefix_to_ins: Dict[str, Set[str]] = defaultdict(set)
    for sh in wb.sheetnames:
        low = sh.lower()
        for suf in ('wt','mt','w','m','кред'):
            if low.endswith(suf):
                prefix_to_ins[normalize_prefix(sh[:-len(suf)])].add(suf)
    insights_prefixes = [p for p, s in prefix_to_ins.items() if required_ins.issubset(s)]

    # Листы инсайтов уже могут быть в файле и их нельзя перезатирать.
    # Также поддерживаем миграцию старого имени листа: "инсайты" -> "инс".
    # Поэтому считаем доступность как:
    # - есть ли хотя бы один префикс-набор, где "инс" ещё НЕ создан
    # - или есть ли хотя бы один старый лист "инсайты", который нужно переименовать.
    insights_existing_titles: List[str] = []
    insights_missing_titles: List[str] = []
    insights_legacy_titles: List[str] = []
    for pfx in insights_prefixes:
        base_new = 'инс' if not pfx else f'{pfx}инс'
        title_new = base_new[:31]
        base_old = 'инсайты' if not pfx else f'инсайты {pfx}'
        title_old = base_old[:31]

        if title_new in wb.sheetnames:
            insights_existing_titles.append(title_new)
        elif title_old in wb.sheetnames:
            insights_existing_titles.append(title_old)
            insights_legacy_titles.append(title_old)
        else:
            insights_missing_titles.append(title_new)

    insights_ok = bool(insights_missing_titles) or bool(insights_legacy_titles)

    # Госы: листы, содержащие "гос" в названии (без учёта регистра)
    gos_sheets = [sh for sh in wb.sheetnames if "гос" in str(sh).lower()]
    gos_ok = bool(gos_sheets)

    return {
        "inventory_accounts_found": inventory_accounts_found,
        "saldo_ok": saldo_ok,
        "contracts_ok": contracts_ok,
        "contracts_prefixes": contracts_prefixes,
        "profit_ok": profit_ok,
        "obsh_sheets": obsh_sheets,
        "insights_ok": insights_ok,
        "insights_prefixes": insights_prefixes,
        "insights_existing_titles": insights_existing_titles,
        "insights_missing_titles": insights_missing_titles,
        "insights_legacy_titles": insights_legacy_titles,
        "gos_ok": gos_ok,
        "gos_sheets": gos_sheets,
    }


def find_existing_saldo_prefixes(wb) -> Dict[str, Set[str]]:
    """For each saldo account (1210/1710/3310/3510), returns existing prefixes found in sheetnames."""
    accounts = {"1210", "1710", "3310", "3510"}
    out: Dict[str, Set[str]] = {a: set() for a in accounts}
    for sh in wb.sheetnames:
        prefix, suf = split_prefix_suffix4(sh)
        if suf in accounts:
            out[suf].add(normalize_prefix(prefix))
    return out


def list_existing_saldo_sheets_with_a1(wb) -> List[Tuple[str, str, str]]:
    """Возвращает кортежи (имя_листа, счет_суффикс, A1_текст) для ОСВ-счетов сальдо."""
    accounts = {"1210", "1710", "3310", "3510"}
    out = []
    for sh in wb.sheetnames:
        prefix, suf = split_prefix_suffix4(sh)
        if suf in accounts:
            try:
                a1 = extract_company_label_from_a1(wb[sh])
            except Exception:
                a1 = ""
            out.append((sh, suf, _short(a1)))
    return out


def list_existing_osv_sheets_with_a1(wb) -> List[Tuple[str, str, str]]:
    """Возвращает кортежи (имя_листа, 4-значный суффикс, A1_текст) для всех ОСВ-листов."""
    out = []
    for sh in wb.sheetnames:
        _prefix, suf = split_prefix_suffix4(sh)
        if suf.isdigit() and len(suf) == 4:
            try:
                a1 = extract_company_label_from_a1(wb[sh])
            except Exception:
                a1 = ""
            out.append((sh, suf, _short(a1)))
    return out


def build_analysis_workbook(
    analysis_file: Optional[Tuple[str, bytes]],
    wh_files: List[Tuple[str, bytes, str]],
    m_files: List[Tuple[str, bytes, str]],
    osv_files: List[Tuple[str, bytes]],
    analysis_name_for_new: str,
    osv_prefix_by_sheet: Optional[Dict[Tuple[str, str], str]] = None,
    progress_cb: Optional[Callable[[float, str], None]] = None,
) -> Tuple[bytes, str, Dict[str, object], Dict[str, List[str]]]:
    report: Dict[str, List[str]] = {"warnings": [], "copied": []}

    def _progress(frac: float, msg: str) -> None:
        if progress_cb:
            try:
                progress_cb(max(0.0, min(1.0, float(frac))), msg)
            except Exception:
                pass

    total_steps = 2 + (1 if analysis_file else 0) + len(wh_files) + len(m_files) + len(osv_files)
    done_steps = 0

    def _step(msg: str) -> None:
        nonlocal done_steps
        done_steps += 1
        _progress(done_steps / max(1, total_steps), f"Сборка ({done_steps}/{total_steps}): {msg}")

    _progress(0.0, "Сборка: старт…")

    if analysis_file:
        analysis_name, analysis_bytes = analysis_file
        analysis_wb, keep_vba = load_wb_from_bytes(analysis_bytes, analysis_name)
        out_name = analysis_name
        placeholder_title = None
        _step(f"Открываю {analysis_name}")
    else:
        analysis_wb = Workbook()
        # Держим видимый "служебный" лист, чтобы сохранение не падало с
        # "At least one sheet must be visible" даже если все входные листы пропущены/пустые.
        placeholder_ws = analysis_wb.active
        placeholder_title = "Сборка"
        placeholder_ws.title = placeholder_title
        placeholder_ws["A1"] = "Служебный лист. Удалится автоматически, если добавятся другие листы."
        keep_vba = False
        base = (analysis_name_for_new or "").strip()
        out_name = safe_filename(f"_Анализ {base}".strip() + ".xlsx")
        _step("Создаю новый _Анализ")

    if wh_files or m_files:
        missing_wh_all: List[str] = []
        missing_m_all: List[str] = []

        for wh_name, wh_bytes, pref in wh_files:
            merge_report = merge_wh_m_into_analysis_with_prefix(analysis_wb, wh_bytes, wh_name, None, "", prefix=pref)
            missing_wh_all.extend([f"{wh_name}:{x}" for x in merge_report["missing_wh"]])
            report["copied"].extend(merge_report["copied"])
            _step(f"WH_KZ: {wh_name} (листов: {len(merge_report.get('copied') or [])})")

        for m_name, m_bytes, pref in m_files:
            merge_report = merge_wh_m_into_analysis_with_prefix(analysis_wb, None, "", m_bytes, m_name, prefix=pref)
            missing_m_all.extend([f"{m_name}:{x}" for x in merge_report["missing_m"]])
            report["copied"].extend(merge_report["copied"])
            _step(f"M_KZ: {m_name} (листов: {len(merge_report.get('copied') or [])})")

        if missing_wh_all or missing_m_all:
            msg = []
            if missing_wh_all:
                msg.append("WH: " + ", ".join(missing_wh_all[:6]) + (" ..." if len(missing_wh_all) > 6 else ""))
            if missing_m_all:
                msg.append("M: " + ", ".join(missing_m_all[:6]) + (" ..." if len(missing_m_all) > 6 else ""))
            report["warnings"].append("WH/M: не найдены листы. " + " | ".join(msg))

    if osv_files:
        osv_prefix_by_sheet = osv_prefix_by_sheet or {}
        # Для ОСВ-счетов (1210/1710/3310/3510) префиксы задаём по каждому листу, чтобы избежать коллизий,
        # но при этом сохранить суффикс счета в названии (это критично для дальнейшей обработки).
        report_local = {"added": [], "skipped": []}
        for fname, fbytes in osv_files:
            _step(f"ОСВ: очищаю {fname}")
            osv_wb, _ = load_wb_from_bytes(fbytes, fname)
            for ws in osv_wb.worksheets:
                clear_outline_for_sheet(ws)
            for idx, ws in enumerate(osv_wb.worksheets):
                original_title = ws.title
                acc = clean_osv_sheet_inplace(ws)
                if not acc:
                    fallback = make_unique_sheet_title(analysis_wb, "OSV")
                    copy_sheet(ws, analysis_wb, fallback)
                    report_local["skipped"].append(f"{fname}:{ws.title} (не найден счет во 2-й строке)")
                    report_local["added"].append(f"{fname}:{ws.title} -> {fallback}")
                    continue

                if acc.isdigit() and len(acc) == 4:
                    p = (osv_prefix_by_sheet.get((fname, original_title, idx)) or "").strip()
                    new_title = make_unique_with_fixed_suffix(analysis_wb, p, acc)
                else:
                    new_title = make_unique_sheet_title(analysis_wb, acc)
                copy_sheet(ws, analysis_wb, new_title)
                report_local["added"].append(f"{fname}:{ws.title} -> {new_title}")

        osv_report = report_local
        if osv_report["skipped"]:
            details = "; ".join(osv_report["skipped"][:8])
            more = f" (+{len(osv_report['skipped'])-8})" if len(osv_report["skipped"]) > 8 else ""
            report["warnings"].append(f"ОСВ: пропущены листы: {len(osv_report['skipped'])}. {details}{more}")
        report["copied"].extend(osv_report["added"])

    # Удаляем служебный лист, если в книге уже появился хотя бы один "нормальный" лист.
    if placeholder_title and len(analysis_wb.sheetnames) > 1 and placeholder_title in analysis_wb.sheetnames:
        del analysis_wb[placeholder_title]

    # Гарантируем, что есть хотя бы один видимый лист (требование openpyxl при сохранении).
    if not analysis_wb.sheetnames:
        ws = analysis_wb.create_sheet("Сборка")
        ws["A1"] = "Пустой файл: не удалось добавить ни одного листа."
    visible = [ws for ws in analysis_wb.worksheets if getattr(ws, "sheet_state", "visible") == "visible"]
    if not visible:
        analysis_wb.worksheets[0].sheet_state = "visible"

    availability = compute_availability_from_wb(analysis_wb)

    _step("Сохраняю файл…")
    out = io.BytesIO()
    analysis_wb.save(out)
    return out.getvalue(), out_name, availability, report


# =========================
# CODE 1 (Сальдо) — несколько компаний через префиксы
# Ищем листы, у которых последние 4 символа = 1210/1710/3310/3510
# Создаём лист на каждый префикс: "<префикс>сальд" или "сальд" (если префикса нет)
# =========================
def run_code_1(file_bytes: bytes) -> bytes:
    wb = load_workbook(io.BytesIO(file_bytes))
    xls = pd.ExcelFile(io.BytesIO(file_bytes))

    target_suffixes = {"1210": 6, "1710": 6, "3310": 7, "3510": 7}

    prefix_to_sheets: Dict[str, Dict[str, str]] = defaultdict(dict)
    for sh in xls.sheet_names:
        prefix, suf = split_prefix_suffix4(sh)
        if suf in target_suffixes:
            prefix = normalize_prefix(prefix)
            if suf not in prefix_to_sheets[prefix]:
                prefix_to_sheets[prefix][suf] = sh

    if not prefix_to_sheets:
        raise ValueError("Код 1: не найдено листов, заканчивающихся на 1210/1710/3310/3510.")

    def _excel_sheet_ref(sheet_name: str) -> str:
        # Если нужно — оборачиваем имя листа в кавычки и экранируем апострофы (правила Excel).
        if sheet_name is None:
            return "''"
        s = str(sheet_name)
        needs = any(ch in s for ch in " []()!,'\"") or (" " in s)
        if "'" in s:
            s = s.replace("'", "''")
            needs = True
        return f"'{s}'" if needs else s

    def _find_prefixed_sheetname(prefix_norm: str, base2: str) -> Optional[str]:
        """
        Ищет лист вида "<префикс>Wr" / "<префикс>Mr" (регистр не важен),
        также допускается суффикс " (n)". Возвращает лучший кандидат или None.
        """
        base2_l = base2.lower()
        best = None
        best_key = None
        for name in wb.sheetnames:
            nm = str(name)
            nm2 = re.sub(r"\s*\(\d+\)$", "", nm).strip()
            if not nm2.lower().endswith(base2_l):
                continue
            pref_part = nm2[: -len(base2_l)]
            if normalize_prefix(pref_part) != prefix_norm:
                continue
            m = re.search(r"\((\d+)\)\s*$", nm)
            n = int(m.group(1)) if m else 0
            key = (1 if m else 0, n, len(nm))  # предпочитаем точное имя без " (n)"
            if best is None or key < best_key:
                best = nm
                best_key = key
        return best

    def _find_prefixed_sheetname_any(prefix_norm: str, base: str) -> Optional[str]:
        """
        Ищет лист вида "<префикс><base>" (регистр не важен),
        также допускается суффикс " (n)". Возвращает лучший кандидат или None.
        """
        base_l = (base or "").lower()
        best = None
        best_key = None
        for name in wb.sheetnames:
            nm = str(name)
            nm2 = re.sub(r"\s*\(\d+\)$", "", nm).strip()
            if not nm2.lower().endswith(base_l):
                continue
            pref_part = nm2[: -len(base_l)]
            if normalize_prefix(pref_part) != prefix_norm:
                continue
            m = re.search(r"\((\d+)\)\s*$", nm)
            n = int(m.group(1)) if m else 0
            key = (1 if m else 0, n, len(nm))  # предпочитаем точное имя без " (n)"
            if best is None or key < best_key:
                best = nm
                best_key = key
        return best

    TOP_N = 30

    def _topn_pos_neg(df: pd.DataFrame, col: str) -> pd.DataFrame:
        if df is None or df.empty or col not in df.columns:
            return df.iloc[0:0].copy()
        pos = df[df[col] > 0].nlargest(TOP_N, col) if (df[col] > 0).any() else df.iloc[0:0]
        neg = df[df[col] < 0].nsmallest(TOP_N, col) if (df[col] < 0).any() else df.iloc[0:0]
        out = pd.concat([pos, neg], ignore_index=True)
        return out.reset_index(drop=True)

    for prefix, suf_map in prefix_to_sheets.items():
        sheet_data = {}

        # Для расширенных колонок в "сальд": начало/обороты/конец по 1210/1710/3310/3510
        # (используем фиксированные позиции как в ОСВ: C..H -> индексы 2..7).
        sheet_metrics: Dict[str, Dict[str, pd.Series]] = {}

        def _clean_contr(v) -> str:
            if v is None:
                return ""
            s = str(v).strip()
            if not s:
                return ""
            if s.lower().startswith("итого"):
                return ""
            if s in ("1210", "1710", "3310", "3510"):
                return ""
            return s

        def _series_by_contr(df: pd.DataFrame, col_i: int) -> pd.Series:
            if df is None or df.empty or int(df.shape[1]) <= int(col_i):
                return pd.Series(dtype=float)
            temp = df.iloc[:, [0, col_i]].copy()
            temp.columns = ["Контрагент", "value"]
            temp["Контрагент"] = temp["Контрагент"].map(_clean_contr)
            temp = temp[temp["Контрагент"] != ""]
            temp["value"] = pd.to_numeric(temp["value"], errors="coerce").fillna(0)
            temp = temp[temp["value"] != 0]
            if temp.empty:
                return pd.Series(dtype=float)
            return temp.groupby("Контрагент")["value"].sum()

        for suf, col_idx in target_suffixes.items():
            if suf not in suf_map:
                continue

            df = pd.read_excel(xls, sheet_name=suf_map[suf], header=0)
            if df is None or df.empty:
                continue

            # Конец периода (как было раньше)
            s_end = _series_by_contr(df, col_idx)
            if not s_end.empty:
                sheet_data[suf] = s_end

            # Расширенные показатели (начало/обороты/конец по нужной стороне)
            m = {
                "bop_deb": _series_by_contr(df, 2),   # C
                "bop_cred": _series_by_contr(df, 3),  # D
                "turn_deb": _series_by_contr(df, 4),  # E
                "turn_cred": _series_by_contr(df, 5), # F
                "eop_deb": _series_by_contr(df, 6),   # G
                "eop_cred": _series_by_contr(df, 7),  # H
            }
            sheet_metrics[suf] = m

        s1210 = sheet_data.get("1210", pd.Series(dtype=float))
        s3510 = sheet_data.get("3510", pd.Series(dtype=float))
        s1710 = sheet_data.get("1710", pd.Series(dtype=float))
        s3310 = sheet_data.get("3310", pd.Series(dtype=float))

        cust_set = set(s1210.index).union(set(s3510.index))
        supp_set = set(s1710.index).union(set(s3310.index))
        all_set = cust_set.union(supp_set)

        if not cust_set and not supp_set:
            continue

        if cust_set:
            df_cust = pd.DataFrame(sorted(cust_set), columns=["Контрагент"])

            m1210 = sheet_metrics.get("1210", {})
            m3510 = sheet_metrics.get("3510", {})

            def _ser(m: Dict[str, pd.Series], key: str) -> pd.Series:
                s = m.get(key)
                return s if isinstance(s, pd.Series) else pd.Series(dtype=float)

            df_cust["1210"] = df_cust["Контрагент"].map(s1210).fillna(0) / 1000
            df_cust["3510"] = df_cust["Контрагент"].map(s3510).fillna(0) / 1000
            df_cust["1210_нач"] = df_cust["Контрагент"].map(_ser(m1210, "bop_deb")).fillna(0) / 1000
            df_cust["3510_нач"] = df_cust["Контрагент"].map(_ser(m3510, "bop_cred")).fillna(0) / 1000
            df_cust["нетто_нач"] = df_cust["1210_нач"] - df_cust["3510_нач"]
            df_cust["оборот_дт"] = (
                df_cust["Контрагент"].map(_ser(m1210, "turn_deb")).fillna(0)
                + df_cust["Контрагент"].map(_ser(m3510, "turn_deb")).fillna(0)
            ) / 1000
            df_cust["оборот_кт"] = (
                df_cust["Контрагент"].map(_ser(m1210, "turn_cred")).fillna(0)
                + df_cust["Контрагент"].map(_ser(m3510, "turn_cred")).fillna(0)
            ) / 1000

            df_cust["сальдо заказчики"] = df_cust["1210"] - df_cust["3510"]
            df_cust = df_cust.sort_values(by="сальдо заказчики", ascending=False).reset_index(drop=True)
        else:
            df_cust = pd.DataFrame(columns=[
                "Контрагент", "1210_нач", "3510_нач", "нетто_нач", "оборот_дт", "оборот_кт",
                "1210", "3510", "сальдо заказчики",
            ])

        if supp_set:
            df_supp = pd.DataFrame(sorted(supp_set), columns=["Контрагент"])

            m1710 = sheet_metrics.get("1710", {})
            m3310 = sheet_metrics.get("3310", {})

            def _ser(m: Dict[str, pd.Series], key: str) -> pd.Series:
                s = m.get(key)
                return s if isinstance(s, pd.Series) else pd.Series(dtype=float)

            df_supp["1710"] = df_supp["Контрагент"].map(s1710).fillna(0) / 1000
            df_supp["3310"] = df_supp["Контрагент"].map(s3310).fillna(0) / 1000
            df_supp["1710_нач"] = df_supp["Контрагент"].map(_ser(m1710, "bop_deb")).fillna(0) / 1000
            df_supp["3310_нач"] = df_supp["Контрагент"].map(_ser(m3310, "bop_cred")).fillna(0) / 1000
            df_supp["нетто_нач"] = df_supp["1710_нач"] - df_supp["3310_нач"]
            df_supp["оборот_дт"] = (
                df_supp["Контрагент"].map(_ser(m1710, "turn_deb")).fillna(0)
                + df_supp["Контрагент"].map(_ser(m3310, "turn_deb")).fillna(0)
            ) / 1000
            df_supp["оборот_кт"] = (
                df_supp["Контрагент"].map(_ser(m1710, "turn_cred")).fillna(0)
                + df_supp["Контрагент"].map(_ser(m3310, "turn_cred")).fillna(0)
            ) / 1000

            df_supp["сальдо поставщики"] = df_supp["1710"] - df_supp["3310"]
            df_supp = df_supp.sort_values(by="сальдо поставщики", ascending=False).reset_index(drop=True)
        else:
            df_supp = pd.DataFrame(columns=[
                "Контрагент", "1710_нач", "3310_нач", "нетто_нач", "оборот_дт", "оборот_кт",
                "1710", "3310", "сальдо поставщики",
            ])

        if all_set:
            df_total = pd.DataFrame(sorted(all_set), columns=["Контрагент"])
            df_total["1210"] = df_total["Контрагент"].map(s1210).fillna(0) / 1000
            df_total["1710"] = df_total["Контрагент"].map(s1710).fillna(0) / 1000
            df_total["3310"] = df_total["Контрагент"].map(s3310).fillna(0) / 1000
            df_total["3510"] = df_total["Контрагент"].map(s3510).fillna(0) / 1000
            df_total["общее сальдо"] = df_total["1210"] + df_total["1710"] - df_total["3310"] - df_total["3510"]
            df_total = df_total.sort_values(by="общее сальдо", ascending=False).reset_index(drop=True)
        else:
            df_total = pd.DataFrame(columns=["Контрагент", "общее сальдо"])

        # =========================
        # Лист "сальд" — топ-15 (плюс/минус) + блоки оплат/выполнений с формулами по месяцам
        # =========================
        out2_name = safe_sheet_name(f"{prefix}сальд" if prefix else "сальд")
        if out2_name in wb.sheetnames:
            wb.remove(wb[out2_name])

        # Порядок вкладок: "сальд" должен идти сразу после листа "общ" соответствующего префикса.
        # Если "общ" не найден — добавляем в конец.
        pref_norm = normalize_prefix(prefix)
        obsh_name = _find_prefixed_sheetname_any(pref_norm, "общ") or (f"{prefix}общ" if prefix else "общ")
        insert_index = None
        if obsh_name in wb.sheetnames:
            try:
                insert_index = wb.sheetnames.index(obsh_name) + 1
            except Exception:
                insert_index = None
        ws2 = wb.create_sheet(out2_name, insert_index) if insert_index is not None else wb.create_sheet(out2_name)

        ws2["A1"] = "Все значения указаны в тысячах тенге"
        ws2["A1"].font = Font(name="Aptos Narrow", size=9, bold=True)

        # На "сальд" везде используем Aptos Narrow 9.
        font_h = Font(name="Aptos Narrow", size=9, bold=True)
        font_b = Font(name="Aptos Narrow", size=9)
        font_bb = Font(name="Aptos Narrow", size=9, bold=True)
        align_c = Alignment(horizontal="center")
        align_l = Alignment(horizontal="left")
        num_fmt = "#,##0;[Red](#,##0)"

        # A2 = 6 (влево + мягкая серая заливка)
        ws2["A2"] = 6
        ws2["A2"].alignment = Alignment(horizontal="left")
        ws2["A2"].fill = PatternFill("solid", fgColor="D9D9D9")

        # Используем конкатенацию через &, чтобы Excel не добавлял неявное пересечение ("@")
        ws2["L2"] = '="Опл L"&$A$2&"M"'
        ws2["M2"] = '="Вып L"&$A$2&"M"'
        ws2["K2"] = "коммент"

        # Заголовки месяцев (сдвинуты вправо из-за доп. колонок после "Контрагент")
        months = [f"2025_{m:02d}" for m in range(1, 13)] + [f"2026_{m:02d}" for m in range(1, 13)]
        start_col_pay = 14  # N
        start_col_perf = start_col_pay + len(months) + 1  # + разделитель
        last_used_col = start_col_perf + len(months) - 1

        ws2.cell(row=1, column=start_col_pay, value="Оплаты").font = font_h
        ws2.cell(row=1, column=start_col_perf, value="Выполнения").font = font_h

        for i, m in enumerate(months):
            ws2.cell(row=2, column=start_col_pay + i, value=m).font = font_h
            ws2.cell(row=2, column=start_col_pay + i).alignment = align_c
            ws2.cell(row=2, column=start_col_perf + i, value=m).font = font_h
            ws2.cell(row=2, column=start_col_perf + i).alignment = align_c

        # Заголовки таблиц (Заказчики)
        ws2["B2"] = "Контрагент"
        ws2["C2"] = "1210 нач"
        ws2["D2"] = "3510 нач"
        ws2["E2"] = "Нетто ДЗ (нач)"
        ws2["F2"] = "оборот ДТ"
        ws2["G2"] = "оборот КТ"
        ws2["H2"] = "1210"
        ws2["I2"] = "3510"
        ws2["J2"] = "Нетто ДЗ (кон)"
        for c in ("B2","C2","D2","E2","F2","G2","H2","I2","J2","K2","L2","M2"):
            ws2[c].font = font_h
            ws2[c].alignment = align_c if c != "B2" else align_l

        # Визуальные разделители: "ЗАКАЗЧИКИ" (до колонок L/M, не залезаем на блок месяцев)
        sec_fill = PatternFill(fill_type="solid", fgColor="1F2937")
        sec_font = Font(name="Aptos Narrow", size=9, bold=True, color="FFFFFF")
        for col in range(2, last_used_col + 1):  # B..(конец таблицы)
            c = ws2.cell(row=1, column=col)
            c.fill = sec_fill
            c.font = sec_font
        ws2["B1"] = "ЗАКАЗЧИКИ"
        ws2["B1"].alignment = align_l

        # Блоки "Поставщики" и "Общее сальдо" размещаем ниже динамически (так как добавляем строки сумм/итого).

        # Подбираем Wr/Mr для текущего префикса (если такие листы есть) — имена нужны внутри формул.
        wr_name = _find_prefixed_sheetname(pref_norm, "Wr") or (f"{prefix}Wr" if prefix else "Wr")
        mr_name = _find_prefixed_sheetname(pref_norm, "Mr") or (f"{prefix}Mr" if prefix else "Mr")
        wr_ref = _excel_sheet_ref(wr_name)
        mr_ref = _excel_sheet_ref(mr_name)

        pay_start = get_column_letter(start_col_pay)
        pay_end = get_column_letter(start_col_pay + len(months) - 1)
        perf_start = get_column_letter(start_col_perf)
        perf_end = get_column_letter(start_col_perf + len(months) - 1)

        def _comment_formula_for_row(r: int, which: str) -> str:
            """
            which: 'cust' or 'supp'
            Возвращает Excel-формулу для колонки F на основе:
            - оплат J:AG
            - выполнений AI:BF (нужно только для правила по 3510)
            """
            last_pos = f"IFERROR(LOOKUP(2,1/(${pay_start}{r}:${pay_end}{r}<>0),COLUMN(${pay_start}{r}:${pay_end}{r}))-COLUMN(${pay_start}{r})+1,0)"
            months_since_last_pay = (
                f"IF(({last_pos})=0,999,"
                f"DATEDIF("
                f"DATE(LEFT(INDEX(${pay_start}$2:${pay_end}$2,1,({last_pos})),4),RIGHT(INDEX(${pay_start}$2:${pay_end}$2,1,({last_pos})),2),1),"
                f"DATE(YEAR(TODAY()),MONTH(TODAY()),1),\"m\"))"
            )

            def _rule_by_last_pay(t_ok: int, t_warn: int, warn_label: str) -> str:
                return (
                    f"IF(({last_pos})=0,\"списание\","
                    f"IF(({months_since_last_pay})<={t_ok},\"ОК\","
                    f"IF(({months_since_last_pay})<={t_warn},\"{warn_label}\",\"списание\")))"
                )

            idx_now_pay = (
                f'IFERROR(MATCH(TEXT(TODAY(),"yyyy")&"_"&TEXT(TODAY(),"mm"),${pay_start}$2:${pay_end}$2,0),'
                f'COUNTA(${pay_start}$2:${pay_end}$2))'
            )
            idx_now_perf = (
                f'IFERROR(MATCH(TEXT(TODAY(),"yyyy")&"_"&TEXT(TODAY(),"mm"),${perf_start}$2:${perf_end}$2,0),'
                f'COUNTA(${perf_start}$2:${perf_end}$2))'
            )
            sum_pay_0_3 = f"SUM(INDEX(${pay_start}{r}:${pay_end}{r},1,MAX(1,({idx_now_pay})-2)):INDEX(${pay_start}{r}:${pay_end}{r},1,({idx_now_pay})))"
            sum_perf_0_3 = f"SUM(INDEX(${perf_start}{r}:${perf_end}{r},1,MAX(1,({idx_now_perf})-2)):INDEX(${perf_start}{r}:${perf_end}{r},1,({idx_now_perf})))"
            sum_pay_0_12 = f"SUM(INDEX(${pay_start}{r}:${pay_end}{r},1,MAX(1,({idx_now_pay})-11)):INDEX(${pay_start}{r}:${pay_end}{r},1,({idx_now_pay})))"
            sum_perf_0_12 = f"SUM(INDEX(${perf_start}{r}:${perf_end}{r},1,MAX(1,({idx_now_perf})-11)):INDEX(${perf_start}{r}:${perf_end}{r},1,({idx_now_perf})))"
            rule_3510 = (
                f"IF(AND(({sum_pay_0_3})<>0,({sum_perf_0_3})<>0),\"ОК\","
                f"IF(OR(({sum_pay_0_12})<>0,({sum_perf_0_12})<>0),\"сомнительно\",\"списание\"))"
            )

            if which == "cust":
                rule_1210 = _rule_by_last_pay(3, 6, "сомнительный")
                return f"=IF($B{r}=\"\",\"\",IF($J{r}>0,{rule_1210},{rule_3510}))"
            # поставщики
            rule_1710 = _rule_by_last_pay(3, 6, "сомнительный")
            rule_3310 = _rule_by_last_pay(3, 12, "сомнительный")
            return f"=IF($B{r}=\"\",\"\",IF($J{r}>0,{rule_1710},{rule_3310}))"

        def _rollup_pay_formula(r: int) -> str:
            return (
                f'=IF($B{r}="","",SUM('
                f'INDEX(${pay_start}{r}:${pay_end}{r},1,MAX(1,IFERROR(MATCH(TEXT(TODAY(),"yyyy")&"_"&TEXT(TODAY(),"mm"),${pay_start}$2:${pay_end}$2,0),COUNTA(${pay_start}$2:${pay_end}$2))-$A$2+1)):'
                f'INDEX(${pay_start}{r}:${pay_end}{r},1,IFERROR(MATCH(TEXT(TODAY(),"yyyy")&"_"&TEXT(TODAY(),"mm"),${pay_start}$2:${pay_end}$2,0),COUNTA(${pay_start}$2:${pay_end}$2)))))'
            )

        def _rollup_perf_formula(r: int) -> str:
            return (
                f'=IF($B{r}="","",SUM('
                f'INDEX(${perf_start}{r}:${perf_end}{r},1,MAX(1,IFERROR(MATCH(TEXT(TODAY(),"yyyy")&"_"&TEXT(TODAY(),"mm"),${perf_start}$2:${perf_end}$2,0),COUNTA(${perf_start}$2:${perf_end}$2))-$A$2+1)):'
                f'INDEX(${perf_start}{r}:${perf_end}{r},1,IFERROR(MATCH(TEXT(TODAY(),"yyyy")&"_"&TEXT(TODAY(),"mm"),${perf_start}$2:${perf_end}$2,0),COUNTA(${perf_start}$2:${perf_end}$2)))))'
            )

        def _write_customer_row(
            r: int,
            contr: str,
            v1210_bop: int,
            v3510_bop: int,
            vnet_bop: int,
            vturn_deb: int,
            vturn_cred: int,
            v1210: int,
            v3510: int,
            vsaldo: int,
        ) -> None:
            ws2.cell(row=r, column=2, value=contr).font = font_b
            ws2.cell(row=r, column=2).alignment = align_l

            # BOP/обороты/конец
            for c, val in [
                (3, v1210_bop),
                (4, v3510_bop),
                (5, vnet_bop),
                (6, vturn_deb),
                (7, vturn_cred),
                (8, v1210),
                (9, v3510),
                (10, vsaldo),
            ]:
                cell = ws2.cell(row=r, column=c, value=val)
                cell.number_format = num_fmt
                cell.font = font_b
                cell.alignment = align_c

            for c in range(3, 11):
                ws2.cell(row=r, column=c).alignment = align_c

            ws2.cell(row=r, column=11, value=_comment_formula_for_row(r, "cust")).alignment = align_l
            ws2.cell(row=r, column=11).font = font_b

            ws2.cell(row=r, column=12, value=_rollup_pay_formula(r)).alignment = align_c
            ws2.cell(row=r, column=12).font = font_b
            ws2.cell(row=r, column=12).number_format = num_fmt

            ws2.cell(row=r, column=13, value=_rollup_perf_formula(r)).alignment = align_c
            ws2.cell(row=r, column=13).font = font_b
            ws2.cell(row=r, column=13).number_format = num_fmt

            # J..AG: оплаты из Wr
            for i in range(len(months)):
                ci = start_col_pay + i
                col_letter = get_column_letter(ci)
                ws2.cell(
                    row=r,
                    column=ci,
                    value=(
                        f'=IF($B{r}="","",'
                        f'SUMIFS({wr_ref}!$F:$F,{wr_ref}!$G:$G,"1210",{wr_ref}!$P:$P,{col_letter}$2,{wr_ref}!$R:$R,$B{r})'
                        f'+SUMIFS({wr_ref}!$F:$F,{wr_ref}!$G:$G,"3510",{wr_ref}!$P:$P,{col_letter}$2,{wr_ref}!$R:$R,$B{r})'
                        f')'
                    ),
                ).number_format = num_fmt
                ws2.cell(row=r, column=ci).alignment = align_c

            # AI..AT (2025) *1.12 и AU..BF (2026) *1.16 — выполнения из Mr
            for i in range(len(months)):
                ci = start_col_perf + i
                col_letter = get_column_letter(ci)
                coef = 1.12 if i < 12 else 1.16
                ws2.cell(
                    row=r,
                    column=ci,
                    value=(
                        f'=IF($B{r}="","",SUMIFS({mr_ref}!$H:$H,{mr_ref}!$P:$P,{col_letter}$2,{mr_ref}!$Q:$Q,$B{r},{mr_ref}!$G:$G,"6010")*{coef})'
                    ),
                ).number_format = num_fmt
                ws2.cell(row=r, column=ci).alignment = align_c

        def _write_supplier_row(
            r: int,
            contr: str,
            v1710_bop: int,
            v3310_bop: int,
            vnet_bop: int,
            vturn_deb: int,
            vturn_cred: int,
            v1710: int,
            v3310: int,
            vsaldo: int,
        ) -> None:
            ws2.cell(row=r, column=2, value=contr).font = font_b
            ws2.cell(row=r, column=2).alignment = align_l

            for c, val in [
                (3, v1710_bop),
                (4, v3310_bop),
                (5, vnet_bop),
                (6, vturn_deb),
                (7, vturn_cred),
                (8, v1710),
                (9, v3310),
                (10, vsaldo),
            ]:
                cell = ws2.cell(row=r, column=c, value=val)
                cell.number_format = num_fmt
                cell.font = font_b
                cell.alignment = align_c

            ws2.cell(row=r, column=11, value=_comment_formula_for_row(r, "supp")).alignment = align_l
            ws2.cell(row=r, column=11).font = font_b

            ws2.cell(row=r, column=12, value=_rollup_pay_formula(r)).alignment = align_c
            ws2.cell(row=r, column=12).font = font_b
            ws2.cell(row=r, column=12).number_format = num_fmt

            # Для поставщиков: только оплаты J..AG (выполнений нет)
            for i in range(len(months)):
                ci = start_col_pay + i
                col_letter = get_column_letter(ci)
                ws2.cell(
                    row=r,
                    column=ci,
                    value=(
                        f'=IF($B{r}="","",'
                        f'SUMIFS({wr_ref}!$H:$H,{wr_ref}!$E:$E,"1710",{wr_ref}!$P:$P,{col_letter}$2,{wr_ref}!$Q:$Q,$B{r})'
                        f'+SUMIFS({wr_ref}!$H:$H,{wr_ref}!$E:$E,"3310",{wr_ref}!$P:$P,{col_letter}$2,{wr_ref}!$Q:$Q,$B{r})'
                        f')'
                    ),
                ).number_format = num_fmt
                ws2.cell(row=r, column=ci).alignment = align_c

        def _write_summary_rows(kind: str, df_total_src: pd.DataFrame, pos_start: int, pos_end: int, sum_row: int, other_row: int, total_row: int) -> None:
            """
            Writes TOP-N sum, прочее, ИТОГО rows for a block.
            kind: 'cust' or 'supp'
            """
            # Подписи строк
            # Жирным выделяем только строки "ТОП-15" и "ИТОГО"; "прочее" оставляем обычным.
            ws2.cell(row=sum_row, column=2, value=f"ТОП-{TOP_N}").font = font_h
            ws2.cell(row=other_row, column=2, value="прочее").font = font_b
            ws2.cell(row=total_row, column=2, value="ИТОГО").font = font_h
            for rr in (sum_row, other_row, total_row):
                ws2.cell(row=rr, column=2).alignment = align_l

            def _sum_formula(col_letter: str) -> str:
                return f"=SUM({col_letter}{pos_start}:{col_letter}{pos_end})"

            # Суммы по ТОП-N (числовые колонки + роллапы + помесячные)
            base_cols = ["C", "D", "E", "F", "G", "H", "I", "J", "L"]
            if kind == "cust":
                base_cols.append("M")
            for col_letter in base_cols:
                c = ws2[f"{col_letter}{sum_row}"]
                c.value = _sum_formula(col_letter)
                c.font = font_bb
                c.alignment = align_c
                c.number_format = num_fmt

            for i in range(len(months)):
                ci = start_col_pay + i
                col_letter = get_column_letter(ci)
                c = ws2.cell(row=sum_row, column=ci, value=_sum_formula(col_letter))
                c.alignment = align_c
                c.number_format = num_fmt
                c.font = font_bb
            if kind == "cust":
                for i in range(len(months)):
                    ci = start_col_perf + i
                    col_letter = get_column_letter(ci)
                    c = ws2.cell(row=sum_row, column=ci, value=_sum_formula(col_letter))
                    c.alignment = align_c
                    c.number_format = num_fmt
                    c.font = font_bb

            # ИТОГО:
            # - по сальдо (C/D/E) пишем числом (посчитано в pandas для текущего блока: + или -)
            # - оплаты/выполнения по месяцам считаем формулами без критерия контрагента
            if kind == "cust":
                def _sum_df(col: str) -> int:
                    return int(round(float(df_total_src[col].sum() if df_total_src is not None and not df_total_src.empty and col in df_total_src.columns else 0.0)))

                ws2[f"C{total_row}"] = _sum_df("1210_нач")
                ws2[f"D{total_row}"] = _sum_df("3510_нач")
                ws2[f"E{total_row}"] = _sum_df("нетто_нач")
                ws2[f"F{total_row}"] = _sum_df("оборот_дт")
                ws2[f"G{total_row}"] = _sum_df("оборот_кт")
                ws2[f"H{total_row}"] = _sum_df("1210")
                ws2[f"I{total_row}"] = _sum_df("3510")
                tot_net = _sum_df("сальдо заказчики")
                if df_total_src is not None and not df_total_src.empty:
                    try:
                        if float(df_total_src["сальдо заказчики"].sum()) < 0:
                            tot_net = abs(int(tot_net))
                    except Exception:
                        pass
                ws2[f"J{total_row}"] = tot_net
                for addr in (f"C{total_row}", f"D{total_row}", f"E{total_row}", f"F{total_row}", f"G{total_row}", f"H{total_row}", f"I{total_row}", f"J{total_row}"):
                    ws2[addr].font = font_bb
                    ws2[addr].alignment = align_c
                    ws2[addr].number_format = num_fmt

                # ИТОГО по месяцам (оплаты Wr, выполнения Mr) без фильтра по контрагенту
                for i in range(len(months)):
                    ci = start_col_pay + i
                    col_letter = get_column_letter(ci)
                    ws2.cell(
                        row=total_row,
                        column=ci,
                        value=(
                            f'=SUMIFS({wr_ref}!$F:$F,{wr_ref}!$G:$G,"1210",{wr_ref}!$P:$P,{col_letter}$2)'
                            f'+SUMIFS({wr_ref}!$F:$F,{wr_ref}!$G:$G,"3510",{wr_ref}!$P:$P,{col_letter}$2)'
                        ),
                    ).number_format = num_fmt
                    ws2.cell(row=total_row, column=ci).alignment = align_c
                    ws2.cell(row=total_row, column=ci).font = font_bb

                for i in range(len(months)):
                    ci = start_col_perf + i
                    col_letter = get_column_letter(ci)
                    coef = 1.12 if i < 12 else 1.16
                    ws2.cell(
                        row=total_row,
                        column=ci,
                        value=f'=SUMIFS({mr_ref}!$H:$H,{mr_ref}!$P:$P,{col_letter}$2,{mr_ref}!$G:$G,"6010")*{coef}',
                    ).number_format = num_fmt
                    ws2.cell(row=total_row, column=ci).alignment = align_c
                    ws2.cell(row=total_row, column=ci).font = font_bb

                # G/H суммы за "последние A2 месяцев" для строки ИТОГО
                ws2.cell(row=total_row, column=12, value=_rollup_pay_formula(total_row)).alignment = align_c
                ws2.cell(row=total_row, column=12).number_format = num_fmt
                ws2.cell(row=total_row, column=12).font = font_bb
                ws2.cell(row=total_row, column=13, value=_rollup_perf_formula(total_row)).alignment = align_c
                ws2.cell(row=total_row, column=13).number_format = num_fmt
                ws2.cell(row=total_row, column=13).font = font_bb

            else:
                def _sum_df(col: str) -> int:
                    return int(round(float(df_total_src[col].sum() if df_total_src is not None and not df_total_src.empty and col in df_total_src.columns else 0.0)))

                ws2[f"C{total_row}"] = _sum_df("1710_нач")
                ws2[f"D{total_row}"] = _sum_df("3310_нач")
                ws2[f"E{total_row}"] = _sum_df("нетто_нач")
                ws2[f"F{total_row}"] = _sum_df("оборот_дт")
                ws2[f"G{total_row}"] = _sum_df("оборот_кт")
                ws2[f"H{total_row}"] = _sum_df("1710")
                ws2[f"I{total_row}"] = _sum_df("3310")
                tot_net = _sum_df("сальдо поставщики")
                if df_total_src is not None and not df_total_src.empty:
                    try:
                        if float(df_total_src["сальдо поставщики"].sum()) < 0:
                            tot_net = abs(int(tot_net))
                    except Exception:
                        pass
                ws2[f"J{total_row}"] = tot_net
                for addr in (f"C{total_row}", f"D{total_row}", f"E{total_row}", f"F{total_row}", f"G{total_row}", f"H{total_row}", f"I{total_row}", f"J{total_row}"):
                    ws2[addr].font = font_bb
                    ws2[addr].alignment = align_c
                    ws2[addr].number_format = num_fmt

                for i in range(len(months)):
                    ci = start_col_pay + i
                    col_letter = get_column_letter(ci)
                    ws2.cell(
                        row=total_row,
                        column=ci,
                        value=(
                            f'=SUMIFS({wr_ref}!$H:$H,{wr_ref}!$E:$E,"1710",{wr_ref}!$P:$P,{col_letter}$2)'
                            f'+SUMIFS({wr_ref}!$H:$H,{wr_ref}!$E:$E,"3310",{wr_ref}!$P:$P,{col_letter}$2)'
                        ),
                    ).number_format = num_fmt
                    ws2.cell(row=total_row, column=ci).alignment = align_c
                    ws2.cell(row=total_row, column=ci).font = font_bb

                ws2.cell(row=total_row, column=12, value=_rollup_pay_formula(total_row)).alignment = align_c
                ws2.cell(row=total_row, column=12).number_format = num_fmt
                ws2.cell(row=total_row, column=12).font = font_bb

            # прочее = ИТОГО - ТОП-N (для чисел и для помесячных колонок)
            other_cols = ["C", "D", "E", "F", "G", "H", "I", "J", "L"]
            if kind == "cust":
                other_cols.append("M")
            for col_letter in other_cols:
                ws2[f"{col_letter}{other_row}"] = f"={col_letter}{total_row}-{col_letter}{sum_row}"
                ws2[f"{col_letter}{other_row}"].font = font_b
                ws2[f"{col_letter}{other_row}"].alignment = align_c
                ws2[f"{col_letter}{other_row}"].number_format = num_fmt

            for i in range(len(months)):
                ci = start_col_pay + i
                col_letter = get_column_letter(ci)
                ws2.cell(row=other_row, column=ci, value=f"={col_letter}{total_row}-{col_letter}{sum_row}").number_format = num_fmt
                ws2.cell(row=other_row, column=ci).alignment = align_c
            if kind == "cust":
                for i in range(len(months)):
                    ci = start_col_perf + i
                    col_letter = get_column_letter(ci)
                    ws2.cell(row=other_row, column=ci, value=f"={col_letter}{total_row}-{col_letter}{sum_row}").number_format = num_fmt
                    ws2.cell(row=other_row, column=ci).alignment = align_c

        # === Заказчики (фиксированная раскладка + строки ТОП-N/прочее/ИТОГО)
        df_cust_pos_all = df_cust[df_cust["сальдо заказчики"] >= 0] if not df_cust.empty else df_cust
        df_cust_neg_all = df_cust[df_cust["сальдо заказчики"] < 0] if not df_cust.empty else df_cust
        cust_pos = df_cust_pos_all.nlargest(TOP_N, "сальдо заказчики") if df_cust_pos_all is not None and not df_cust_pos_all.empty else df_cust_pos_all
        cust_neg = df_cust_neg_all.nsmallest(TOP_N, "сальдо заказчики") if df_cust_neg_all is not None and not df_cust_neg_all.empty else df_cust_neg_all

        cust_pos_start = 3
        cust_pos_end = cust_pos_start + TOP_N - 1
        cust_pos_sum, cust_pos_other, cust_pos_total = cust_pos_end + 1, cust_pos_end + 2, cust_pos_end + 3
        # Отступ между блоками: 1 пустая строка, затем строка заголовков, затем данные.
        cust_gap = cust_pos_total + 1
        cust_neg_hdr = cust_gap + 1
        cust_neg_start = cust_neg_hdr + 1
        cust_neg_end = cust_neg_start + TOP_N - 1
        cust_neg_sum, cust_neg_other, cust_neg_total = cust_neg_end + 1, cust_neg_end + 2, cust_neg_end + 3
        cust_end_gap = cust_neg_total + 1

        # Для отрицательного блока заказчиков делаем отдельную строку заголовков (чтобы было явно "Нетто КЗ").
        for col in range(2, 14):  # B..M
            src = ws2.cell(row=2, column=col)
            dst = ws2.cell(row=cust_neg_hdr, column=col)
            dst.value = src.value
            dst.font = _copy(src.font)
            dst.alignment = _copy(src.alignment)
        ws2[f"J{cust_neg_hdr}"] = "Нетто КЗ"
        ws2[f"B{cust_neg_hdr}"] = "Контрагент"
        ws2[f"E{cust_neg_hdr}"] = "Нетто КЗ (нач)"
        ws2[f"J{cust_neg_hdr}"] = "Нетто КЗ (кон)"

        for i in range(TOP_N):
            if cust_pos is None or i >= len(cust_pos.index):
                continue
            rr = cust_pos_start + i
            r0 = cust_pos.iloc[i]
            contr = str(r0.get("Контрагент", "")).strip()
            if not contr:
                continue
            _write_customer_row(
                rr,
                contr,
                int(round(float(r0.get("1210_нач", 0) or 0))),
                int(round(float(r0.get("3510_нач", 0) or 0))),
                int(round(float(r0.get("нетто_нач", 0) or 0))),
                int(round(float(r0.get("оборот_дт", 0) or 0))),
                int(round(float(r0.get("оборот_кт", 0) or 0))),
                int(round(float(r0.get("1210", 0) or 0))),
                int(round(float(r0.get("3510", 0) or 0))),
                int(round(float(r0.get("сальдо заказчики", 0) or 0))),
            )

        _write_summary_rows("cust", df_cust_pos_all, cust_pos_start, cust_pos_end, cust_pos_sum, cust_pos_other, cust_pos_total)

        for i in range(TOP_N):
            if cust_neg is None or i >= len(cust_neg.index):
                continue
            rr = cust_neg_start + i
            r0 = cust_neg.iloc[i]
            contr = str(r0.get("Контрагент", "")).strip()
            if not contr:
                continue
            _write_customer_row(
                rr,
                contr,
                int(round(float(r0.get("1210_нач", 0) or 0))),
                int(round(float(r0.get("3510_нач", 0) or 0))),
                abs(int(round(float(r0.get("нетто_нач", 0) or 0)))),
                int(round(float(r0.get("оборот_дт", 0) or 0))),
                int(round(float(r0.get("оборот_кт", 0) or 0))),
                int(round(float(r0.get("1210", 0) or 0))),
                int(round(float(r0.get("3510", 0) or 0))),
                abs(int(round(float(r0.get("сальдо заказчики", 0) or 0)))),
            )

        _write_summary_rows("cust", df_cust_neg_all, cust_neg_start, cust_neg_end, cust_neg_sum, cust_neg_other, cust_neg_total)

        # === Поставщики (ниже блока заказчиков)
        supp_header_row = cust_end_gap + 2  # пустая строка после блока заказчиков
        # Визуальный разделитель
        for col in range(2, last_used_col + 1):  # B..(конец таблицы)
            c = ws2.cell(row=supp_header_row - 1, column=col)
            c.fill = sec_fill
            c.font = sec_font
        ws2.cell(row=supp_header_row - 1, column=2, value="ПОСТАВЩИКИ").alignment = align_l

        ws2[f"B{supp_header_row}"] = "Контрагент"
        ws2[f"C{supp_header_row}"] = "1710 нач"
        ws2[f"D{supp_header_row}"] = "3310 нач"
        ws2[f"E{supp_header_row}"] = "Нетто ДЗ (нач)"
        ws2[f"F{supp_header_row}"] = "оборот ДТ"
        ws2[f"G{supp_header_row}"] = "оборот КТ"
        ws2[f"H{supp_header_row}"] = "1710"
        ws2[f"I{supp_header_row}"] = "3310"
        # Для поставщиков в шапке по умолчанию пишем "Нетто ДЗ (кон)" (для положительного блока),
        # а для отрицательного блока сделаем отдельную шапку ниже.
        ws2[f"J{supp_header_row}"] = "Нетто ДЗ (кон)"
        ws2[f"K{supp_header_row}"] = "коммент"
        ws2[f"L{supp_header_row}"] = '="Опл L"&$A$2&"M"'
        ws2[f"M{supp_header_row}"] = ""  # выполнение не заполняем для поставщиков
        for caddr in (f"B{supp_header_row}", f"C{supp_header_row}", f"D{supp_header_row}", f"E{supp_header_row}", f"F{supp_header_row}", f"G{supp_header_row}", f"H{supp_header_row}", f"I{supp_header_row}", f"J{supp_header_row}", f"K{supp_header_row}", f"L{supp_header_row}", f"M{supp_header_row}"):
            ws2[caddr].font = font_h
            ws2[caddr].alignment = align_l if caddr.startswith("B") else align_c

        df_supp_pos_all = df_supp[df_supp["сальдо поставщики"] >= 0] if not df_supp.empty else df_supp
        df_supp_neg_all = df_supp[df_supp["сальдо поставщики"] < 0] if not df_supp.empty else df_supp
        supp_pos = df_supp_pos_all.nlargest(TOP_N, "сальдо поставщики") if df_supp_pos_all is not None and not df_supp_pos_all.empty else df_supp_pos_all
        supp_neg = df_supp_neg_all.nsmallest(TOP_N, "сальдо поставщики") if df_supp_neg_all is not None and not df_supp_neg_all.empty else df_supp_neg_all

        supp_pos_start = supp_header_row + 1
        supp_pos_end = supp_pos_start + TOP_N - 1
        supp_pos_sum, supp_pos_other, supp_pos_total = supp_pos_end + 1, supp_pos_end + 2, supp_pos_end + 3
        # Для отрицательного блока поставщиков тоже добавим отдельную строку заголовков.
        supp_neg_hdr = supp_pos_total + 2
        # Копируем заголовки поставщиков и меняем Нетто на "Нетто КЗ".
        for col in range(2, 14):  # B..M
            src = ws2.cell(row=supp_header_row, column=col)
            dst = ws2.cell(row=supp_neg_hdr, column=col)
            dst.value = src.value
            dst.font = _copy(src.font)
            dst.alignment = _copy(src.alignment)
        ws2[f"E{supp_neg_hdr}"] = "Нетто КЗ (нач)"
        ws2[f"J{supp_neg_hdr}"] = "Нетто КЗ (кон)"
        ws2[f"B{supp_neg_hdr}"] = "Контрагент"

        supp_neg_start = supp_neg_hdr + 1
        supp_neg_end = supp_neg_start + TOP_N - 1
        supp_neg_sum, supp_neg_other, supp_neg_total = supp_neg_end + 1, supp_neg_end + 2, supp_neg_end + 3

        for i in range(TOP_N):
            if supp_pos is None or i >= len(supp_pos.index):
                continue
            rr = supp_pos_start + i
            r0 = supp_pos.iloc[i]
            contr = str(r0.get("Контрагент", "")).strip()
            if not contr:
                continue
            _write_supplier_row(
                rr,
                contr,
                int(round(float(r0.get("1710_нач", 0) or 0))),
                int(round(float(r0.get("3310_нач", 0) or 0))),
                int(round(float(r0.get("нетто_нач", 0) or 0))),
                int(round(float(r0.get("оборот_дт", 0) or 0))),
                int(round(float(r0.get("оборот_кт", 0) or 0))),
                int(round(float(r0.get("1710", 0) or 0))),
                int(round(float(r0.get("3310", 0) or 0))),
                int(round(float(r0.get("сальдо поставщики", 0) or 0))),
            )
        _write_summary_rows("supp", df_supp_pos_all, supp_pos_start, supp_pos_end, supp_pos_sum, supp_pos_other, supp_pos_total)

        for i in range(TOP_N):
            if supp_neg is None or i >= len(supp_neg.index):
                continue
            rr = supp_neg_start + i
            r0 = supp_neg.iloc[i]
            contr = str(r0.get("Контрагент", "")).strip()
            if not contr:
                continue
            _write_supplier_row(
                rr,
                contr,
                int(round(float(r0.get("1710_нач", 0) or 0))),
                int(round(float(r0.get("3310_нач", 0) or 0))),
                abs(int(round(float(r0.get("нетто_нач", 0) or 0)))),
                int(round(float(r0.get("оборот_дт", 0) or 0))),
                int(round(float(r0.get("оборот_кт", 0) or 0))),
                int(round(float(r0.get("1710", 0) or 0))),
                int(round(float(r0.get("3310", 0) or 0))),
                abs(int(round(float(r0.get("сальдо поставщики", 0) or 0)))),
            )
        _write_summary_rows("supp", df_supp_neg_all, supp_neg_start, supp_neg_end, supp_neg_sum, supp_neg_other, supp_neg_total)

        # === Общее сальдо (ниже поставщиков): только значение сальдо + строки ТОП-15/прочее/ИТОГО
        total_header_row = supp_neg_total + 5  # пустая строка после блока поставщиков
        ws2[f"B{total_header_row}"] = "Контрагент"
        ws2[f"E{total_header_row}"] = "общее сальдо"
        ws2[f"B{total_header_row}"].font = font_h
        ws2[f"E{total_header_row}"].font = font_h
        ws2[f"B{total_header_row}"].alignment = align_l
        ws2[f"E{total_header_row}"].alignment = align_c

        df_total_pos_all = df_total[df_total["общее сальдо"] > 0] if not df_total.empty else df_total
        df_total_neg_all = df_total[df_total["общее сальдо"] < 0] if not df_total.empty else df_total
        total_pos = df_total_pos_all.nlargest(TOP_N, "общее сальдо") if df_total_pos_all is not None and not df_total_pos_all.empty else df_total_pos_all
        total_neg = df_total_neg_all.nsmallest(TOP_N, "общее сальдо") if df_total_neg_all is not None and not df_total_neg_all.empty else df_total_neg_all

        total_pos_start = total_header_row + 1
        total_pos_end = total_pos_start + TOP_N - 1
        total_pos_sum, total_pos_other, total_pos_total = total_pos_end + 1, total_pos_end + 2, total_pos_end + 3
        total_neg_start = total_pos_total + 2
        total_neg_end = total_neg_start + TOP_N - 1
        total_neg_sum, total_neg_other, total_neg_total = total_neg_end + 1, total_neg_end + 2, total_neg_end + 3

        def _write_total_row(r: int, contr: str, vtot: int) -> None:
            ws2.cell(row=r, column=2, value=contr).font = font_b
            ws2.cell(row=r, column=2).alignment = align_l
            # Для строк с контрагентами (топ-15) числа НЕ делаем жирными.
            ws2.cell(row=r, column=5, value=vtot).font = font_b
            ws2.cell(row=r, column=5).alignment = align_c
            ws2.cell(row=r, column=5).number_format = num_fmt

        for i in range(TOP_N):
            if total_pos is None or i >= len(total_pos.index):
                continue
            rr = total_pos_start + i
            r0 = total_pos.iloc[i]
            contr = str(r0.get("Контрагент", "")).strip()
            if not contr:
                continue
            _write_total_row(rr, contr, int(round(float(r0.get("общее сальдо", 0) or 0))))

        # Сводные строки (положительный блок)
        ws2.cell(row=total_pos_sum, column=2, value=f"ТОП-{TOP_N}").font = font_h
        ws2.cell(row=total_pos_other, column=2, value="прочее").font = font_h
        ws2.cell(row=total_pos_total, column=2, value="ИТОГО").font = font_h
        for rr in (total_pos_sum, total_pos_other, total_pos_total):
            ws2.cell(row=rr, column=2).alignment = align_l

        ws2[f"E{total_pos_sum}"] = f"=SUM(E{total_pos_start}:E{total_pos_end})"
        ws2[f"E{total_pos_sum}"].font = font_bb
        ws2[f"E{total_pos_sum}"].alignment = align_c
        ws2[f"E{total_pos_sum}"].number_format = num_fmt

        total_all = int(round(float(df_total_pos_all["общее сальдо"].sum() if df_total_pos_all is not None and not df_total_pos_all.empty else 0.0)))
        ws2[f"E{total_pos_total}"] = total_all
        ws2[f"E{total_pos_total}"].font = font_bb
        ws2[f"E{total_pos_total}"].alignment = align_c
        ws2[f"E{total_pos_total}"].number_format = num_fmt

        ws2[f"E{total_pos_other}"] = f"=E{total_pos_total}-E{total_pos_sum}"
        ws2[f"E{total_pos_other}"].font = font_b
        ws2[f"E{total_pos_other}"].alignment = align_c
        ws2[f"E{total_pos_other}"].number_format = num_fmt

        for i in range(TOP_N):
            if total_neg is None or i >= len(total_neg.index):
                continue
            rr = total_neg_start + i
            r0 = total_neg.iloc[i]
            contr = str(r0.get("Контрагент", "")).strip()
            if not contr:
                continue
            _write_total_row(rr, contr, int(round(float(r0.get("общее сальдо", 0) or 0))))

        # Сводные строки (отрицательный блок)
        ws2.cell(row=total_neg_sum, column=2, value=f"ТОП-{TOP_N}").font = font_h
        ws2.cell(row=total_neg_other, column=2, value="прочее").font = font_h
        ws2.cell(row=total_neg_total, column=2, value="ИТОГО").font = font_h
        for rr in (total_neg_sum, total_neg_other, total_neg_total):
            ws2.cell(row=rr, column=2).alignment = align_l

        ws2[f"E{total_neg_sum}"] = f"=SUM(E{total_neg_start}:E{total_neg_end})"
        ws2[f"E{total_neg_sum}"].font = font_bb
        ws2[f"E{total_neg_sum}"].alignment = align_c
        ws2[f"E{total_neg_sum}"].number_format = num_fmt

        total_all_neg = int(round(float(df_total_neg_all["общее сальдо"].sum() if df_total_neg_all is not None and not df_total_neg_all.empty else 0.0)))
        ws2[f"E{total_neg_total}"] = total_all_neg
        ws2[f"E{total_neg_total}"].font = font_bb
        ws2[f"E{total_neg_total}"].alignment = align_c
        ws2[f"E{total_neg_total}"].number_format = num_fmt

        ws2[f"E{total_neg_other}"] = f"=E{total_neg_total}-E{total_neg_sum}"
        ws2[f"E{total_neg_other}"].font = font_b
        ws2[f"E{total_neg_other}"].alignment = align_c
        ws2[f"E{total_neg_other}"].number_format = num_fmt

        # Минимальные ширины колонок для читаемости
        ws2.column_dimensions["A"].width = 6
        ws2.column_dimensions["B"].width = 35
        for col in ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]:
            ws2.column_dimensions[col].width = 18

        # Пунктирные границы серого цвета вокруг блоков (для читабельности)
        dotted_side = Side(border_style="dotted", color="A0A0A0")
        dotted_border = Border(left=dotted_side, right=dotted_side, top=dotted_side, bottom=dotted_side)

        def _dotted_box(r1: int, r2: int, c1: int, c2: int) -> None:
            r1 = max(1, int(r1))
            r2 = max(1, int(r2))
            c1 = max(1, int(c1))
            c2 = max(1, int(c2))
            for rr in range(r1, r2 + 1):
                for cc in range(c1, c2 + 1):
                    ws2.cell(row=rr, column=cc).border = dotted_border

        last_used_col = start_col_perf + len(months) - 1
        # Заказчики: + и - блоки (включая строки ТОП/прочее/ИТОГО)
        _dotted_box(cust_pos_start, cust_pos_total, 2, last_used_col)
        _dotted_box(cust_neg_start, cust_neg_total, 2, last_used_col)
        # Поставщики
        _dotted_box(supp_header_row - 1, supp_pos_total, 2, last_used_col)
        _dotted_box(supp_neg_start, supp_neg_total, 2, last_used_col)

        # Чёрные разделители колонок: после "оборот КТ" и вокруг "Нетто" (для всех 4 таблиц)
        black = Side(border_style="thin", color="000000")

        def _apply_vline(row_from: int, row_to: int, col: int, which: str) -> None:
            for rr in range(int(row_from), int(row_to) + 1):
                cell = ws2.cell(row=rr, column=int(col))
                b = cell.border or Border()
                if which == "right":
                    cell.border = Border(left=b.left, right=black, top=b.top, bottom=b.bottom)
                else:
                    cell.border = Border(left=black, right=b.right, top=b.top, bottom=b.bottom)

        # Столбцы: G (оборот КТ) -> правый разделитель; J (Нетто) -> левый и правый.
        col_after_turn_kt = 7   # G
        col_net = 10            # J
        for r1, r2 in [
            (cust_pos_start - 1, cust_pos_total),
            (cust_neg_hdr, cust_neg_total),
            (supp_header_row - 1, supp_pos_total),
            (supp_neg_hdr, supp_neg_total),
        ]:
            _apply_vline(r1, r2, col_after_turn_kt, "right")
            _apply_vline(r1, r2, col_net, "left")
            _apply_vline(r1, r2, col_net, "right")

        def _apply_hline(col_from: int, col_to: int, row: int, which: str) -> None:
            for cc in range(int(col_from), int(col_to) + 1):
                cell = ws2.cell(row=int(row), column=int(cc))
                b = cell.border or Border()
                if which == "bottom":
                    cell.border = Border(left=b.left, right=b.right, top=b.top, bottom=black)
                else:
                    cell.border = Border(left=b.left, right=b.right, top=black, bottom=b.bottom)

        # Горизонтальные чёрные разделители между Нетто ДЗ и Нетто КЗ
        _apply_hline(2, last_used_col, cust_pos_total, "bottom")
        _apply_hline(2, last_used_col, cust_neg_hdr, "top")
        _apply_hline(2, last_used_col, supp_pos_total, "bottom")
        _apply_hline(2, last_used_col, supp_neg_hdr, "top")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# =========================
# CODE 2 (Контракты) — несколько компаний через префиксы (Wd/Md, регистр не важен)
# Создаём лист на каждый префикс: "<префикс>контр" или "контр" (если префикса нет)
# =========================
def run_code_2(file_bytes: bytes) -> bytes:
    wb = load_workbook(io.BytesIO(file_bytes))

    prefix_to_pair: Dict[str, Dict[str, str]] = defaultdict(dict)
    for sh in wb.sheetnames:
        prefix, suf2 = split_prefix_suffix2(sh)
        if suf2 in ("wd", "md"):
            prefix = normalize_prefix(prefix)
            if suf2 not in prefix_to_pair[prefix]:
                prefix_to_pair[prefix][suf2] = sh

    valid_prefixes = [p for p, d in prefix_to_pair.items() if "wd" in d and "md" in d]
    if not valid_prefixes:
        raise ValueError("Код 2: не найдено пар листов (Wd/Md) по префиксам.")

    source_sheets_to_delete = set()
    for prefix in valid_prefixes:
        md_name = prefix_to_pair[prefix]["md"]
        wd_name = prefix_to_pair[prefix]["wd"]
        md_ws = wb[md_name]
        wd_ws = wb[wd_name]
        source_sheets_to_delete.add(md_name)
        source_sheets_to_delete.add(wd_name)

        def _as_int_year(v) -> Optional[int]:
            if v is None:
                return None
            if isinstance(v, (int, float)) and float(v).is_integer():
                y = int(v)
                return y if 1900 <= y <= 2500 else None
            s = str(v).strip()
            m = re.search(r"(\d{4})", s)
            if not m:
                return None
            try:
                y = int(m.group(1))
                return y if 1900 <= y <= 2500 else None
            except Exception:
                return None

        def _scan_month_headers(sheet) -> List[Tuple[int, str]]:
            """Ищет заголовки вида YYYY_MM в 1-й строке и возвращает [(col_idx, label), ...] по порядку колонок."""
            out: List[Tuple[int, str]] = []
            max_col = int(sheet.max_column or 1)
            for c in range(1, max_col + 1):
                v = sheet.cell(row=1, column=c).value
                if v is None:
                    continue
                s = str(v).strip()
                if re.fullmatch(r"\d{4}_\d{2}", s):
                    out.append((c, s))
            return out

        years_pay = [_as_int_year(wd_ws["C1"].value), _as_int_year(wd_ws["D1"].value), _as_int_year(wd_ws["E1"].value)]
        years_perf = [_as_int_year(md_ws["C1"].value), _as_int_year(md_ws["D1"].value), _as_int_year(md_ws["E1"].value)]
        years_pay = [y for y in years_pay if y is not None]
        years_perf = [y for y in years_perf if y is not None]

        # По текущей структуре "Договоры" ожидаем 3 годовых колонки (C/D/E). Если что-то не так — не ломаемся,
        # просто подставляем "пустые" годы для сохранения структуры.
        while len(years_pay) < 3:
            years_pay.append(0)
        while len(years_perf) < 3:
            years_perf.append(0)
        years_pay = years_pay[:3]
        years_perf = years_perf[:3]

        payments_year = defaultdict(lambda: [0.0, 0.0, 0.0])
        performance_year = defaultdict(lambda: [0.0, 0.0, 0.0])

        pay_month_cols = _scan_month_headers(wd_ws)
        perf_month_cols = _scan_month_headers(md_ws)
        payments_monthly = defaultdict(lambda: [0.0] * len(pay_month_cols))
        performance_monthly = defaultdict(lambda: [0.0] * len(perf_month_cols))

        def collect_yearly(sheet, target_dict):
            for row in range(2, sheet.max_row + 1):
                n = sheet[f"A{row}"].value
                c = sheet[f"B{row}"].value
                if not n and not c:
                    continue
                key = (str(n).strip() if n else "", str(c).strip() if c else "")
                for idx, col in enumerate(["C", "D", "E"]):
                    v = sheet[f"{col}{row}"].value
                    if v is None:
                        continue
                    try:
                        target_dict[key][idx] += float(v)
                    except:
                        pass

        def collect_monthly(sheet, target_dict, month_cols: List[Tuple[int, str]]):
            for row in range(2, sheet.max_row + 1):
                n = sheet[f"A{row}"].value
                c = sheet[f"B{row}"].value
                if not n and not c:
                    continue
                key = (str(n).strip() if n else "", str(c).strip() if c else "")
                for i, (col_idx, _label) in enumerate(month_cols):
                    v = sheet.cell(row=row, column=col_idx).value
                    if v is None:
                        continue
                    try:
                        target_dict[key][i] += float(v)
                    except:
                        pass

        collect_yearly(wd_ws, payments_year)
        collect_yearly(md_ws, performance_year)
        collect_monthly(wd_ws, payments_monthly, pay_month_cols)
        collect_monthly(md_ws, performance_monthly, perf_month_cols)

        all_keys = sorted(
            set(payments_year.keys())
            | set(performance_year.keys())
            | set(payments_monthly.keys())
            | set(performance_monthly.keys()),
            key=lambda x: (x[0], x[1]),
        )

        out_sheet_name = safe_sheet_name(f"{prefix}контр" if prefix else "контр")
        if out_sheet_name in wb.sheetnames:
            del wb[out_sheet_name]
        ws = wb.create_sheet(out_sheet_name)

        ws["A1"] = "ИТОГО в тыс тенге"
        ws["A2"] = "Контрагент"
        ws["B2"] = "Договор"

        ws["C1"] = "оплата"
        ws["C2"] = years_pay[0] or ""
        ws["D2"] = years_pay[1] or ""
        ws["E2"] = years_pay[2] or ""
        ws["F2"] = "Total"

        ws["G1"] = "выполнения с ндс"
        ws["G2"] = years_perf[0] or ""
        ws["H2"] = years_perf[1] or ""
        ws["I2"] = years_perf[2] or ""
        ws["J2"] = "Total"

        ws["K2"] = "дз/(аванс)"

        ws["M1"] = "оплата"
        start_col_pay = column_index_from_string("M")
        pay_month_labels = [lbl for _c, lbl in pay_month_cols]
        for i, label in enumerate(pay_month_labels):
            ws[f"{get_column_letter(start_col_pay + i)}2"] = label

        start_col_perf = start_col_pay + max(1, len(pay_month_labels))
        ws[f"{get_column_letter(start_col_perf)}1"] = "выполнения с ндс"
        perf_month_labels = [lbl for _c, lbl in perf_month_cols]
        for i, label in enumerate(perf_month_labels):
            ws[f"{get_column_letter(start_col_perf + i)}2"] = label

        start_row = 3
        row = start_row
        for key in all_keys:
            name, contract = key

            py = payments_year.get(key, [0, 0, 0])
            pf = performance_year.get(key, [0, 0, 0])

            # Убираем строки, где одновременно:
            # - F (итого оплаты) = 0
            # - J (итого выполнения) = 0
            # Проверяем по исходным суммам (а не по формуле), чтобы не зависеть от пересчёта Excel.
            if abs(float(py[0]) + float(py[1]) + float(py[2])) < 1e-9 and abs(float(pf[0]) + float(pf[1]) + float(pf[2])) < 1e-9:
                continue

            ws[f"A{row}"] = name
            ws[f"B{row}"] = contract

            ws[f"C{row}"] = py[0]
            ws[f"D{row}"] = py[1]
            ws[f"E{row}"] = py[2]

            def _vat_coef_for_year(y: int) -> float:
                # НДС 12% до 2025 включительно, 16% начиная с 2026 (и для лет > 2026 тоже 16%).
                return 1.16 if int(y or 0) >= 2026 else 1.12

            ws[f"G{row}"] = pf[0] * _vat_coef_for_year(years_perf[0])
            ws[f"H{row}"] = pf[1] * _vat_coef_for_year(years_perf[1])
            ws[f"I{row}"] = pf[2] * _vat_coef_for_year(years_perf[2])

            ws[f"F{row}"] = f"=SUM(C{row}:E{row})"
            ws[f"J{row}"] = f"=SUM(G{row}:I{row})"
            ws[f"K{row}"] = f"=J{row}-F{row}"

            mp = payments_monthly.get(key, [0.0] * len(pay_month_labels))
            for i in range(len(pay_month_labels)):
                ws[f"{get_column_letter(start_col_pay + i)}{row}"] = mp[i]

            mf = performance_monthly.get(key, [0.0] * len(perf_month_labels))
            for i, label in enumerate(perf_month_labels):
                y = _as_int_year(label.split("_", 1)[0]) or 0
                ws[f"{get_column_letter(start_col_perf + i)}{row}"] = mf[i] * _vat_coef_for_year(y)

            row += 1

        last_row = row - 1 if row > start_row else 2

        regular = Font(name="Arial", size=10)
        bold = Font(name="Arial", size=10, bold=True)

        last_col = max(start_col_perf + len(perf_month_labels) - 1, column_index_from_string("K"))
        for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=last_col):
            for c in row:
                c.font = regular

        for col in range(1, last_col + 1):
            ws[f"{get_column_letter(col)}2"].font = bold
        for addr in ["A1", "C1", "G1", "M1", f"{get_column_letter(start_col_perf)}1", "K2"]:
            ws[addr].font = bold

        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")
        for col in range(1, last_col + 1):
            addr = f"{get_column_letter(col)}2"
            ws[addr].alignment = left if addr in ["A2", "B2"] else center

        num_format = "#,##0;[Red](#,##0)"
        numeric_cols = list("CDEFGHIJK")
        numeric_cols += [get_column_letter(c) for c in range(start_col_pay, start_col_pay + len(pay_month_labels))]
        numeric_cols += [get_column_letter(c) for c in range(start_col_perf, start_col_perf + len(perf_month_labels))]

        for col in numeric_cols:
            for r in range(3, last_row + 1):
                cell = ws[f"{col}{r}"]
                cell.alignment = center
                cell.number_format = num_format

        for r in range(1, last_row + 1):
            ws[f"A{r}"].alignment = left
            ws[f"B{r}"].alignment = left

        for addr in ["C1", "G1", "M1"]:
            try:
                ws[addr].alignment = left
            except Exception:
                pass
        try:
            ws[f"{get_column_letter(start_col_perf)}1"].alignment = left
        except Exception:
            pass

        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 38
        for col in numeric_cols:
            ws.column_dimensions[col].width = 12.2 if column_index_from_string(col) >= start_col_pay else 12.6

        thin = Side(border_style="thin", color="000000")
        border_cols = ["C", "G", "K", "M", get_column_letter(start_col_perf)]
        for r in range(1, last_row + 1):
            for col in border_cols:
                cell = ws[f"{col}{r}"]
                cell.border = Border(
                    left=thin,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom,
                )

    # После формирования листа(ов) "контр" удаляем исходные листы Wd/Md из выходного файла.
    for sh in sorted(source_sheets_to_delete):
        if sh in wb.sheetnames:
            del wb[sh]

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# =========================
# CODE 3 (Запасы)
# Для каждого выбранного счета 13**:
# - находим лист(ы), где в названии встречается номер счета
# - в листе ищем строку в колонке A, где ячейка равна номеру счета
# - ставим I1=400 и суммарную "шапку" корзин в K6:O7 (как в эталонном файле)
# - в колонке I проставляем формулу-классификацию по строкам до первой пустой в колонке G
# Отдельных "отчётных" листов не создаём: UI показывает предупреждения по результату.
# =========================
def run_code_3_inventory(file_bytes: bytes, accounts: List[str]) -> Tuple[bytes, Dict[str, List[str]]]:
    wb = load_workbook(io.BytesIO(file_bytes))

    processed: List[str] = []
    missing_sheets: List[str] = []
    missing_markers: List[str] = []

    def _is_blank(v) -> bool:
        return v is None or (isinstance(v, str) and v.strip() == "")

    def _find_account_row(ws, account: str) -> Optional[int]:
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if _is_blank(v):
                continue
            if isinstance(v, (int, float)) and float(v).is_integer():
                s = str(int(v))
            else:
                s = str(v).strip()
            if s == account:
                return r
        return None

    def _apply_inventory_bucket_layout(ws) -> None:
        """
        Новая методика ликвидности запасов (как в эталонном файле):
        - I1 = 400
        - K6:O6 — корзины
        - K7:O7 — суммы по корзинам через SUMIFS (G по значениям I)
        """
        ws["I1"].value = 400
        ws["I1"].number_format = "#,##0"
        ws["I1"].font = Font(name="Arial", size=8, bold=True, color="FF4B5563")  # тёмно-серый
        ws["I1"].fill = PatternFill(fill_type="solid", fgColor="FFD9D9D9")  # светло-серый фон
        ws["I1"].alignment = Alignment(horizontal="center")

        labels = [
            "0-1 мес",
            "1-3 мес",
            "3-6 мес",
            "6-12 мес",
            "ДТ BoP = 0",
            "более 12 мес",
            "Бесконечность (ост нач)",
            "Бесконечность (обор ДТ)",
        ]
        cols = ["K", "L", "M", "N", "O", "P", "Q", "R"]

        fill_hdr = PatternFill(fill_type="solid", fgColor="E4F0DD")
        font_hdr = Font(name="Aptos Narrow", size=9, bold=True, color="003F2F")
        align_hdr = Alignment(horizontal="center", vertical="top", wrap_text=True)
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col, label in zip(cols, labels):
            c = ws[f"{col}6"]
            c.value = label
            c.font = font_hdr
            c.fill = fill_hdr
            c.alignment = align_hdr
            c.border = border
            c.number_format = "#,##0.00"

        font_sum = Font(name="Aptos Narrow", size=9, bold=True)
        align_sum = Alignment(horizontal="center")
        for col in cols:
            c = ws[f"{col}7"]
            c.value = f"=SUMIFS($G:$G,$I:$I,{col}6)"
            c.font = font_sum
            c.alignment = align_sum
            c.number_format = "#,##0"

    def _bucket_formula_for_row(r: int) -> str:
        return (
            f'=IF(G{r}=0,\"Нет остатка\",\n'
            f' IF(G{r}<0,\"Отрицательное\",\n'
            f' IF(C{r}=0,\"ДТ BoP = 0\",\n'
            f' IF(AND(F{r}=0,C{r}<>0,E{r}=0,G{r}=C{r}),\"Бесконечность (ост нач)\",\n'
            f' IF(AND(F{r}=0,C{r}<>0,E{r}<>0),\"Бесконечность (обор ДТ)\",\n'
            f' _xlfn.LET(_xlpm.turn,$I$1*(G{r}/F{r}),\n'
            f'     IF(_xlpm.turn<=31,\"0-1 мес\",\n'
            f'        IF(_xlpm.turn<=92,\"1-3 мес\",\n'
            f'           IF(_xlpm.turn<=183,\"3-6 мес\",\n'
            f'              IF(_xlpm.turn<=366,\"6-12 мес\",\"более 12 мес\"))))))))))'
        )

    for account in accounts:
        matched_sheets = [sh for sh in wb.sheetnames if account in sh]
        if not matched_sheets:
            missing_sheets.append(account)
            continue

        for sh in matched_sheets:
            ws = wb[sh]
            found_row = _find_account_row(ws, account)
            if found_row is None:
                missing_markers.append(f"{account}: {sh}")
                continue

            _apply_inventory_bucket_layout(ws)

            # Заполняем формулы по строкам запасов: от строки ниже маркера до первой пустой в колонке G.
            start_row = found_row + 1
            last = start_row - 1
            for r in range(start_row, ws.max_row + 1):
                # Останавливаемся перед строкой "Итого" (чтобы не тянуть хвосты/сводные строки в запасы).
                a = ws.cell(row=r, column=1).value
                if a is not None and "итого" in str(a).strip().lower():
                    break
                g = ws.cell(row=r, column=7).value
                if _is_blank(g):
                    break
                last = r

            if last >= start_row:
                for rr in range(start_row, last + 1):
                    c = ws.cell(row=rr, column=9)  # I
                    c.value = _bucket_formula_for_row(rr)
                    c.font = Font(name="Arial", size=8)
                processed.append(f"{account}: {sh} (строки {start_row}-{last})")
            else:
                processed.append(f"{account}: {sh} (нет строк с данными в G ниже маркера)")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), {"processed": processed, "missing_sheets": missing_sheets, "missing_markers": missing_markers}


# =========================
# CODE 4 (Чистая прибыль по "общ*")
# Рассчитывает ЧП из движения денег (1030), исключая CFF/CFI/NWC/Амортизацию.
# Для каждого выбранного листа "общ*" добавляет справа-внизу таблицу (формулами INDEX/MATCH).
# =========================
def run_code_4_net_profit(file_bytes: bytes, obsh_sheetnames: List[str]) -> bytes:
    wb = load_workbook(io.BytesIO(file_bytes))

    if not obsh_sheetnames:
        raise ValueError("Чистая прибыль: не выбраны листы «общ*».")

    for sh in obsh_sheetnames:
        if sh not in wb.sheetnames:
            raise ValueError(f"Чистая прибыль: лист не найден: {sh}")

    def _is_blank(v) -> bool:
        if v is None:
            return True
        if isinstance(v, str) and not v.strip():
            return True
        return False

    def _excel_sheet_ref(title: str) -> str:
        # Excel-референс листа в формулах: 'Лист 1'!A1. Одинарные кавычки внутри имени удваиваем.
        t = str(title).replace("'", "''")
        return f"'{t}'"

    def _find_last_used_col_in_header(ws) -> int:
        maxc = 1
        top = min(int(ws.max_row or 1), 12)
        for r in range(1, top + 1):
            for c in range(1, int(ws.max_column or 1) + 1):
                v = ws.cell(row=r, column=c).value
                if not _is_blank(v):
                    maxc = max(maxc, c)
        return maxc

    def _find_last_used_row_in_col_a(ws) -> int:
        max_scan = min(int(ws.max_row or 1), 10000)
        last = 1
        for r in range(1, max_scan + 1):
            if not _is_blank(ws.cell(row=r, column=1).value):
                last = r
        return last

    def _find_debit_credit_pairs(ws) -> Tuple[int, int, int, int, int, int]:
        """
        Возвращает (bop_deb, bop_cred, turn_deb, turn_cred, eop_deb, eop_cred) — номера колонок.
        Ищем строку заголовков, где встречаются "Дебет"/"Кредит" 3 раза подряд (6 колонок).
        """
        # Ищем по верхним строкам (обычно 4..8).
        best_row = None
        for r in range(1, 20):
            deb_cols = []
            cred_cols = []
            for c in range(1, int(ws.max_column or 1) + 1):
                v = ws.cell(row=r, column=c).value
                if v is None:
                    continue
                s = str(v).strip().lower()
                if "дебет" == s or s.startswith("дебет"):
                    deb_cols.append(c)
                if "кредит" == s or s.startswith("кредит"):
                    cred_cols.append(c)
            # Кандидат: минимум 3 дебета и 3 кредита, и они идут вперемешку по возрастанию.
            if len(deb_cols) >= 3 and len(cred_cols) >= 3:
                best_row = r
                break

        if best_row is None:
            raise ValueError(f"Чистая прибыль ({ws.title}): не нашёл заголовки «Дебет/Кредит» (3 пары) в верхней части листа.")

        labels = []
        for c in range(1, int(ws.max_column or 1) + 1):
            v = ws.cell(row=best_row, column=c).value
            if v is None:
                continue
            s = str(v).strip().lower()
            if s.startswith("дебет"):
                labels.append((c, "D"))
            elif s.startswith("кредит"):
                labels.append((c, "C"))
        labels.sort(key=lambda x: x[0])

        # Берём первые 6 в формате D,C,D,C,D,C
        seq = []
        for c, t in labels:
            if not seq:
                if t == "D":
                    seq.append((c, t))
            else:
                want = "C" if seq[-1][1] == "D" else "D"
                if t == want:
                    seq.append((c, t))
            if len(seq) == 6:
                break

        if len(seq) < 6 or [t for _c, t in seq] != ["D", "C", "D", "C", "D", "C"]:
            raise ValueError(f"Чистая прибыль ({ws.title}): не смог определить 3 пары «Дебет/Кредит» из заголовков.")

        bop_deb, bop_cred, turn_deb, turn_cred, eop_deb, eop_cred = [c for c, _t in seq]
        return bop_deb, bop_cred, turn_deb, turn_cred, eop_deb, eop_cred

    def _find_account_row(ws, acc: str) -> Optional[int]:
        max_scan = min(int(ws.max_row or 1), 10000)
        for r in range(3, max_scan + 1):
            v = ws.cell(row=r, column=1).value
            if v is None:
                continue
            if acc in str(v):
                return r
        return None

    def _name_or_acc(ws, row: Optional[int], acc: str) -> str:
        if row is None:
            return acc
        v = ws.cell(row=row, column=1).value
        return str(v).strip() if v is not None else acc

    def _delta_formula(sheet_ref: str, acc_row: int, cols: Tuple[int, int, int, int, int, int]) -> str:
        bop_d, bop_c, _t_d, _t_c, eop_d, eop_c = cols
        # Δ = (Кт_кон - Кт_нач) - (Дт_кон - Дт_нач)
        bd = f"INDEX({sheet_ref}!$1:$1048576,{acc_row},{bop_d})"
        bc = f"INDEX({sheet_ref}!$1:$1048576,{acc_row},{bop_c})"
        ed = f"INDEX({sheet_ref}!$1:$1048576,{acc_row},{eop_d})"
        ec = f"INDEX({sheet_ref}!$1:$1048576,{acc_row},{eop_c})"
        return f"=(({ec})-({bc}))-(({ed})-({bd}))"

    thin_black = Side(style="thin", color="000000")
    dotted_gray = Side(style="dotted", color="A0A0A0")
    header_fill = PatternFill("solid", fgColor="EDEDED")
    font_h = Font(name="Aptos Narrow", size=9, bold=True)
    font_b = Font(name="Aptos Narrow", size=9)
    num_fmt = "#,##0; -#,##0"

    for ws_name in obsh_sheetnames:
        ws = wb[ws_name]

        cols = _find_debit_credit_pairs(ws)
        last_row = _find_last_used_row_in_col_a(ws)
        start_row = last_row + 3
        # Таблицу ставим на 2 столбца правее ПОСЛЕДНЕГО столбца "Сальдо на конец (Кредит)"
        # (то есть правее третьей пары "Дебет/Кредит").
        start_col = int(cols[5]) + 2
        sheet_ref = _excel_sheet_ref(ws.title)

        # Заголовки таблицы
        headers = ["Счет", "Полный", "ВГО", "Без ВГО"]
        for j, h in enumerate(headers):
            cell = ws.cell(row=start_row, column=start_col + j, value=h)
            cell.font = font_h
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(left=dotted_gray, right=dotted_gray, top=dotted_gray, bottom=dotted_gray)

        # Строки (в таком порядке, как в ТЗ)
        # Для каждой "строки счета": если счет найден — пишем полное название, иначе — 4 цифры.
        # Для "итогов": CF/CFF/CFI/NWC/ЧП — считаем формулой по значениям выше.
        row = start_row + 1

        def _write_line(label: str, full_text: str, full_formula: Optional[str]) -> int:
            nonlocal row
            ws.cell(row=row, column=start_col + 0, value=full_text).font = font_b

            c_full = ws.cell(row=row, column=start_col + 1, value=(full_formula or 0))
            c_full.font = font_b
            c_full.number_format = num_fmt

            ws.cell(row=row, column=start_col + 2, value="").font = font_b  # ВГО пустой

            c_no = ws.cell(row=row, column=start_col + 3, value=f"=({get_column_letter(start_col+1)}{row})-({get_column_letter(start_col+2)}{row})")
            c_no.font = font_b
            c_no.number_format = num_fmt

            # Пунктирные границы вокруг строки
            for j in range(4):
                ws.cell(row=row, column=start_col + j).border = Border(left=dotted_gray, right=dotted_gray, top=dotted_gray, bottom=dotted_gray)
            cur = row
            row += 1
            return cur

        def _write_sum_line(text: str, ref_rows: List[int]) -> int:
            nonlocal row
            ws.cell(row=row, column=start_col + 0, value=text).font = font_h
            # Полный
            letters = [f"{get_column_letter(start_col+1)}{r}" for r in ref_rows]
            ws.cell(row=row, column=start_col + 1, value=("=SUM(" + ",".join(letters) + ")") if letters else 0).number_format = num_fmt
            ws.cell(row=row, column=start_col + 1).font = font_h
            # ВГО пустой
            ws.cell(row=row, column=start_col + 2, value="").font = font_h
            # Без ВГО
            letters2 = [f"{get_column_letter(start_col+3)}{r}" for r in ref_rows]
            ws.cell(row=row, column=start_col + 3, value=("=SUM(" + ",".join(letters2) + ")") if letters2 else 0).number_format = num_fmt
            ws.cell(row=row, column=start_col + 3).font = font_h
            for j in range(4):
                ws.cell(row=row, column=start_col + j).border = Border(left=dotted_gray, right=dotted_gray, top=dotted_gray, bottom=dotted_gray)
            cur = row
            row += 1
            return cur

        def _top_border(rr: int) -> None:
            for j in range(4):
                cell = ws.cell(row=rr, column=start_col + j)
                cell.border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=thin_black,
                    bottom=cell.border.bottom,
                )

        # 1030
        r1030 = _find_account_row(ws, "1030")
        if r1030:
            base = _delta_formula(sheet_ref, r1030, cols)
            neg = f"=-({base[1:]})" if isinstance(base, str) and base.startswith("=") else f"=-({base})"
        else:
            neg = None
        r_1030_line = _write_line("1030", _name_or_acc(ws, r1030, "1030"), neg)
        r_cf = _write_sum_line("CF", [r_1030_line])
        _top_border(r_cf)

        row += 1  # пропуск строки

        # CFF: 3010 + 3030 + 4010
        r3010 = _find_account_row(ws, "3010")
        r3030 = _find_account_row(ws, "3030")
        r4010 = _find_account_row(ws, "4010")
        r_3010_line = _write_line("3010", _name_or_acc(ws, r3010, "3010"), _delta_formula(sheet_ref, r3010, cols) if r3010 else None)
        r_3030_line = _write_line("3030", _name_or_acc(ws, r3030, "3030"), _delta_formula(sheet_ref, r3030, cols) if r3030 else None)
        r_4010_line = _write_line("4010", _name_or_acc(ws, r4010, "4010"), _delta_formula(sheet_ref, r4010, cols) if r4010 else None)
        r_cff = _write_sum_line("CFF", [r_3010_line, r_3030_line, r_4010_line])
        _top_border(r_cff)

        row += 1  # пропуск строки

        # CFI: 2410
        r2410 = _find_account_row(ws, "2410")
        r_2410_line = _write_line("2410", _name_or_acc(ws, r2410, "2410"), _delta_formula(sheet_ref, r2410, cols) if r2410 else None)
        r_cfi = _write_sum_line("CFI", [r_2410_line])
        _top_border(r_cfi)

        row += 1

        # NWC: 1210,1310,1320,1330,1400,1710,1711,3100,3310,3311,3320,3350,3510
        nwc_rows = []
        for acc in ["1210", "1310", "1320", "1330", "1400", "1710", "1711", "3100", "3310", "3311", "3320", "3350", "3510"]:
            rr = _find_account_row(ws, acc)
            nwc_rows.append(_write_line(acc, _name_or_acc(ws, rr, acc), _delta_formula(sheet_ref, rr, cols) if rr else None))
        r_nwc = _write_sum_line("NWC", nwc_rows)
        _top_border(r_nwc)

        row += 1

        # Амортизация: 2420
        r2420 = _find_account_row(ws, "2420")
        r_am_line = _write_line("2420", _name_or_acc(ws, r2420, "2420"), _delta_formula(sheet_ref, r2420, cols) if r2420 else None)

        row += 1

        # ЧП = CF - CFF - CFI - NWC - Аморт
        # Делается формулами по строкам итогов.
        ws.cell(row=row, column=start_col + 0, value="ЧП").font = font_h
        ws.cell(
            row=row,
            column=start_col + 1,
            value=f"=({get_column_letter(start_col+1)}{r_cf})-({get_column_letter(start_col+1)}{r_cff})-({get_column_letter(start_col+1)}{r_cfi})-({get_column_letter(start_col+1)}{r_nwc})-({get_column_letter(start_col+1)}{r_am_line})",
        ).font = font_h
        ws.cell(row=row, column=start_col + 1).number_format = num_fmt
        ws.cell(row=row, column=start_col + 2, value="").font = font_h
        ws.cell(
            row=row,
            column=start_col + 3,
            value=f"=({get_column_letter(start_col+3)}{r_cf})-({get_column_letter(start_col+3)}{r_cff})-({get_column_letter(start_col+3)}{r_cfi})-({get_column_letter(start_col+3)}{r_nwc})-({get_column_letter(start_col+3)}{r_am_line})",
        ).font = font_h
        ws.cell(row=row, column=start_col + 3).number_format = num_fmt
        for j in range(4):
            ws.cell(row=row, column=start_col + j).border = Border(left=dotted_gray, right=dotted_gray, top=thin_black, bottom=dotted_gray)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()




# =========================
# CODE 5 (Инсайты)
# Генерирует лист(ы) «инсайты» на основе W/M/Wt/Mt (и общ/кред при наличии).
# Реализовано в отдельном модуле `insights.py`.
# =========================

def run_code_5_insights(file_bytes: bytes) -> bytes:
    """Генерация листа(ов) «инсайты» (см. модуль `insights.py`)."""
    from insights import generate_insights

    return generate_insights(file_bytes)


# =========================
# CODE 6 (Госы)
# Для каждого выбранного листа «гос*»:
#   - находит столбцы «Номер договора» и «Статус» в строке 1
#   - по каждой строке со статусом «Действует» запрашивает данные из API goszakup
#   - записывает результат правее существующей таблицы
# =========================

_GOS_BASE_OWS = "https://ows.goszakup.gov.kz"
_GOS_BASE_WEB = "https://www.goszakup.gov.kz"

_GOS_NEW_COLS = [
    "Ссылка",
    "Многолетний",
    "Год окончания",
    "Сумма 2026",
    "Сумма 2027",
    "% аванса",
    "Сумма аванса",
    "Обеспечение аванса требуется",
    "Обеспечение аванса прикреплено",
    "Факт сумма",
]


def _gos_money(x) -> float:
    if x in [None, "", "null"]:
        return 0.0
    try:
        return float(Decimal(str(x)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
    except Exception:
        return 0.0


def _gos_parse_money(text: str) -> float:
    if text is None:
        return 0.0
    text = str(text).replace("\xa0", " ").replace(" ", "").replace(",", ".")
    text = re.sub(r"[^\d.]", "", text)
    return _gos_money(text) if text else 0.0


def _gos_clean_text(text: str) -> str:
    if not text:
        return ""
    text = text.replace("\xa0", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n+", "\n", text)
    return text.strip()


def _gos_html_to_text(html: str) -> str:
    try:
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(html, "html.parser")
        return _gos_clean_text(soup.get_text("\n", strip=True))
    except Exception:
        text = re.sub(r"<script.*?</script>", " ", html, flags=re.S | re.I)
        text = re.sub(r"<style.*?</style>", " ", text, flags=re.S | re.I)
        text = re.sub(r"<[^>]+>", "\n", text)
        return _gos_clean_text(text)


def _gos_get_year_from_date(date_str):
    if not date_str:
        return None
    try:
        return int(str(date_str)[:4])
    except Exception:
        return None


def _gos_get_max_end_year(contract: Dict[str, Any]) -> Optional[int]:
    years = []
    for field in ["contract_end_date", "plan_exec_date", "ec_end_date"]:
        y = _gos_get_year_from_date(contract.get(field))
        if y:
            years.append(y)
    return max(years) if years else None


def _gos_get_additions_url(contract: Dict[str, Any]) -> str:
    additions_id = contract.get("root_id") or contract.get("id")
    if not additions_id:
        return ""
    return f"{_GOS_BASE_WEB}/ru/egzcontract/cpublic/additions/{additions_id}"


def _gos_get_json(url: str, token: str, params: Optional[Dict[str, Any]] = None) -> Any:
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
        "Accept": "application/json",
        "User-Agent": "Mozilla/5.0",
    }
    r = _requests_lib.get(url, headers=headers, params=params, timeout=40)
    if r.status_code == 401:
        raise Exception("401: неверный или просроченный токен")
    if r.status_code == 403:
        raise Exception(f"403: нет доступа к API: {url}")
    if r.status_code == 404:
        raise Exception(f"404: не найдено: {url}")
    r.raise_for_status()
    return r.json()


def _gos_get_html_text(url: str) -> str:
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    }
    r = _requests_lib.get(url, headers=headers, timeout=40)
    r.raise_for_status()
    return _gos_html_to_text(r.text)


def _gos_unwrap_contract(data: Any) -> Dict[str, Any]:
    if isinstance(data, dict) and "data" in data:
        data = data["data"]
    if isinstance(data, list):
        if not data:
            raise Exception("Договор не найден")
        return data[0]
    if isinstance(data, dict):
        return data
    raise Exception(f"Неожиданный формат ответа: {type(data)}")


def _gos_get_contract_by_number(contract_number: str, token: str) -> Dict[str, Any]:
    if "/" in contract_number:
        url = f"{_GOS_BASE_OWS}/v3/contract/number-sys/"
    else:
        url = f"{_GOS_BASE_OWS}/v3/contract/number/"
    data = _gos_get_json(url, token, params={"number": contract_number})
    return _gos_unwrap_contract(data)


def _gos_get_contract_units(contract_id: int, token: str) -> Dict[str, Any]:
    url = f"{_GOS_BASE_OWS}/v3/contract/{contract_id}/units"
    return _gos_get_json(url, token)


def _gos_get_plan_point(pln_point_id: int, token: str) -> Optional[Dict[str, Any]]:
    url = f"{_GOS_BASE_OWS}/v3/plans/view/{pln_point_id}"
    try:
        data = _gos_get_json(url, token)
        if isinstance(data, dict) and "data" in data:
            return data["data"]
        return data
    except Exception:
        return None


def _gos_extract_current_block(additions_text: str, contract_number: str) -> str:
    start_pattern = re.escape(contract_number) + r"\s+Статус договора:"
    start_match = re.search(start_pattern, additions_text)
    if not start_match:
        return additions_text
    start = start_match.start()
    next_match = re.search(r"\n\d{12}/\d{6}/\d{2}\s+Статус договора:", additions_text[start_match.end():])
    if next_match:
        return additions_text[start: start_match.end() + next_match.start()]
    footer_patterns = ["АИИС ЭГЗ", "Техническая поддержка", "Чем я могу Вам помочь"]
    end = len(additions_text)
    for p in footer_patterns:
        m = re.search(p, additions_text[start:])
        if m:
            end = min(end, start + m.start())
    return additions_text[start:end]


def _gos_parse_year_amounts(block: str) -> Dict[str, float]:
    result: Dict[str, float] = {}
    pattern = r"Сумма\s+на\s+(\d{4})\s+год\s*[-:]\s*([\d\s]+(?:[.,]\d+)?)"
    for year, amount in re.findall(pattern, block, flags=re.I):
        result[year] = result.get(year, 0.0) + _gos_parse_money(amount)
    return {k: _gos_money(v) for k, v in result.items()}


def _gos_parse_advance_security(block: str) -> Dict[str, Any]:
    result: Dict[str, Any] = {"advance_security_required": None, "advance_security_attached": None}
    pattern = (
        r"Обеспечение исполнения договора на сумму аванса по договору\s*"
        r"Требуется:\s*(Да|Нет)\s*,\s*Прикреплен:\s*(Да|Нет)"
    )
    m = re.search(pattern, block, flags=re.I)
    if m:
        result["advance_security_required"] = 1 if m.group(1).lower() == "да" else 0
        result["advance_security_attached"] = 1 if m.group(2).lower() == "да" else 0
    return result


def _gos_parse_totals(block: str) -> Dict[str, Any]:
    result: Dict[str, Any] = {"contract_sum": None, "fact_sum": None}
    m = re.search(r"Общая итоговая сумма договора\s*:\s*([\d\s]+(?:[.,]\d+)?)", block, flags=re.I)
    if m:
        result["contract_sum"] = _gos_parse_money(m.group(1))
    m = re.search(r"Общая фактическая сумма договора\s*:\s*([\d\s]+(?:[.,]\d+)?)", block, flags=re.I)
    if m:
        result["fact_sum"] = _gos_parse_money(m.group(1))
    return result


def _gos_extract_advance_percent(contract: Dict[str, Any], token: str) -> Dict[str, Any]:
    advance_percent = 0.0
    unresolved: List[int] = []
    units = contract.get("units", {}).get("items", []) or []
    for unit in units:
        pln_point_id = unit.get("pln_point_id")
        if not pln_point_id:
            continue
        pp = _gos_get_plan_point(pln_point_id, token)
        if not pp:
            unresolved.append(pln_point_id)
            continue
        prepayment = pp.get("prepayment")
        if prepayment in [None, "", "null"]:
            prepayment = 0
        try:
            prepayment = float(prepayment)
        except Exception:
            prepayment = 0.0
        if prepayment > advance_percent:
            advance_percent = prepayment
    if advance_percent > 0:
        return {"status": "ok", "advance_percent": _gos_money(advance_percent), "message": "", "unresolved": unresolved}
    if unresolved:
        return {
            "status": "error",
            "advance_percent": None,
            "message": f"ОШИБКА: аванс не удалось проверить. Не раскрылись plans/view: {unresolved}",
            "unresolved": unresolved,
        }
    return {"status": "ok", "advance_percent": 0.0, "message": "", "unresolved": []}


def _gos_build_row(contract_number: str, token: str) -> Dict[str, Any]:
    contract = _gos_get_contract_by_number(contract_number, token)
    contract_id = contract.get("id")
    if not contract_id:
        raise Exception("В договоре нет contract_id")
    try:
        contract["units"] = _gos_get_contract_units(contract_id, token)
    except Exception:
        contract["units"] = {"items": []}

    additions_url = _gos_get_additions_url(contract)
    is_long_term = 1 if contract.get("ref_contract_year_type_id") == 2 else 0
    contract_sum = _gos_money(contract.get("contract_sum_wnds"))
    fact_sum = _gos_money(contract.get("fakt_sum_wnds"))
    amount_2026 = 0.0
    amount_2027 = 0.0
    advance_security_required = None
    advance_security_attached = None
    max_end_year = _gos_get_max_end_year(contract)
    current_block = ""
    html_status = "not_used"
    html_error = ""

    if is_long_term == 1:
        try:
            additions_text = _gos_get_html_text(additions_url) if additions_url else ""
            current_block = _gos_extract_current_block(
                additions_text,
                contract.get("contract_number_sys") or contract_number,
            )
            amounts_by_year = _gos_parse_year_amounts(current_block)
            totals = _gos_parse_totals(current_block)
            adv_sec = _gos_parse_advance_security(current_block)
            amount_2026 = _gos_money(amounts_by_year.get("2026", 0))
            amount_2027 = 0.0 if (max_end_year is not None and max_end_year <= 2026) else _gos_money(amounts_by_year.get("2027", 0))
            if totals.get("contract_sum") is not None:
                contract_sum = _gos_money(totals["contract_sum"])
            if totals.get("fact_sum") is not None:
                fact_sum = _gos_money(totals["fact_sum"])
            advance_security_required = adv_sec["advance_security_required"]
            advance_security_attached = adv_sec["advance_security_attached"]
            html_status = "ok"
        except Exception as e:
            html_error = str(e)
            html_status = "error"
            amount_2026 = f"ОШИБКА HTML: {html_error}"
            amount_2027 = f"ОШИБКА HTML: {html_error}"
    else:
        fin_year = str(contract.get("fin_year") or "")
        if fin_year == "2026":
            amount_2026 = contract_sum
        elif fin_year == "2027":
            amount_2027 = contract_sum

    advance_data = _gos_extract_advance_percent(contract, token)
    if advance_data["status"] == "error":
        advance_percent = advance_data["message"]
        advance_sum = advance_data["message"]
        advance_security_required = advance_data["message"]
        advance_security_attached = advance_data["message"]
    else:
        advance_percent = advance_data["advance_percent"]
        if advance_percent > 0:
            advance_sum = _gos_money(contract_sum * advance_percent / 100)
            if not current_block and additions_url:
                try:
                    additions_text = _gos_get_html_text(additions_url)
                    current_block = _gos_extract_current_block(
                        additions_text,
                        contract.get("contract_number_sys") or contract_number,
                    )
                    adv_sec = _gos_parse_advance_security(current_block)
                    advance_security_required = adv_sec["advance_security_required"]
                    advance_security_attached = adv_sec["advance_security_attached"]
                    html_status = "ok"
                except Exception as e:
                    html_status = "error"
                    html_error = str(e)
            if html_status != "ok":
                msg = f"ОШИБКА: HTML не получен: {html_error}"
                advance_security_required = msg
                advance_security_attached = msg
            elif advance_security_required is None:
                advance_security_required = "ОШИБКА: строка обеспечения не найдена"
            elif advance_security_attached is None:
                advance_security_attached = "ОШИБКА: строка обеспечения не найдена"
        else:
            advance_percent = 0.0
            advance_sum = 0.0
            advance_security_required = 0
            advance_security_attached = 0

    return {
        "Ссылка": additions_url,
        "Многолетний": is_long_term,
        "Год окончания": max_end_year,
        "Сумма 2026": amount_2026,
        "Сумма 2027": amount_2027,
        "% аванса": advance_percent,
        "Сумма аванса": advance_sum,
        "Обеспечение аванса требуется": advance_security_required,
        "Обеспечение аванса прикреплено": advance_security_attached,
        "Факт сумма": fact_sum,
    }


def run_code_6_gos(file_bytes: bytes, selected_gos_sheets: List[str], token: str) -> bytes:
    wb = load_workbook(io.BytesIO(file_bytes))

    for sheet_name in selected_gos_sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        # Найти столбцы «Номер договора» и «Статус» в строке 1
        col_contract: Optional[int] = None
        col_status: Optional[int] = None
        for col_idx in range(1, (ws.max_column or 1) + 1):
            cell_val = ws.cell(row=1, column=col_idx).value
            if cell_val is None:
                continue
            cv = str(cell_val).strip()
            if cv == "Номер договора":
                col_contract = col_idx
            elif cv == "Статус":
                col_status = col_idx

        if col_contract is None or col_status is None:
            continue

        # Начало новых столбцов — после последнего используемого столбца
        start_col = (ws.max_column or 1) + 1

        # Заголовки новых столбцов в строке 1
        for j, col_name in enumerate(_GOS_NEW_COLS):
            ws.cell(row=1, column=start_col + j, value=col_name)

        # Обработка строк
        for row_idx in range(2, (ws.max_row or 1) + 1):
            status_val = ws.cell(row=row_idx, column=col_status).value
            if str(status_val or "").strip() != "Действует":
                continue
            contract_number = ws.cell(row=row_idx, column=col_contract).value
            if not contract_number:
                continue
            contract_number = str(contract_number).strip()
            try:
                row_data = _gos_build_row(contract_number, token)
                for j, col_name in enumerate(_GOS_NEW_COLS):
                    val = row_data.get(col_name)
                    c = ws.cell(row=row_idx, column=start_col + j)
                    if col_name == "Ссылка" and val:
                        c.value = "Открыть"
                        c.hyperlink = str(val)
                    else:
                        c.value = val
            except Exception as e:
                ws.cell(row=row_idx, column=start_col, value=f"ОШИБКА: {e}")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# =========================
# ?????? ? ????
# =========================
st.set_page_config(page_title="", page_icon=None, layout="wide", initial_sidebar_state="collapsed")

# Тема (современный dark UI)
BG = "#101824"          # фон
BG_2 = "#101824"        # без градиента
TEXT = "#E9EEF5"        # основной текст
MUTED = "#A7B3C4"       # подписи/вторичный текст
CARD = "#0F1722"        # карточки/поля
BORDER = "#223042"      # границы
ACCENT = "#ADD6FF"      # акцент (чекбоксы/фокус)
ACCENT_2 = "#2EE9A6"    # второй акцент (для подсветок)
BTN_BG = "#163042"      # кнопки (тёмный сине-серый)
BTN_TEXT = TEXT         # текст кнопок (светлый)
BTN_BORDER = "rgba(255,255,255,0.12)"
BTN_HOVER = "#1B3A50"

st.markdown(
    f"""
    <style>
      #MainMenu {{visibility: hidden;}}
      footer {{visibility: hidden;}}
      header {{visibility: hidden;}}

      @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
      @import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&display=swap');
      @import url('https://fonts.googleapis.com/css2?family=Playfair+Display:wght@400;600;700&display=swap');

      .stApp {{
        background: {BG};
        color: {TEXT};
      }}

      html, body, [class*="css"], .stApp {{
        font-size: 15px !important;
        font-family: "Playfair Display", Georgia, "Times New Roman", serif !important;
      }}

      /* Streamlit/BaseWeb часто задают шрифты точечно — фиксируем на всей типографике. */
      button, input, textarea, select, option,
      label,
      [data-baseweb],
      [data-baseweb] *,
      [data-testid] * {{
        font-family: "Playfair Display", Georgia, "Times New Roman", serif !important;
      }}

      .block-container {{
        max-width: 980px;
        padding-top: 1.25rem;
        padding-bottom: 1.25rem;
      }}

      .stMarkdown h4 {{
        font-size: 18px !important;
        margin: 0 0 0.55rem 0 !important;
        color: {TEXT} !important;
        font-weight: 700 !important;
        letter-spacing: 0.2px !important;
      }}

      /* Карточки/контейнеры */
      [data-testid="stFileUploader"] section,
      [data-testid="stForm"],
      .stExpander,
      [data-testid="stMetric"] {{
        border-radius: 14px !important;
      }}

      [data-testid="stFileUploader"] section {{
        padding: 14px;
        border: 1px solid rgba(255,255,255,0.10);
        background: {CARD};
        box-shadow: none;
      }}
      [data-testid="stFileUploader"] * {{
        color: {TEXT} !important;
      }}
      [data-testid="stFileUploader"] small {{
        color: {MUTED} !important;
      }}

      /* Универсальные кнопки (Streamlit / BaseWeb / kind=primary/secondary) */
      div.stButton > button,
      div.stDownloadButton > button,
      button[kind="primary"],
      button[kind="secondary"],
      div[data-baseweb="button"] > button,
      div[data-baseweb="button"] button {{
        background: {BTN_BG} !important;
        color: {BTN_TEXT} !important;
        border: 1px solid {BTN_BORDER} !important;
        border-radius: 12px !important;
        padding: 0.60rem 0.90rem !important;
        font-weight: 700 !important;
        box-shadow: none !important;
        transition: transform 120ms ease, filter 120ms ease, background-color 120ms ease !important;
      }}
      div.stButton > button:hover,
      div.stDownloadButton > button:hover,
      button[kind="primary"]:hover,
      button[kind="secondary"]:hover,
      div[data-baseweb="button"] > button:hover,
      div[data-baseweb="button"] button:hover {{
        background: {BTN_HOVER} !important;
        transform: translateY(-1px) !important;
      }}
      div.stButton > button:active,
      div.stDownloadButton > button:active,
      button[kind="primary"]:active,
      button[kind="secondary"]:active,
      div[data-baseweb="button"] > button:active,
      div[data-baseweb="button"] button:active {{
        transform: translateY(0px) !important;
        filter: brightness(0.98) !important;
      }}
      div.stButton > button:focus-visible,
      div.stDownloadButton > button:focus-visible,
      button[kind="primary"]:focus-visible,
      button[kind="secondary"]:focus-visible,
      div[data-baseweb="button"] > button:focus-visible,
      div[data-baseweb="button"] button:focus-visible {{
        outline: none !important;
        box-shadow: 0 0 0 3px rgba(173, 214, 255, 0.22) !important;
        border-color: rgba(173, 214, 255, 0.55) !important;
      }}

      /* Uploader: в разных версиях кнопка живёт в dropzone/section */
      [data-testid="stFileUploader"] button,
      [data-testid="stFileUploaderDropzone"] button {{
        background: {BTN_BG} !important;
        color: {BTN_TEXT} !important;
        border: 1px solid {BTN_BORDER} !important;
        border-radius: 12px !important;
        font-weight: 700 !important;
      }}
      [data-testid="stFileUploader"] button:hover,
      [data-testid="stFileUploaderDropzone"] button:hover {{
        background: {BTN_HOVER} !important;
      }}

      [data-testid="stFileUploaderDeleteBtn"] button {{
        background: rgba(255,255,255,0.06) !important;
        color: {TEXT} !important;
        border: 1px solid {BTN_BORDER} !important;
      }}
      [data-testid="stFileUploaderDeleteBtn"] svg {{
        fill: {TEXT} !important;
        color: {TEXT} !important;
      }}

      input, textarea, [data-baseweb="input"] input {{
        background: {CARD} !important;
        color: {TEXT} !important;
        border-color: rgba(255,255,255,0.12) !important;
        border-radius: 12px !important;
      }}
      input::placeholder, textarea::placeholder {{
        color: {MUTED} !important;
      }}
      input:focus, textarea:focus, [data-baseweb="input"] input:focus {{
        outline: none !important;
        box-shadow: 0 0 0 3px rgba(173, 214, 255, 0.24) !important;
        border-color: rgba(173, 214, 255, 0.55) !important;
      }}

      /* Ширина для основных action-кнопок */
      div.stButton > button, div.stDownloadButton > button {{
        width: 100%;
        font-size: 16px !important;
      }}

      .stMarkdown, .stMarkdown p, .stCaption, label {{
        color: {TEXT} !important;
      }}
      .stCaption {{
        color: {MUTED} !important;
      }}

      /* Чекбоксы/радио */
      [data-testid="stCheckbox"] * {{
        color: {TEXT} !important;
      }}
      input[type="checkbox"], input[type="radio"] {{
        accent-color: {ACCENT} !important;
      }}

      /* Алерты */
      [data-testid="stAlert"] {{
        border-radius: 14px !important;
        border: 1px solid rgba(255,255,255,0.10) !important;
        box-shadow: none !important;
      }}
    </style>
    """,
    unsafe_allow_html=True,
)

def classify_uploads(uploads) -> Dict[str, object]:
    analysis = []
    wh = []
    mkz = []
    osv = []
    other = []

    for u in uploads:
        name = u.name or ""
        lower = name.lower()
        if lower.startswith("_анализ"):
            analysis.append(u)
        elif lower.startswith("wh_kz"):
            wh.append(u)
        elif lower.startswith("m_kz"):
            mkz.append(u)
        else:
            ext = lower.rsplit(".", 1)[-1] if "." in lower else ""
            if ext in ("xlsx", "xlsm", "xls"):
                osv.append(u)
            else:
                other.append(u)

    return {"analysis": analysis, "wh": wh, "m": mkz, "osv": osv, "other": other}


st.markdown("#### Загрузка")
upload_types = ["xlsx", "xlsm"] + (["xls"] if os.name == "nt" else [])
uploads = st.file_uploader(
    "Загрузите файлы Excel",
    type=upload_types,
    accept_multiple_files=True,
    label_visibility="collapsed",
    key="uploads",
)

if not uploads:
    st.stop()

if os.name != "nt":
    st.caption(".xls на Streamlit Cloud (Linux) не поддерживается. Используй .xlsx/.xlsm или запускай локально на Windows для авто-конвертации.")

if any(is_xls_filename(u.name) for u in uploads):
    st.info("Обнаружены .xls. Перед обработкой они будут конвертированы в .xlsx через установленный Microsoft Excel (может занять время).")

sig = tuple(sorted((u.name, getattr(u, "size", None)) for u in uploads))
if st.session_state.get("upload_sig") != sig:
    st.session_state["upload_sig"] = sig
    for k in ["prepared_bytes", "prepared_name", "availability", "prep_report", "processed_bytes", "obsh_meta"]:
        st.session_state.pop(k, None)

cls = classify_uploads(uploads)
analysis_list: List = cls["analysis"]
wh_list: List = cls["wh"]
m_list: List = cls["m"]
osv_list: List = cls["osv"]

if cls["other"]:
    st.warning("Некоторые файлы не подходят (нужны .xlsx/.xlsm). Они будут проигнорированы.")

if len(analysis_list) > 1:
    st.error("Файл _Анализ должен быть один.")
    st.stop()

analysis_u = analysis_list[0] if analysis_list else None

if (wh_list and not m_list) or (m_list and not wh_list):
    st.error("Для сценариев с WH_KZ/M_KZ нужно загрузить оба файла: WH_KZ и M_KZ.")
    st.stop()

needs_new_name = analysis_u is None
analysis_title = ""
if needs_new_name:
    st.markdown("#### Имя _Анализ")
    analysis_title = st.text_input("Название", value="", placeholder="Например: Алтын Мунай ЛТД 170740018427")

out_name_preview = analysis_u.name if analysis_u else safe_filename(f"_Анализ {analysis_title}".strip() + ".xlsx")
st.caption("Выходной файл: " + out_name_preview)

prep_disabled = needs_new_name and not analysis_title.strip()

# Запрос префиксов для дубликатов ОСВ (на каждый загруженный лист ОСВ)
osv_prefix_by_sheet: Dict[Tuple[str, str, int], str] = {}
wh_prefix_by_file: Dict[str, str] = {}
m_prefix_by_file: Dict[str, str] = {}

xls_cache = st.session_state.setdefault("xls_cache", {})

analysis_wb_tmp = None
if analysis_u is not None:
    try:
        an_name, an_bytes = ensure_openpyxl_bytes(analysis_u.name, analysis_u.getvalue(), cache=xls_cache)
        analysis_wb_tmp, _ = load_wb_from_bytes(an_bytes, an_name)
    except Exception:
        analysis_wb_tmp = None

existing_suffixes: Set[str] = set()
if analysis_wb_tmp is not None:
    for sh in analysis_wb_tmp.sheetnames:
        existing_suffixes.add(split_prefix_suffix4(sh)[1])

    existing_osv_info = list_existing_osv_sheets_with_a1(analysis_wb_tmp)
    if existing_osv_info:
        st.markdown("#### _Анализ: найденные ОСВ")
        for sh, suf, a1 in existing_osv_info[:12]:
            tail = f" | {_short(a1)}" if a1 else ""
            st.caption(f"{sh} → {suf}{tail}")
        if len(existing_osv_info) > 12:
            st.caption(f"... и еще {len(existing_osv_info) - 12}")

# Находим все загруженные ОСВ-листы, у которых определился 4-значный номер счета.
osv_items: List[Tuple[str, str, int, str, str]] = []
for u in osv_list:
    try:
        u_name, u_bytes = ensure_openpyxl_bytes(u.name, u.getvalue(), cache=xls_cache)
        wb_tmp, _ = load_wb_from_bytes(u_bytes, u_name)
        for idx, ws in enumerate(wb_tmp.worksheets):
            acc = get_account_number(ws)
            if acc and acc.isdigit() and len(acc) == 4:
                company = _short(extract_company_label_from_a1(ws))
                osv_items.append((u.name, ws.title, idx, acc, company))
    except Exception:
        pass

acc_to_items: Dict[str, List[Tuple[str, str, int]]] = defaultdict(list)
for fname, sheet_title, idx, acc, _company in osv_items:
    acc_to_items[acc].append((fname, sheet_title, idx))

need_prefix_items: List[Tuple[str, str, int, str]] = []
for acc, items in acc_to_items.items():
    if len(items) > 1 or (analysis_wb_tmp is not None and acc in existing_suffixes):
        for fname, sheet_title, idx in items:
            need_prefix_items.append((fname, sheet_title, idx, acc))

if need_prefix_items:
    st.markdown("#### Префиксы ОСВ (повторы/конфликт)")
    st.caption("Обнаружены повторяющиеся ОСВ по счетам или такие счета уже есть в _Анализ. Для КАЖДОГО ОСВ укажи префикс или отметь «без префикса».")

    # Валидация: в рамках одного счета нельзя использовать один и тот же префикс дважды;
    # "без префикса" тоже должен быть максимум один раз на счет.
    chosen_by_acc: Dict[str, List[str]] = defaultdict(list)
    for fname, sheet_title, idx, acc in need_prefix_items:
        company = ""
        for f2, s2, i2, a2, comp2 in osv_items:
            if f2 == fname and s2 == sheet_title and i2 == idx and a2 == acc:
                company = comp2
                break
        if company:
            st.caption(f"{fname} / {sheet_title} → {acc} | {company}")
        else:
            st.caption(f"{fname} / {sheet_title} → {acc}")
        no_pref = st.checkbox("без префикса", value=False, key=f"osv_nopref::{fname}::{idx}")
        pref = st.text_input("префикс", value="", key=f"osv_pref::{fname}::{idx}", disabled=no_pref)
        if no_pref:
            osv_prefix_by_sheet[(fname, sheet_title, idx)] = ""
            chosen_by_acc[acc].append("")
        else:
            if not pref.strip():
                prep_disabled = True
                st.caption("Нужен префикс или отметь «без префикса».")
            else:
                p = pref.strip()
                osv_prefix_by_sheet[(fname, sheet_title, idx)] = p
                chosen_by_acc[acc].append(p)

    for acc, prefs in chosen_by_acc.items():
        if len(prefs) != len(set(prefs)):
            prep_disabled = True
            st.caption(f"Для счета {acc} префиксы должны быть уникальными (включая «без префикса»).")

# Префиксы для WH_KZ / M_KZ, если загружено несколько файлов
if len(wh_list) > 1:
    st.markdown("#### Префиксы WH_KZ")
    st.caption("Загружено несколько WH_KZ. Для каждого укажи префикс или «без префикса» (префиксы должны быть уникальными).")
    seen: Set[str] = set()
    for u in wh_list:
        no_pref = st.checkbox(f"{u.name}: без префикса", value=False, key=f"wh_nopref::{u.name}")
        pref = st.text_input(f"{u.name}: префикс", value="", key=f"wh_pref::{u.name}", disabled=no_pref)
        p = "" if no_pref else pref.strip()
        if not no_pref and not p:
            prep_disabled = True
            st.caption(f"{u.name}: нужен префикс или «без префикса».")
        if p in seen:
            prep_disabled = True
            st.caption("Префиксы WH_KZ должны быть уникальными.")
        seen.add(p)
        wh_prefix_by_file[u.name] = p

if len(m_list) > 1:
    st.markdown("#### Префиксы M_KZ")
    st.caption("Загружено несколько M_KZ. Для каждого укажи префикс или «без префикса» (префиксы должны быть уникальными).")
    seen: Set[str] = set()
    for u in m_list:
        no_pref = st.checkbox(f"{u.name}: без префикса", value=False, key=f"m_nopref::{u.name}")
        pref = st.text_input(f"{u.name}: префикс", value="", key=f"m_pref::{u.name}", disabled=no_pref)
        p = "" if no_pref else pref.strip()
        if not no_pref and not p:
            prep_disabled = True
            st.caption(f"{u.name}: нужен префикс или «без префикса».")
        if p in seen:
            prep_disabled = True
            st.caption("Префиксы M_KZ должны быть уникальными.")
        seen.add(p)
        m_prefix_by_file[u.name] = p

prep_btn = st.button("Собрать _Анализ", disabled=prep_disabled)

if prep_btn:
    status = st.empty()
    prog = st.progress(0)

    def _ui_progress(frac: float, msg: str) -> None:
        status.info(msg)
        try:
            prog.progress(int(round(frac * 100)))
        except Exception:
            try:
                prog.progress(frac)
            except Exception:
                pass

    status.info("Сборка…")
    try:
        analysis_file = None
        if analysis_u:
            an_name, an_bytes = ensure_openpyxl_bytes(analysis_u.name, analysis_u.getvalue(), cache=xls_cache)
            analysis_file = (an_name, an_bytes)

        wh_files = []
        for u in wh_list:
            n, b = ensure_openpyxl_bytes(u.name, u.getvalue(), cache=xls_cache)
            wh_files.append((n, b, wh_prefix_by_file.get(u.name, "")))

        m_files = []
        for u in m_list:
            n, b = ensure_openpyxl_bytes(u.name, u.getvalue(), cache=xls_cache)
            m_files.append((n, b, m_prefix_by_file.get(u.name, "")))

        osv_files = []
        for u in osv_list:
            n, b = ensure_openpyxl_bytes(u.name, u.getvalue(), cache=xls_cache)
            osv_files.append((u.name, b))

        out_bytes, out_name, availability, prep_report = build_analysis_workbook(
            analysis_file=analysis_file,
            wh_files=wh_files,
            m_files=m_files,
            osv_files=osv_files,
            analysis_name_for_new=analysis_title,
            osv_prefix_by_sheet=osv_prefix_by_sheet,
            progress_cb=_ui_progress,
        )

        st.session_state["prepared_bytes"] = out_bytes
        st.session_state["prepared_name"] = out_name
        st.session_state["availability"] = availability
        st.session_state["prep_report"] = prep_report
        st.session_state.pop("processed_bytes", None)
        _ui_progress(1.0, "Сборка: готово.")
        status.success("Готово.")
    except Exception as e:
        status.error(f"Ошибка сборки: {e}")

if "prepared_bytes" in st.session_state:
    prep_report = st.session_state.get("prep_report") or {}
    if prep_report.get("warnings"):
        for w in prep_report["warnings"]:
            st.warning(w)

    st.write("")
    st.markdown("#### Обработка")

    availability = st.session_state.get("availability") or {}
    inv_accounts_found = availability.get("inventory_accounts_found") or []
    saldo_ok = bool(availability.get("saldo_ok"))
    contracts_ok = bool(availability.get("contracts_ok"))
    profit_ok = bool(availability.get("profit_ok"))
    obsh_sheets = availability.get("obsh_sheets") or []
    insights_ok = bool(availability.get("insights_ok"))
    insights_existing = availability.get("insights_existing_titles") or []
    insights_missing = availability.get("insights_missing_titles") or []
    insights_legacy = availability.get("insights_legacy_titles") or []
    gos_ok = bool(availability.get("gos_ok"))
    gos_sheets = availability.get("gos_sheets") or []

    opt_profit = st.checkbox("Чистая прибыль", value=False, disabled=(not profit_ok))
    if not profit_ok:
        st.caption("Чистая прибыль недоступна: не найдены листы, содержащие «общ» в названии (например: «общ26», «Gобщ26»).")

    selected_obsh: List[str] = []
    if opt_profit and obsh_sheets:
        # Показываем выбор листов "общ*" + подпись компании из A1
        if "obsh_meta" not in st.session_state:
            meta = {}
            try:
                wb_tmp = load_workbook(io.BytesIO(st.session_state["prepared_bytes"]), read_only=True, data_only=True)
                for sh in obsh_sheets:
                    try:
                        a1 = wb_tmp[sh].cell(row=1, column=1).value
                        meta[sh] = _short(a1) if a1 is not None else ""
                    except Exception:
                        meta[sh] = ""
            except Exception:
                meta = {sh: "" for sh in obsh_sheets}
            st.session_state["obsh_meta"] = meta
        meta = st.session_state.get("obsh_meta") or {}

        st.caption("Выбери листы «общ*», по которым строить расчёт:")
        for sh in obsh_sheets:
            pad, col1, col2 = st.columns([1, 2, 4])
            with pad:
                st.write("")
            with col1:
                on = st.checkbox(sh, value=False, key=f"obsh_pick::{sh}")
            with col2:
                st.caption(meta.get(sh, ""))
            if on:
                selected_obsh.append(sh)

    opt_contracts = st.checkbox("Контракты", value=False, disabled=(not contracts_ok))
    if not contracts_ok:
        st.caption("Контракты недоступны: не найдены пары листов *Wd/*Md.")

    opt_saldo = st.checkbox("Сальдо", value=False, disabled=(not saldo_ok))
    if not saldo_ok:
        st.caption("Сальдо недоступно: не найдены листы, заканчивающиеся на 1210/1710/3310/3510.")

    inv_available_any = bool(inv_accounts_found)
    opt_inventory = st.checkbox("Запасы", value=False, disabled=(not inv_available_any))
    if not inv_available_any:
        st.caption("Запасы недоступны: не найдены ОСВ-листы по счетам 13**.")

    inventory_accounts: List[str] = inv_accounts_found if opt_inventory else []

    opt_insights = st.checkbox("Инсайты", value=False, disabled=(not insights_ok))
    if insights_existing:
        st.warning("Инс уже есть в файле: " + ", ".join(insights_existing) + ". Повторно не генерирую.")
    if insights_legacy:
        st.caption("Найдены старые листы «инсайты» — при запуске будут переименованы в «инс» и перенесены в конец файла.")
    if not insights_ok:
        if insights_existing:
            st.caption("Инс недоступен: для всех найденных префиксов лист «инс» уже существует.")
        else:
            st.caption("Инсайты недоступны: нужны листы W, M, Mt, Wt (лист «общ» — только для блока 6).")
    elif insights_missing and insights_existing:
        st.caption("Будут добавлены только отсутствующие: " + ", ".join(insights_missing) + ".")

    opt_gos = st.checkbox("Госы", value=False, disabled=(not gos_ok))
    if not gos_ok:
        st.caption("Госы недоступны: не найдены листы, содержащие «гос» в названии.")

    selected_gos_sheets: List[str] = []
    gos_token = ""
    if opt_gos and gos_sheets:
        st.caption("Выбери листы «гос*» для обработки:")
        for sh in gos_sheets:
            pad, col1 = st.columns([1, 6])
            with pad:
                st.write("")
            with col1:
                on = st.checkbox(sh, value=False, key=f"gos_pick::{sh}")
            if on:
                selected_gos_sheets.append(sh)
        gos_token = st.text_input(
            "Токен API goszakup.gov.kz",
            value="",
            placeholder="Вставьте токен…",
            type="password",
            key="gos_token_input",
        )

    selected_modes: List[str] = []
    if opt_saldo:
        selected_modes.append("Сальдо")
    if opt_contracts:
        selected_modes.append("Контракты")

    st.write("")
    st.markdown("#### Запуск")

    has_any_mode = bool(selected_modes) or opt_inventory or opt_profit or opt_insights or opt_gos
    inventory_ok = (not opt_inventory) or bool(inventory_accounts)
    profit_sel_ok = (not opt_profit) or bool(selected_obsh)
    gos_sel_ok = (not opt_gos) or (bool(selected_gos_sheets) and bool(gos_token))
    run_btn = st.button("Обработать", disabled=((not has_any_mode) or (not inventory_ok) or (not profit_sel_ok) or (not gos_sel_ok)))

    status_box = st.empty()
    progress = st.progress(0)

    if run_btn:
        try:
            status_box.info("Подготовка…")
            progress.progress(10)
            time.sleep(0.05)

            out_bytes = st.session_state["prepared_bytes"]

            if opt_profit:
                status_box.info("Обработка: Чистая прибыль…")
                progress.progress(30)
                out_bytes = run_code_4_net_profit(out_bytes, selected_obsh)
                progress.progress(40)

            if "Контракты" in selected_modes:
                status_box.info("Обработка: Контракты…")
                progress.progress(45)
                out_bytes = run_code_2(out_bytes)
                progress.progress(60)

            if "Сальдо" in selected_modes:
                status_box.info("Обработка: Сальдо…")
                progress.progress(65)
                out_bytes = run_code_1(out_bytes)
                progress.progress(78)

            if opt_inventory:
                status_box.info("Обработка: Запасы…")
                progress.progress(82)
                out_bytes, inv_report = run_code_3_inventory(out_bytes, inventory_accounts)
                progress.progress(94)
                if inv_report.get("missing_sheets") or inv_report.get("missing_markers"):
                    parts = []
                    if inv_report.get("missing_sheets"):
                        parts.append("не найдены листы: " + ", ".join(inv_report["missing_sheets"]))
                    if inv_report.get("missing_markers"):
                        parts.append("в листах не найден счет в колонке A")
                    st.warning("Запасы: " + "; ".join(parts))

            if opt_insights:
                status_box.info("Обработка: Инсайты…")
                progress.progress(97)
                out_bytes = run_code_5_insights(out_bytes)
                progress.progress(99)

            if opt_gos and selected_gos_sheets:
                status_box.info("Обработка: Госы (запросы к API, может занять время)…")
                progress.progress(99)
                out_bytes = run_code_6_gos(out_bytes, selected_gos_sheets, gos_token)

            st.session_state["processed_bytes"] = out_bytes
            st.session_state["processed_name"] = st.session_state.get("prepared_name") or "output.xlsx"
            status_box.success("Готово.")
            progress.progress(100)
        except Exception as e:
            progress.progress(0)
            status_box.error(f"Ошибка: {e}")

    st.write("")
    st.markdown("#### Скачать")
    download_bytes = st.session_state.get("processed_bytes") or st.session_state.get("prepared_bytes")
    download_name = st.session_state.get("processed_name") or st.session_state.get("prepared_name") or "output.xlsx"
    # Если внутри есть проект VBA, Excel ожидает .xlsm; иначе сохраняем как .xlsx.
    want_xlsm = bool(download_bytes and has_vba_project(download_bytes))
    if want_xlsm:
        if not download_name.lower().endswith(".xlsm"):
            download_name = re.sub(r"\.(xlsx|xlsm)$", ".xlsm", download_name, flags=re.IGNORECASE)
            if not download_name.lower().endswith(".xlsm"):
                download_name += ".xlsm"
        download_mime = "application/vnd.ms-excel.sheet.macroEnabled.12"
    else:
        if not download_name.lower().endswith(".xlsx"):
            download_name = re.sub(r"\.(xlsx|xlsm)$", ".xlsx", download_name, flags=re.IGNORECASE)
            if not download_name.lower().endswith(".xlsx"):
                download_name += ".xlsx"
        download_mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    if not has_any_mode:
        confirm_merge = st.checkbox("Только объединить (без обработок)", value=False)
        st.download_button(
            label="Скачать",
            data=download_bytes,
            file_name=download_name,
            mime=download_mime,
            use_container_width=True,
            disabled=(not confirm_merge),
        )
    else:
        st.download_button(
            label="Скачать",
            data=download_bytes,
            file_name=download_name,
            mime=download_mime,
            use_container_width=True,
        )
